"""
Escritura incremental del checklist de validación documental en formato Excel.

Columnas:
  Identificación : ID_unico, modalidad, unidad_doc, integrante
  Carpeta 00     : CEDULA, COMERCIO, RUT, TENENCIA
  Carpeta 01     : 01_DOCUMENTOS, 01_FOTOS_VIDEOS, 01_REVISION_IA
  General        : observaciones

Color rojo  : documentos obligatorios ausentes (00) o faltantes en 01.
Color amarillo: alertas de la revisión IA en documentos de 01.
"""

import os
from typing import Dict, Set

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.core.reglas import Estado

# ── Paleta de colores (ARGB) ──────────────────────────────────────────────────
_LILA     = "FFCC8FCC"
_AZUL_OSC = "FF2E4A7A"   # encabezado de grupo (carpeta)
_AZUL_MED = "FF4472C4"   # encabezado Identificación
_VERDE    = "FF375623"   # encabezado 01 Visita
_NARANJA  = "FF833C00"   # encabezado 00 Documentación
_TEAL     = "FF1F5C6B"   # encabezado 03 Capacitación
_GRIS     = "FF595959"   # encabezado General
_ROJO     = "FFFF6B6B"
_AMARILLO = "FFFFF59D"
_NARANJA_AT = "FFFFB74D"   # naranja de "atención" (distinto de rojo y amarillo)
_BLANC    = "FFFFFFFF"
_NEGRO    = "FF000000"

FILL_LILA       = PatternFill(start_color=_LILA,       end_color=_LILA,       fill_type="solid")
FILL_ROJO       = PatternFill(start_color=_ROJO,       end_color=_ROJO,       fill_type="solid")
FILL_AMARILLO   = PatternFill(start_color=_AMARILLO,   end_color=_AMARILLO,   fill_type="solid")
FILL_ATENCION   = PatternFill(start_color=_NARANJA_AT, end_color=_NARANJA_AT, fill_type="solid")

FONT_HEADER       = Font(name="Arial", size=9,  bold=True,  color=_BLANC)
FONT_GRUPO        = Font(name="Arial", size=10, bold=True,  color=_BLANC)
FONT_NORMAL       = Font(name="Arial", size=9)
FONT_FALTA        = Font(name="Arial", size=9,  bold=True,  color=_BLANC)
FONT_IA_WARN      = Font(name="Arial", size=9,  bold=False, color=_NEGRO)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ── Columnas ──────────────────────────────────────────────────────────────────

_MORADO   = "FF5C3273"   # encabezado 02 Visita 2
_CAFE     = "FF7B3F00"   # encabezado 04 Capitalización

COLUMNAS = [
    # Identificación
    "ID_unico", "modalidad", "unidad_doc",
    # Carpeta 00_DOCUMENTACION
    "CEDULA", "COMERCIO", "RUT", "TENENCIA",
    # Carpeta 01_VISITA_1_CARACTERIZACION
    "01_DOCUMENTOS",
    "01_FOTOS_VIDEOS",
    "ACTA_COMPROMISO",
    "ACTA_VISITA",
    "GESTOR",
    "TRATAMIENTO_DATOS",
    # Carpeta 02_VISITA_2_DIAGNOSTICO
    "02_DOCUMENTOS",
    "02_ACTA_VISITA",
    "02_DIAGNOSTICO",
    "02_PLAN_NEGOCIO",
    # Carpeta 03 (Capacitación)
    "03_ENCUESTAS",
    "03_GRUPAL",
    "03_INDIVIDUAL",
    "03_MODULOS",
    "03_ASISTENCIA",
    # Carpeta 04 (Capitalización)
    "04_PLAN_INVERSION_ENCONTRADO",
    "04_PDF_APROBACION_ENCONTRADO",
    "04_PDF_APROBACION_FIRMADO",
    "04_CONSISTENCIA_COTIZACION_GANADORA",
    "04_COTIZACIONES (carpeta 02)",
    "04_VALIDACION_PRECIO_MERCADO",
    # General
    "observaciones",
]

_ANCHOS_MIN = {
    "ID_unico":          18, "modalidad":       10, "unidad_doc":      22,
    "CEDULA":            18, "COMERCIO":        18, "RUT":             14,
    "TENENCIA":          18,
    "01_DOCUMENTOS":     38, "01_FOTOS_VIDEOS": 22,
    "ACTA_COMPROMISO":   35, "ACTA_VISITA":     35,
    "GESTOR":            40, "TRATAMIENTO_DATOS": 35,
    "02_DOCUMENTOS":     35, "02_ACTA_VISITA":  35,
    "02_DIAGNOSTICO":    35, "02_PLAN_NEGOCIO": 35,
    "03_ENCUESTAS":      25, "03_GRUPAL":       18,
    "03_INDIVIDUAL":     18, "03_MODULOS":      50,
    "03_ASISTENCIA":     55,
    "04_PLAN_INVERSION_ENCONTRADO":           45,
    "04_PDF_APROBACION_ENCONTRADO":            30,
    "04_PDF_APROBACION_FIRMADO":         30,
    "04_CONSISTENCIA_COTIZACION_GANADORA":     55,
    "04_COTIZACIONES (carpeta 02)":    55,
    "04_VALIDACION_PRECIO_MERCADO":            45,
    "observaciones":     50,
}

_COLS_00_IDX = {
    doc: COLUMNAS.index(doc) + 1
    for doc in ("CEDULA", "COMERCIO", "RUT", "TENENCIA")
}

_COL_DOCS_IDX   = COLUMNAS.index("01_DOCUMENTOS")   + 1
_COL_FOTOS_IDX  = COLUMNAS.index("01_FOTOS_VIDEOS") + 1
_COL_GESTOR_IDX = COLUMNAS.index("GESTOR")          + 1
_COLS_DOC_REV_IDX = {
    col: COLUMNAS.index(col) + 1
    for col in ("ACTA_COMPROMISO", "ACTA_VISITA", "GESTOR", "TRATAMIENTO_DATOS")
}
_COL_DOCS2_IDX = COLUMNAS.index("02_DOCUMENTOS") + 1
_COLS_DOC_REV2_IDX = {
    col: COLUMNAS.index(col) + 1
    for col in ("02_ACTA_VISITA", "02_DIAGNOSTICO", "02_PLAN_NEGOCIO")
}
_COLS_03_IDX = {
    col: COLUMNAS.index(col) + 1
    for col in ("03_ENCUESTAS", "03_GRUPAL", "03_INDIVIDUAL", "03_MODULOS", "03_ASISTENCIA")
}
_COLS_04_IDX = {col: COLUMNAS.index(col) + 1 for col in ("04_PLAN_INVERSION_ENCONTRADO", "04_PDF_APROBACION_ENCONTRADO", "04_PDF_APROBACION_FIRMADO", "04_CONSISTENCIA_COTIZACION_GANADORA", "04_COTIZACIONES (carpeta 02)", "04_VALIDACION_PRECIO_MERCADO")}

# Orden y color de cada grupo-carpeta para la fila de encabezado superior
_GRUPOS_ORDEN = [
    ("Identificación",   ["ID_unico", "modalidad", "unidad_doc"],                 _AZUL_MED),
    ("00 Documentación", ["CEDULA", "COMERCIO", "RUT", "TENENCIA"],               _NARANJA),
    ("01 Visita",        ["01_DOCUMENTOS", "01_FOTOS_VIDEOS",
                          "ACTA_COMPROMISO", "ACTA_VISITA", "GESTOR",
                          "TRATAMIENTO_DATOS"],                                    _VERDE),
    ("02 Visita 2",      ["02_DOCUMENTOS", "02_ACTA_VISITA",
                          "02_DIAGNOSTICO", "02_PLAN_NEGOCIO"],                   _MORADO),
    ("03 Capacitación",  ["03_ENCUESTAS", "03_GRUPAL", "03_INDIVIDUAL",
                          "03_MODULOS", "03_ASISTENCIA"],                          _TEAL),
    ("04 Capitalización", ["04_PLAN_INVERSION_ENCONTRADO", "04_PDF_APROBACION_ENCONTRADO", "04_PDF_APROBACION_FIRMADO", "04_CONSISTENCIA_COTIZACION_GANADORA", "04_COTIZACIONES (carpeta 02)", "04_VALIDACION_PRECIO_MERCADO"], _CAFE),
    ("General",          ["observaciones"],                                        _GRIS),
]


class ChecklistWriter:
    """Mantiene y escribe el archivo Excel del checklist de validación."""

    def __init__(self, ruta: str, lote_guardado: int = 50):
        self.ruta          = ruta
        self.lote_guardado = lote_guardado
        self._sin_guardar  = 0
        self._max_ancho    = [_ANCHOS_MIN[c] for c in COLUMNAS]

        if os.path.exists(ruta):
            self._wb = openpyxl.load_workbook(ruta)
            self._ws = self._wb.active
            self._ids_procesados: Set[str] = self._leer_ids_existentes()
        else:
            self._wb = openpyxl.Workbook()
            self._ws = self._wb.active
            self._ws.title = "Checklist"
            self._ids_procesados = set()
            self._escribir_encabezado()

    def ids_procesados(self) -> Set[str]:
        return self._ids_procesados

    def agregar_fila(self, resultado: dict) -> None:
        docs_estado: Dict[str, Estado] = resultado.get("docs_estado_raw", {})

        valores = [
            resultado.get("ID_unico",      ""),
            resultado.get("modalidad",     ""),
            resultado.get("unidad_doc",    ""),
            # 00
            resultado.get("docs", {}).get("CEDULA",   ""),
            resultado.get("docs", {}).get("COMERCIO", ""),
            resultado.get("docs", {}).get("RUT",      ""),
            resultado.get("docs", {}).get("TENENCIA", ""),
            # 01
            resultado.get("01_documentos",      "N/A"),
            resultado.get("01_fotos_videos",    "N/A"),
            resultado.get("01_acta_compromiso", "N/A"),
            resultado.get("01_acta_visita",     "N/A"),
            resultado.get("01_gestor",          "N/A"),
            resultado.get("01_tratamiento",     "N/A"),
            # 02
            resultado.get("02_documentos",      "N/A"),
            resultado.get("02_acta_visita",     "N/A"),
            resultado.get("02_diagnostico",     "N/A"),
            resultado.get("02_plan_negocio",    "N/A"),
            # 03
            resultado.get("03_encuestas",       "N/A"),
            resultado.get("03_grupal",          "N/A"),
            resultado.get("03_individual",      "N/A"),
            resultado.get("03_modulos",         "N/A"),
            resultado.get("03_asistencia",      "N/A"),
            # 04
            resultado.get("04_xlsx",            "N/A"),
            resultado.get("04_pdf",             "N/A"),
            resultado.get("04_firmas",          "N/A"),
            resultado.get("04_cotizacion",      "N/A"),
            resultado.get("04_cotizacion_carpeta", "N/A"),
            resultado.get("04_web",             "N/A"),
            # General
            resultado.get("observaciones", ""),
        ]

        self._ws.append(valores)
        fila_num = self._ws.max_row

        for col_idx in range(1, len(COLUMNAS) + 1):
            celda           = self._ws.cell(row=fila_num, column=col_idx)
            celda.font      = FONT_NORMAL
            celda.alignment = ALIGN_CENTER

        # Rojo en docs 00 ausentes; amarillo si cédula existe pero no coincide
        for doc, col_idx in _COLS_00_IDX.items():
            if docs_estado.get(doc) == Estado.FALTA:
                celda      = self._ws.cell(row=fila_num, column=col_idx)
                celda.fill = FILL_ROJO
                celda.font = FONT_FALTA

        # Rojo en 01_DOCUMENTOS si hay alguno faltante
        val_docs = str(valores[COLUMNAS.index("01_DOCUMENTOS")])
        if "FALTA" in val_docs.upper():
            celda      = self._ws.cell(row=fila_num, column=_COL_DOCS_IDX)
            celda.fill = FILL_ROJO
            celda.font = FONT_FALTA

         # Rojo en 02_DOCUMENTOS si hay alguno faltante
        val_docs2 = str(valores[COLUMNAS.index("02_DOCUMENTOS")])
        if "FALTA" in val_docs2.upper():
            celda = self._ws.cell(
                row=fila_num,
                column=COLUMNAS.index("02_DOCUMENTOS") + 1,
            )
            celda.fill = FILL_ROJO
            celda.font = FONT_FALTA

        # Rojo en 01_FOTOS_VIDEOS si insuficientes
        val_fotos = str(valores[COLUMNAS.index("01_FOTOS_VIDEOS")])
        if "FALTA" in val_fotos.upper():
            celda      = self._ws.cell(row=fila_num, column=_COL_FOTOS_IDX)
            celda.fill = FILL_ROJO
            celda.font = FONT_FALTA

        # Rojo en GESTOR si no está registrado en BD
        val_gestor = str(valores[COLUMNAS.index("GESTOR")])
        if "NO REGISTRADO" in val_gestor.upper():
            celda      = self._ws.cell(row=fila_num, column=_COL_GESTOR_IDX)
            celda.fill = FILL_ROJO
            celda.font = FONT_FALTA

        # Amarillo en cada columna de revisión IA de 01 si tiene alerta
        alertas_por_col = resultado.get("01_alertas_por_doc", {})
        for col_nombre, col_idx in _COLS_DOC_REV_IDX.items():
            if alertas_por_col.get(col_nombre, False):
                celda      = self._ws.cell(row=fila_num, column=col_idx)
                celda.fill = FILL_AMARILLO
                celda.font = FONT_IA_WARN

        # Rojo/amarillo en columnas 03
        alertas_03 = resultado.get("03_alertas", {})
        for col_nombre, col_idx in _COLS_03_IDX.items():
            if alertas_03.get(col_nombre, False):
                val = str(valores[col_idx - 1])
                if "FALTA" in val.upper():
                    celda      = self._ws.cell(row=fila_num, column=col_idx)
                    celda.fill = FILL_ROJO
                    celda.font = FONT_FALTA
                else:
                    celda      = self._ws.cell(row=fila_num, column=col_idx)
                    celda.fill = FILL_AMARILLO
                    celda.font = FONT_IA_WARN

        # Amarillo en columnas de revisión de la carpeta 02
        alertas_por_col_02 = resultado.get("02_alertas_por_doc", {})

        for col_nombre, col_idx in _COLS_DOC_REV2_IDX.items():
            if alertas_por_col_02.get(col_nombre, False):
                celda = self._ws.cell(row=fila_num, column=col_idx)
                celda.fill = FILL_AMARILLO
                celda.font = FONT_IA_WARN

        # Columnas 04: rojo si FALTA/INCOMPLETA; amarillo si requiere completación manual
        # La columna de cotizaciones de carpeta 02 se colorea aparte (amarillo).
        _COL_CARPETA02 = "04_COTIZACIONES (carpeta 02)"
        for col_nombre, col_idx in _COLS_04_IDX.items():
            if col_nombre == _COL_CARPETA02:
                continue
            val = str(valores[COLUMNAS.index(col_nombre)])
            val_up = val.upper()
            if "COMPLETACIÓN MANUAL" in val_up or "COMPLETACION MANUAL" in val_up:
                # Caso 04_WEB sin las 3 cotizaciones: amarillo (revisión/completación manual)
                celda      = self._ws.cell(row=fila_num, column=col_idx)
                celda.fill = FILL_AMARILLO
                celda.font = FONT_IA_WARN
            elif "FALTA" in val_up or "INCOMPLETA" in val_up or "DETENER" in val_up or "NO COINCIDE" in val_up:
                celda      = self._ws.cell(row=fila_num, column=col_idx)
                celda.fill = FILL_ROJO
                celda.font = FONT_FALTA

        # Columna 04_COTIZACIONES (carpeta 02): amarillo si hay algo que revisar
        val_c02 = str(valores[COLUMNAS.index(_COL_CARPETA02)])
        if val_c02 and not val_c02.startswith("OK") and not val_c02.startswith("N/A") and val_c02 != "—":
            celda      = self._ws.cell(row=fila_num, column=_COLS_04_IDX[_COL_CARPETA02])
            celda.fill = FILL_AMARILLO
            celda.font = FONT_IA_WARN

        # Naranja de atención en "observaciones" si hay alerta del Excel de capturas
        # (desviación de precio "04_VALIDACION_PRECIO_MERCADO [...] ALERTA"). Color distinto de rojo/amarillo.
        obs_val = str(valores[COLUMNAS.index("observaciones")])
        if "04_VALIDACION_PRECIO_MERCADO" in obs_val and "ALERTA" in obs_val.upper():
            celda      = self._ws.cell(row=fila_num, column=COLUMNAS.index("observaciones") + 1)
            celda.fill = FILL_ATENCION
            celda.font = FONT_IA_WARN

        # Auto-fit
        for i, valor in enumerate(valores):
            longitud = len(str(valor)) if valor else 0
            if longitud + 3 > self._max_ancho[i]:
                self._max_ancho[i] = longitud + 3

        self._ids_procesados.add(str(resultado.get("ID_unico", "")))
        self._sin_guardar += 1

        if self._sin_guardar >= self.lote_guardado:
            self.guardar()

    def guardar(self) -> None:
        self._ajustar_anchos()
        self._wb.save(self.ruta)
        self._sin_guardar = 0

    def _escribir_encabezado(self) -> None:
        from openpyxl.styles import Border, Side
        borde_blanco = Border(
            left=Side(style="thin", color=_BLANC),
            right=Side(style="thin", color=_BLANC),
        )

        # ── Fila 1: encabezados de grupo (carpeta) ────────────────────────────
        col_cursor = 1
        for nombre_grupo, cols_grupo, color_hex in _GRUPOS_ORDEN:
            ancho  = len(cols_grupo)
            fill   = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
            inicio = col_cursor
            fin    = col_cursor + ancho - 1

            # Celda principal con el texto
            celda           = self._ws.cell(row=1, column=inicio)
            celda.value     = nombre_grupo
            celda.fill      = fill
            celda.font      = FONT_GRUPO
            celda.alignment = ALIGN_CENTER
            celda.border    = borde_blanco

            # Rellenar celdas intermedias antes de combinar
            for c in range(inicio + 1, fin + 1):
                cx = self._ws.cell(row=1, column=c)
                cx.fill   = fill
                cx.border = borde_blanco

            if ancho > 1:
                self._ws.merge_cells(
                    start_row=1, start_column=inicio,
                    end_row=1,   end_column=fin,
                )
            col_cursor += ancho

        self._ws.row_dimensions[1].height = 22

        # ── Fila 2: nombres de columna ────────────────────────────────────────
        self._ws.append(COLUMNAS)
        for col_idx in range(1, len(COLUMNAS) + 1):
            celda           = self._ws.cell(row=2, column=col_idx)
            celda.fill      = FILL_LILA
            celda.font      = FONT_HEADER
            celda.alignment = ALIGN_CENTER
        self._ws.row_dimensions[2].height = 28
        self._ws.freeze_panes = "A3"

    def _leer_ids_existentes(self) -> Set[str]:
        ids: Set[str] = set()
        for fila in self._ws.iter_rows(min_row=3, values_only=True):
            if fila[0]:
                ids.add(str(fila[0]))
        return ids

    def _ajustar_anchos(self) -> None:
        for col_idx, ancho in enumerate(self._max_ancho, start=1):
            letra = get_column_letter(col_idx)
            self._ws.column_dimensions[letra].width = min(ancho, 65)
