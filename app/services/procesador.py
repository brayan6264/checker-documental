"""
Motor principal de validación documental.

Flujo por fila:
  1. Leer matriz Excel (hoja activa).
  2. Descargar y validar 00_DOCUMENTACION (CEDULA, COMERCIO, RUT, TENENCIA).
  3. Descargar y validar 01_VISITA_1_CARACTERIZACION:
       - Existencia de ACTA_COMPROMISO, ACTA_VISITA_1, TRATAMIENTO_DATOS.
       - Conteo mínimo de 4 fotos/videos.
       - Análisis IA de cada documento (si IA_HABILITADO=true).
  4. Registrar resultado en ChecklistWriter (incremental + reanudable).
  5. Borrar carpetas temporales.
  6. Nunca abortar por error individual: los errores van a "observaciones".
"""

import logging
import re
import shutil
import time
from pathlib import Path
from typing import Callable, Dict, Optional

import openpyxl
import pandas as pd
from app.config import CHECKLIST_LOTE_GUARDADO, DOWNLOADS_DIR, IA_HABILITADO
from app.core.normalizacion import normalizar
from app.core.reglas import (
    DOCS_VISITA_ORDEN,
    DOCS_VISITA2_ORDEN,
    DOCUMENTOS_ORDEN,
    EXTENSIONES_MEDIA,
    MIN_ARCHIVOS_MEDIA,
    MIN_ASISTENTES_03,
    PALABRAS_CLAVE,
    PALABRAS_CLAVE_VISITA,
    REGLAS,
    UNIDAD_DOC_TIPO,
    Estado,
    evaluar_documentos,
    estados_por_defecto,
)
from app.ia.analizador_visita import (
    ResultadoAnalisis,
    analizar_acta_compromiso,
    analizar_acta_visita,
    analizar_tratamiento_datos,
)
from app.ia.analizador_visita2 import analizar_diagnostico, analizar_plan_negocio
from app.ia.analizador_03 import extraer_nombre_encuesta, validar_nombres_en_txrx
from app.core.gestores import buscar_gestor
from app.services.checklist import ChecklistWriter
from app.services.sharepoint import (
    descargar_carpeta_doc,
    descargar_carpeta_03,
    descargar_carpeta_04,
    descargar_visita_selectiva,
    descargar_visita2_selectiva,
)

logger = logging.getLogger(__name__)

_NO_ANALIZADO = ResultadoAnalisis(ok=True, alerta="No analizado (IA desactivada)")

# Mapa de keyword → función de análisis IA
_ANALIZADORES = {
    "ACTA_COMPROMISO":   analizar_acta_compromiso,
    "ACTA_VISITA_1":     analizar_acta_visita,
    "TRATAMIENTO_DATOS": analizar_tratamiento_datos,

    # Carpeta 02
    "ACTA_VISITA_2":     analizar_acta_visita,
}


def _armar_04(resultado_04: dict) -> tuple[str, str, str, str, str]:
    """Retorna (xlsx, pdf, firmas, cotizacion, web) para el checklist."""
    _FALTA_04  = "FALTA (04 no encontrada)"
    _FALTA_SUB = "FALTA (subcarpeta 01 no encontrada)"

    if not resultado_04.get("encontrada"):
        return _FALTA_04, _FALTA_04, _FALTA_04, _FALTA_04, _FALTA_04
    if not resultado_04.get("sub01_encontrada"):
        return _FALTA_SUB, _FALTA_SUB, _FALTA_SUB, _FALTA_SUB, _FALTA_SUB

    xlsx_encontrado = resultado_04.get("xlsx_encontrado", False)
    xlsx_valido     = resultado_04.get("xlsx_valido",     False)
    xlsx_resumen    = resultado_04.get("xlsx_resumen",    "FALTA")
    pdf_ok          = resultado_04.get("pdf_plan",        False)
    firmas_resumen  = resultado_04.get("firmas_resumen",  "N/A")
    cot_resumen     = resultado_04.get("cot_pdf_resumen", "N/A")
    web_txt         = resultado_04.get("web_excel",       "N/A")

    if not xlsx_encontrado:
        xlsx_txt = "FALTA PLAN_INVERSION.xlsx"
    elif xlsx_valido:
        xlsx_txt = f"OK — {xlsx_resumen}"
    else:
        xlsx_txt = f"Cotizaciones incompletas — {xlsx_resumen}"

    pdf_txt = "OK — Plan de inversión encontrado" if pdf_ok else "FALTA — PDF con 'Plan de inversión' no encontrado"

    return xlsx_txt, pdf_txt, firmas_resumen, cot_resumen, web_txt


class ValidadorDocumental:
    """
    Componente de validación documental.

    Uso mínimo:
        v = ValidadorDocumental("checklist.xlsx")
        v.procesar_matriz("matriz.xlsx")
    """

    # Flujos válidos disponibles
    FLUJOS_VALIDOS = {"00", "01", "02", "03", "04"}

    def __init__(
        self,
        ruta_checklist: str,
        callback_progreso: Optional[Callable[[int, int, int], None]] = None,
        lote_guardado: int = CHECKLIST_LOTE_GUARDADO,
        flujos: Optional[set] = None,
    ):
        self.ruta_checklist    = ruta_checklist
        self.callback_progreso = callback_progreso
        self.lote_guardado     = lote_guardado
        # None → todos los flujos; set vacío → ninguno (no tiene sentido, se trata como todos)
        self.flujos: set[str] = flujos if flujos else self.FLUJOS_VALIDOS

    # ── Punto de entrada principal ────────────────────────────────────────────

    def procesar_matriz(self, ruta_excel: str) -> None:
        t_inicio = time.perf_counter()
        logger.info("Cargando matriz: %s", ruta_excel)
        df    = self._cargar_matriz(ruta_excel)
        total = len(df)
        logger.info("Total filas: %d", total)

        checklist     = ChecklistWriter(self.ruta_checklist, lote_guardado=self.lote_guardado)
        ya_procesados = checklist.ids_procesados()
        procesadas    = len(ya_procesados)
        errores       = 0
        filas_omitidas: list[dict] = []   # filas que fallaron por timeout/error crítico

        if ya_procesados:
            logger.info("Reanudando: %d filas ya procesadas.", procesadas)

        for _, fila in df.iterrows():
            id_unico = str(fila.get("ID_unico", "")).strip()
            if not id_unico or id_unico.lower() in ("nan", "none", ""):
                continue
            if id_unico in ya_procesados:
                continue

            logger.info("[%d/%d] ID: %s", procesadas + 1, total, id_unico)
            t_fila = time.perf_counter()

            try:
                resultado = self._procesar_fila(fila)
                elapsed_fila = time.perf_counter() - t_fila

                if resultado.get("_tiene_error"):
                    errores += 1
                    logger.warning("  Alerta: %s", resultado.get("observaciones"))

                logger.info("  [%s] Fila procesada en %.1fs", id_unico, elapsed_fila)
                checklist.agregar_fila(resultado)
                procesadas += 1

            except (TimeoutError, ConnectionError, OSError) as exc:
                elapsed_fila = time.perf_counter() - t_fila
                logger.error(
                    "  [%s] TIMEOUT/RED tras %.1fs — omitida: %s",
                    id_unico, elapsed_fila, exc,
                )
                filas_omitidas.append(fila.to_dict())
                errores += 1
                # Limpiar temporales si quedaron
                dest_base = DOWNLOADS_DIR / id_unico
                if dest_base.exists():
                    try:
                        shutil.rmtree(dest_base)
                    except Exception:
                        pass

            except Exception as exc:
                elapsed_fila = time.perf_counter() - t_fila
                logger.error(
                    "  [%s] ERROR CRÍTICO tras %.1fs — omitida: %s",
                    id_unico, elapsed_fila, exc,
                )
                filas_omitidas.append(fila.to_dict())
                errores += 1
                dest_base = DOWNLOADS_DIR / id_unico
                if dest_base.exists():
                    try:
                        shutil.rmtree(dest_base)
                    except Exception:
                        pass

            if self.callback_progreso:
                self.callback_progreso(procesadas, total, errores)

        checklist.guardar()

        # Guardar pendientes si hubo omisiones
        ruta_pendientes = None
        if filas_omitidas:
            ruta_pendientes = str(self.ruta_checklist).replace("checklist.xlsx", "pendientes.xlsx")
            self._guardar_pendientes(filas_omitidas, ruta_pendientes)
            logger.warning(
                "%d fila(s) omitidas por error/timeout → %s",
                len(filas_omitidas), ruta_pendientes,
            )

        elapsed_total = time.perf_counter() - t_inicio
        promedio = elapsed_total / procesadas if procesadas else 0
        logger.info(
            "Terminado. %d/%d filas | %d alertas | %d omitidas | total: %.1fs | prom: %.1fs",
            procesadas, total, errores, len(filas_omitidas), elapsed_total, promedio,
        )
        return {
            "filas_omitidas":   len(filas_omitidas),
            "ruta_pendientes":  ruta_pendientes,
        }

    # ── Procesamiento de una fila ─────────────────────────────────────────────

    def _procesar_fila(self, fila: pd.Series) -> dict:
        id_unico   = str(fila.get("ID_unico",      "")).strip()
        modalidad  = str(fila.get("modalidad",     "")).strip().upper()
        unidad_doc = str(fila.get("unidad_doc",    "")).strip()
        link       = str(fila.get("carpetas_link", "")).strip()
        integrante = self._componer_integrante(fila)

        observaciones: list[str] = []
        tiene_error = False
        dest_base   = DOWNLOADS_DIR / id_unico

        # ── 1. Validar modalidad ──────────────────────────────────────────────
        if modalidad not in REGLAS:
            observaciones.append(f"Modalidad inválida: '{modalidad}'")
            return self._armar_resultado(
                id_unico, modalidad, unidad_doc, integrante,
                estados_por_defecto(modalidad), {}, {}, {}, {}, observaciones, tiene_error=True,
                flujos=self.flujos,
            )

        _omitido = {"encontrada": False, "observaciones": [], "tiene_error": False}

        # ── 2. Carpeta 00_DOCUMENTACION ───────────────────────────────────────
        docs_estado = estados_por_defecto(modalidad)
        if "00" in self.flujos:
            ruta_doc: Optional[Path] = None
            try:
                logger.info("  [%s] Buscando 00_DOCUMENTACION...", id_unico)
                ruta_doc, _ = descargar_carpeta_doc(link, dest_base / "doc")
                if ruta_doc is None:
                    observaciones.append("00_DOCUMENTACION no encontrada en SharePoint")
                    tiene_error = True
                else:
                    archivos   = self._listar_archivos(str(ruta_doc))
                    rutas_docs = self._detectar_documentos(archivos)
                    self._aplicar_regla_unidad_doc(rutas_docs, archivos, unidad_doc, modalidad)
                    presentes   = {doc: (r is not None) for doc, r in rutas_docs.items()}
                    docs_estado = evaluar_documentos(modalidad, presentes)
                    faltantes   = [d for d, e in docs_estado.items() if e == Estado.FALTA]
                    if faltantes:
                        observaciones.append(f"00 — Obligatorios ausentes: {', '.join(faltantes)}")
                        tiene_error = True
                    logger.info("  [%s] 00 OK — detectados: %s", id_unico, presentes)
            except Exception as exc:
                observaciones.append(f"00 — Descarga fallida: {exc}")
                tiene_error = True
        else:
            logger.info("  [%s] 00 omitido (no incluido en flujos)", id_unico)

        # ── 3. Carpeta 01_VISITA_1_CARACTERIZACION ────────────────────────────
        if "01" in self.flujos:
            resultado_visita = self._procesar_visita(id_unico, link, dest_base / "visita")
            observaciones.extend(resultado_visita["observaciones"])
            if resultado_visita["tiene_error"]:
                tiene_error = True
        else:
            resultado_visita = _omitido
            logger.info("  [%s] 01 omitido (no incluido en flujos)", id_unico)

        # ── 4. Carpeta 02_VISITA_2_DIAGNOSTICO ───────────────────────────────
        if "02" in self.flujos:
            resultado_visita2 = self._procesar_visita2(id_unico, link, dest_base / "visita2")
            observaciones.extend(resultado_visita2["observaciones"])
            if resultado_visita2["tiene_error"]:
                tiene_error = True
        else:
            resultado_visita2 = _omitido
            logger.info("  [%s] 02 omitido (no incluido en flujos)", id_unico)

        # ── 5. Carpeta 03_* (capacitación) ────────────────────────────────────
        if "03" in self.flujos:
            resultado_03 = self._procesar_03(id_unico, modalidad, link, dest_base / "cap03")
            observaciones.extend(resultado_03["observaciones"])
            if resultado_03["tiene_error"]:
                tiene_error = True
        else:
            resultado_03 = _omitido
            logger.info("  [%s] 03 omitido (no incluido en flujos)", id_unico)

        # ── 6. Carpeta 04_* (capitalización) ──────────────────────────────────
        if "04" in self.flujos:
            resultado_04 = self._procesar_04(id_unico, link, dest_base / "cap04")
            observaciones.extend(resultado_04["observaciones"])
            if resultado_04["tiene_error"]:
                tiene_error = True
        else:
            resultado_04 = _omitido
            logger.info("  [%s] 04 omitido (no incluido en flujos)", id_unico)

        # ── 7. Limpiar temporales ─────────────────────────────────────────────
        if dest_base.exists():
            try:
                shutil.rmtree(dest_base)
                logger.info("  [%s] Temporales eliminados.", id_unico)
            except Exception as exc:
                logger.warning("No se pudo eliminar '%s': %s", dest_base, exc)

        return self._armar_resultado(
            id_unico, modalidad, unidad_doc, integrante,
            docs_estado, resultado_visita, resultado_visita2, resultado_03,
            resultado_04, observaciones, tiene_error, self.flujos,
        )

    # ── Validación carpeta 01 ─────────────────────────────────────────────────

    def _procesar_visita(self, id_unico: str, link: str, dest_base: Path) -> dict:
        """
        Descarga selectivamente la carpeta 01_VISITA, cuenta media,
        y ejecuta análisis IA sobre los 3 documentos.
        """
        obs: list[str] = []
        tiene_error    = False

        # Estados iniciales: FALTA para todos
        docs_visita = {k: "FALTA" for k in DOCS_VISITA_ORDEN}
        ia_resultados = {
            "IA_COMPROMISO":  _NO_ANALIZADO,
            "IA_VISITA":      _NO_ANALIZADO,
            "IA_TRATAMIENTO": _NO_ANALIZADO,
        }
        conteo_media  = 0
        gestor_nombre = ""
        gestor_cedula = ""
        gestor_ok     = None  # None = no analizado, True/False = resultado

        try:
            logger.info("  [%s] Buscando 01_VISITA...", id_unico)
            info = descargar_visita_selectiva(link, dest_base)

            if not info["encontrada"]:
                obs.append("01_VISITA_1_CARACTERIZACION no encontrada")
                tiene_error = True
                return {
                    "encontrada": False, "docs_visita": docs_visita,
                    "conteo_media": 0, "ia": ia_resultados,
                    "gestor_nombre": "", "gestor_cedula": "", "gestor_ok": None,
                    "observaciones": obs, "tiene_error": tiene_error,
                }

            conteo_media = info["conteo_media"]
            logger.info("  [%s] 01 — media encontrada: %d", id_unico, conteo_media)

            # Verificar mínimo de fotos/videos
            if conteo_media < MIN_ARCHIVOS_MEDIA:
                obs.append(
                    f"01 — Fotos/videos insuficientes: {conteo_media} "
                    f"(mínimo {MIN_ARCHIVOS_MEDIA})"
                )
                tiene_error = True

            # Verificar y analizar los 3 documentos
            for keyword, ruta_archivo in info["docs_rutas"].items():
                if ruta_archivo is None:
                    docs_visita[keyword] = "FALTA"
                    obs.append(f"01 — {keyword} no encontrado")
                    tiene_error = True
                    continue

                docs_visita[keyword] = "OK"

                if IA_HABILITADO:
                    logger.info("  [%s] Analizando IA: %s", id_unico, keyword)
                    try:
                        res = _ANALIZADORES[keyword](ruta_archivo)
                        ia_key = {
                            "ACTA_COMPROMISO":   "IA_COMPROMISO",
                            "ACTA_VISITA_1":     "IA_VISITA",
                            "TRATAMIENTO_DATOS": "IA_TRATAMIENTO",
                        }[keyword]
                        ia_resultados[ia_key] = res
                        if not res.ok:
                            obs.append(f"IA {keyword}: {res.alerta}")
                            tiene_error = True

                        # Verificar gestor extraído del acta de visita
                        if keyword == "ACTA_VISITA_1":
                            gestor_nombre     = res.gestor_nombre
                            gestor_cedula     = res.gestor_cedula
                            gestor_cedulas_alt = res.gestor_cedulas_alt or []
                            if gestor_nombre or gestor_cedula:
                                encontrado, nombre_bd = buscar_gestor(
                                    gestor_cedula, gestor_nombre, gestor_cedulas_alt
                                )
                                gestor_ok = encontrado
                                if not encontrado:
                                    alt_txt = f" | variantes: {', '.join(gestor_cedulas_alt)}" if gestor_cedulas_alt else ""
                                    obs.append(
                                        f"Gestor no registrado: {gestor_nombre} "
                                        f"(Cédula: {gestor_cedula}{alt_txt})"
                                    )
                                    tiene_error = True
                                else:
                                    logger.info("  [%s] Gestor verificado: %s", id_unico, nombre_bd)
                            else:
                                logger.warning("  [%s] No se extrajo gestor del acta de visita", id_unico)
                    except Exception as exc:
                        logger.error("Error IA %s: %s", keyword, exc)
                        obs.append(f"IA {keyword}: error — {exc}")
                        tiene_error = True

        except Exception as exc:
            obs.append(f"01 — Error procesando visita: {exc}")
            tiene_error = True

        return {
            "encontrada":     True,
            "docs_visita":    docs_visita,
            "conteo_media":   conteo_media,
            "ia":             ia_resultados,
            "gestor_nombre":  gestor_nombre,
            "gestor_cedula":  gestor_cedula,
            "gestor_ok":      gestor_ok,
            "observaciones":  obs,
            "tiene_error":   tiene_error,
        }

    # ── Validación carpeta 02 ─────────────────────────────────────────────────

    def _procesar_visita2(self, id_unico: str, link: str, dest_base: Path) -> dict:
        obs: list[str] = []
        tiene_error = False
        docs_visita2 = {k: "FALTA" for k in DOCS_VISITA2_ORDEN}
        acta_visita2 = diagnostico = plan_negocio = "N/A"

        try:
            logger.info("  [%s] Buscando 02_VISITA...", id_unico)
            info = descargar_visita2_selectiva(link, dest_base)

            if not info["encontrada"]:
                obs.append("02_VISITA_2_DIAGNOSTICO no encontrada")
                tiene_error = True
                return {
                    "encontrada": False, "docs_visita2": docs_visita2,
                    "acta_visita_2": acta_visita2, "diagnostico": diagnostico,
                    "plan_negocio": plan_negocio,
                    "observaciones": obs, "tiene_error": tiene_error,
                }

            for keyword, ruta_archivo in info["docs_rutas"].items():
                if ruta_archivo is None:
                    docs_visita2[keyword] = "FALTA"
                    obs.append(f"02 — {keyword} no encontrado")
                    tiene_error = True
                    continue

                docs_visita2[keyword] = "OK"

                if not IA_HABILITADO:
                    continue

                try:
                    if keyword == "ACTA_VISITA_2":
                        res = analizar_acta_visita(ruta_archivo)
                        acta_visita2 = "OK" if res.ok else res.alerta
                        if not res.ok:
                            obs.append(f"Acta Visita 2: {res.alerta}")
                            tiene_error = True

                    elif keyword == "DIAGNOSTICO":
                        res = analizar_diagnostico(str(ruta_archivo))
                        diagnostico = "OK" if res.ok else res.alerta
                        if not res.ok:
                            obs.append(f"Diagnóstico: {res.alerta}")
                            tiene_error = True

                    elif keyword == "PLAN_NEGOCIO":
                        resultado_plan = analizar_plan_negocio(str(ruta_archivo))
                        if resultado_plan["completo"]:
                            plan_negocio = "OK"
                        else:
                            n = len(resultado_plan["faltantes"])
                            plan_negocio = f"Faltan {n} campos"
                            obs.append(f"Plan de Negocio incompleto ({n} campos)")
                            tiene_error = True

                except Exception as exc:
                    obs.append(f"IA 02 {keyword}: error — {exc}")
                    tiene_error = True

        except Exception as exc:
            obs.append(f"02 — Error procesando visita: {exc}")
            tiene_error = True

        return {
            "encontrada": True, "docs_visita2": docs_visita2,
            "acta_visita_2": acta_visita2, "diagnostico": diagnostico,
            "plan_negocio": plan_negocio,
            "observaciones": obs, "tiene_error": tiene_error,
        }

    # ── Validación carpeta 03 ─────────────────────────────────────────────────

    def _procesar_03(self, id_unico: str, modalidad: str, link: str, dest_base: Path) -> dict:
        _NA = {
            "encontrada": False,
            "encuestas_valor": "N/A", "grupal_valor": "N/A",
            "individual_valor": "N/A", "modulos_valor": "N/A",
            "asistencia_valor": "N/A",
            "alertas": {k: False for k in (
                "03_ENCUESTAS", "03_GRUPAL", "03_INDIVIDUAL",
                "03_MODULOS", "03_ASISTENCIA",
            )},
            "observaciones": [], "tiene_error": False,
        }

        try:
            info = descargar_carpeta_03(link, dest_base, id_unico)
        except Exception as exc:
            _NA["observaciones"] = [f"03 — Error descargando carpeta: {exc}"]
            _NA["tiene_error"] = True
            return _NA

        if not info["encontrada"]:
            _NA["observaciones"] = ["03_CAPACITACION no encontrada en SharePoint"]
            _NA["tiene_error"] = True
            return _NA

        obs: list[str] = []
        tiene_error = False
        alertas = {k: False for k in _NA["alertas"]}
        root = info["root"]
        modulos = info["modulos"]

        min_enc = MIN_ASISTENTES_03.get(modalidad, 1)
        encuestas = root["encuestas"]
        n_enc = len(encuestas)
        if n_enc < min_enc:
            obs.append(f"03 — Encuestas insuficientes: {n_enc}/{min_enc} mínimo")
            tiene_error = True
            alertas["03_ENCUESTAS"] = True
            encuestas_valor = f"FALTA ({n_enc}/{min_enc} mínimo)"
        else:
            encuestas_valor = f"OK ({n_enc} encuestas)"

        if root["grupal"] is None:
            obs.append("03 — Falta informe GRUPAL")
            tiene_error = True
            alertas["03_GRUPAL"] = True
            grupal_valor = "FALTA"
        else:
            grupal_valor = "OK"

        if root["individual"] is None:
            obs.append(f"03 — Falta informe individual (ID_{id_unico})")
            tiene_error = True
            alertas["03_INDIVIDUAL"] = True
            individual_valor = "FALTA"
        else:
            individual_valor = "OK"

        nombres: list[str] = []
        for ruta_enc in encuestas:
            try:
                nombre = extraer_nombre_encuesta(ruta_enc)
                if nombre:
                    nombres.append(nombre)
            except Exception as exc:
                logger.warning("Error extrayendo nombre de '%s': %s", ruta_enc.name, exc)

        if len(modulos) < 3:
            obs.append(f"03 — Solo {len(modulos)} carpeta(s) de módulo (se esperan 3)")
            tiene_error = True

        mod_alertas: list[str] = []
        asistencia_alertas: list[str] = []

        for mod in modulos:
            nombre_mod = mod["nombre"]
            if mod["txrx"] is None:
                mod_alertas.append(f"{nombre_mod}: sin TX_RX")
                obs.append(f"03 {nombre_mod} — Falta archivo TX_RX")
                tiene_error = True
            if mod["conteo_evidencia"] == 0:
                mod_alertas.append(f"{nombre_mod}: sin evidencia")
                obs.append(f"03 {nombre_mod} — Sin evidencia fotográfica")
                tiene_error = True
            if IA_HABILITADO and mod["txrx"] and nombres:
                try:
                    res = validar_nombres_en_txrx(mod["txrx"], nombres, nombre_mod)
                    if not res.ok:
                        asistencia_alertas.append(res.alerta)
                        obs.append(res.alerta)
                        tiene_error = True
                except Exception as exc:
                    obs.append(f"03 {nombre_mod} — Error IA TX_RX: {exc}")

        if mod_alertas:
            alertas["03_MODULOS"] = True
        modulos_valor = (
            f"OK ({len(modulos)} módulos)" if not mod_alertas
            else "ALERTA: " + " | ".join(mod_alertas)
        )
        if asistencia_alertas:
            alertas["03_ASISTENCIA"] = True
        asistencia_valor = "OK" if not asistencia_alertas else "ALERTA: " + " | ".join(asistencia_alertas)

        return {
            "encontrada": True,
            "encuestas_valor": encuestas_valor, "grupal_valor": grupal_valor,
            "individual_valor": individual_valor, "modulos_valor": modulos_valor,
            "asistencia_valor": asistencia_valor, "alertas": alertas,
            "observaciones": obs, "tiene_error": tiene_error,
        }

    # ── Detección de documentos en 00 ────────────────────────────────────────

    @staticmethod
    def _listar_archivos(ruta: str) -> list[Path]:
        import os
        paths: list[Path] = []
        for dirpath, _, archivos in os.walk(ruta):
            for archivo in archivos:
                paths.append(Path(dirpath) / archivo)
        return paths

    @staticmethod
    def _aplicar_regla_unidad_doc(
        rutas: Dict[str, Optional[Path]],
        archivos: list[Path],
        unidad_doc: str,
        modalidad: str,
    ) -> None:
        """
        Si algún archivo contiene el número de unidad_doc en su nombre y el
        tipo de documento objetivo (CEDULA o RUT según modalidad) aún no fue
        detectado, lo asigna como ese tipo.
        Modifica `rutas` en el lugar.
        """
        tipo_objetivo = UNIDAD_DOC_TIPO.get(modalidad)
        if not tipo_objetivo or not unidad_doc:
            return
        if rutas.get(tipo_objetivo) is not None:
            return  # ya fue detectado por palabra clave, no pisar

        # Limpia el número: solo dígitos para comparar
        numero = "".join(c for c in unidad_doc if c.isdigit())
        if not numero:
            return

        for archivo in archivos:
            stem_limpio = "".join(c for c in archivo.stem if c.isdigit())
            if numero in stem_limpio or stem_limpio in numero and len(stem_limpio) >= 6:
                rutas[tipo_objetivo] = archivo
                return

    @staticmethod
    def _detectar_documentos(archivos: list[Path]) -> Dict[str, Optional[Path]]:
        """Retorna {doc: ruta_del_archivo} o None si no se encontró."""
        rutas: Dict[str, Optional[Path]] = {doc: None for doc in PALABRAS_CLAVE}
        for archivo in archivos:
            stem_norm = normalizar(archivo.stem)
            for doc, palabras in PALABRAS_CLAVE.items():
                if rutas[doc] is None and any(p in stem_norm for p in palabras):
                    rutas[doc] = archivo
        return rutas

    # ── Validación carpeta 04 ─────────────────────────────────────────────────

    def _procesar_04(self, id_unico: str, link: str, dest_base: Path) -> dict:
        from app.ia.extractor import extraer_texto

        _FALLO = lambda obs: {
            "encontrada": False, "sub01_encontrada": False,
            "xlsx_plan": False, "pdf_plan": False,
            "observaciones": obs, "tiene_error": True,
        }

        try:
            info = descargar_carpeta_04(link, dest_base, id_unico)
        except Exception as exc:
            return _FALLO([f"04 — Error descargando carpeta: {exc}"])

        if not info["encontrada"]:
            return _FALLO(["04_CAPITALIZACION no encontrada en SharePoint"])

        if not info["sub01_encontrada"]:
            return {
                "encontrada": True, "sub01_encontrada": False,
                "xlsx_plan": False, "pdf_plan": False,
                "observaciones": ["04 — Subcarpeta 01 no encontrada dentro de 04"],
                "tiene_error": True,
            }

        archivos: list[Path] = info.get("archivos", [])
        obs: list[str]       = []
        tiene_error          = False

        from app.core.plan_inversion import (
            validar_plan_inversion,
            verificar_firmas_pdf,
            validar_cotizacion_seleccionada_en_pdf,
        )
        from app.core.cotizacion_web import (
            buscar_links_openai,
            tomar_screenshots,
            generar_excel_cotizaciones,
        )
        from app.config import OPENAI_API_KEY

        # ── 1. Verificar XLSX con "PLAN_INVERSION" en el nombre ──────────────
        xlsx_plan = next(
            (
                a for a in archivos
                if a.suffix.lower() == ".xlsx"
                and "plan_inversion" in normalizar(a.stem)
            ),
            None,
        )
        xlsx_valido      = False
        xlsx_resumen     = "FALTA"
        seleccionadas_xlsx = []

        if xlsx_plan is None:
            obs.append("04 — Falta PLAN_INVERSION.xlsx — no se validarán cotizaciones")
            tiene_error  = True
            xlsx_resumen = "FALTA — cotizaciones no analizadas"
            logger.info("  [%s] 04 — PLAN_INVERSION.xlsx NO encontrado", id_unico)
        else:
            logger.info("  [%s] 04 — PLAN_INVERSION.xlsx encontrado: %s", id_unico, xlsx_plan.name)
            resultado_xlsx     = validar_plan_inversion(xlsx_plan)
            xlsx_valido        = resultado_xlsx.ok
            xlsx_resumen       = resultado_xlsx.resumen
            seleccionadas_xlsx = resultado_xlsx.seleccionadas
            if not resultado_xlsx.ok:
                for alerta in resultado_xlsx.alertas:
                    obs.append(
                        f"04 — PLAN_INVERSION '{alerta.item}' "
                        f"Cot.{alerta.cotizacion}: falta {alerta.campo}"
                    )
                tiene_error = True
            logger.info("  [%s] 04 — PLAN_INVERSION: %s", id_unico, xlsx_resumen)

        # ── 2. Verificar PDF con "Plan de inversión" en su contenido ─────────
        # Prioridad: archivos con "FIRMA" en el nombre van primero.
        _pdfs_raw = [a for a in archivos if a.suffix.lower() == ".pdf"]
        pdfs = sorted(
            _pdfs_raw,
            key=lambda p: (0 if "firma" in p.name.lower() else 1),
        )
        pdf_plan_ok    = False
        pdf_para_cruce = None
        firmas_resumen = "N/A"
        cot_pdf_resumen = "N/A"

        if not pdfs:
            obs.append("04 — No hay PDF en la subcarpeta 01 — firmas y cotizaciones no verificadas")
            tiene_error    = True
            firmas_resumen = "N/A — sin PDF"
            logger.info("  [%s] 04 — Sin PDFs en subcarpeta 01", id_unico)
        else:
            for pdf in pdfs:
                texto = extraer_texto(pdf)
                if "plan de inversion" in normalizar(texto):
                    pdf_plan_ok    = True
                    pdf_para_cruce = pdf
                    logger.info("  [%s] 04 — 'Plan de inversión' encontrado en: %s", id_unico, pdf.name)
                    break
            if not pdf_plan_ok:
                obs.append(
                    "04 — Ningún PDF contiene 'Plan de inversión'"
                    f" (revisados: {', '.join(p.name for p in pdfs)})"
                    " — firmas y cotizaciones no verificadas"
                )
                tiene_error    = True
                firmas_resumen = "N/A — PDF sin contenido de Plan de inversión"
                logger.info("  [%s] 04 — 'Plan de inversión' NO encontrado en PDFs", id_unico)

        # ── 3. Verificar firmas en el PDF ────────────────────────────────────
        if pdf_para_cruce is not None:
            res_firmas     = verificar_firmas_pdf(pdf_para_cruce)
            firmas_resumen = res_firmas.resumen
            if not res_firmas.ok:
                obs.append(f"04 — Firmas: {res_firmas.resumen}")
                tiene_error = True
                logger.info("  [%s] 04 — Firmas insuficientes: %s", id_unico, res_firmas.resumen)
            else:
                logger.info("  [%s] 04 — Firmas OK: %s", id_unico, res_firmas.resumen)

        # ── 4. Cruce cotizaciones seleccionadas vs PDF ────────────────────────
        cot_pdf_ok = False
        if not seleccionadas_xlsx:
            cot_pdf_resumen = "N/A — sin cotizaciones seleccionadas en XLSX"
        elif pdf_para_cruce is None:
            cot_pdf_resumen = "N/A — PDF no disponible"
        else:
            resultado_cot_pdf = validar_cotizacion_seleccionada_en_pdf(
                seleccionadas_xlsx, pdf_para_cruce
            )
            cot_pdf_ok      = resultado_cot_pdf.ok
            cot_pdf_resumen = resultado_cot_pdf.resumen
            if not resultado_cot_pdf.ok:
                for alerta in resultado_cot_pdf.alertas:
                    obs.append(f"04 — Cotización seleccionada: {alerta}")
                tiene_error = True
            logger.info("  [%s] 04 — Cruce cotización/PDF: %s", id_unico, cot_pdf_resumen)

        # ── 5. Búsqueda web de precios de referencia ──────────────────────────
        # Solo si el cruce de cotizaciones pasó sin alertas.
        # Si hubo alertas en el paso 4 se solicita revisar primero los documentos.
        web_excel_nombre = "N/A"
        if not seleccionadas_xlsx:
            logger.info("  [%s] 04 — Búsqueda web omitida: sin cotizaciones seleccionadas", id_unico)
        elif not cot_pdf_ok:
            obs.append(
                "04 — Búsqueda web de precios de referencia omitida: "
                "revisar primero las alertas de cotizaciones en el PDF"
            )
            logger.info("  [%s] 04 — Búsqueda web omitida por alertas en cotizaciones", id_unico)
        else:
            logger.info("  [%s] 04 — Iniciando búsqueda web de precios de referencia...", id_unico)
            web_ok, web_productos, web_resumen = buscar_links_openai(seleccionadas_xlsx, OPENAI_API_KEY)
            logger.info("  [%s] 04 — GPT-4.1 web: %s", id_unico, web_resumen)

            if web_ok and web_productos:
                dir_ss = dest_base / "_screenshots"
                web_productos, screenshots = tomar_screenshots(web_productos, dir_ss, OPENAI_API_KEY)

                # Nombre del Excel = id_unico (sanitizado para nombre de archivo)
                nombre_seguro = re.sub(r'[\\/:*?"<>|]', "_", id_unico)
                dir_salida    = Path(self.ruta_checklist).parent
                ruta_web_xls  = dir_salida / f"{nombre_seguro}.xlsx"
                try:
                    generar_excel_cotizaciones(id_unico, web_productos, screenshots, ruta_web_xls, seleccionadas_xlsx)
                    web_excel_nombre = ruta_web_xls.name
                    logger.info("  [%s] 04 — Excel web generado: %s", id_unico, ruta_web_xls.name)
                except Exception as exc:
                    logger.error("  [%s] 04 — Error generando Excel web: %s", id_unico, exc)
                    obs.append(f"04 — Error generando Excel de precios web: {exc}")
                    tiene_error = True
            else:
                obs.append(f"04 — Búsqueda web: {web_resumen}")
                tiene_error = True

        return {
            "encontrada":       True,
            "sub01_encontrada": True,
            "xlsx_encontrado":  xlsx_plan is not None,
            "xlsx_valido":      xlsx_valido,
            "xlsx_resumen":     xlsx_resumen,
            "pdf_plan":         pdf_plan_ok,
            "firmas_resumen":   firmas_resumen,
            "cot_pdf_resumen":  cot_pdf_resumen,
            "web_excel":        web_excel_nombre,
            "observaciones":    obs,
            "tiene_error":      tiene_error,
        }

    # ── Helpers ───────────────────────────────────────────────────────────────

    @staticmethod
    def _componer_integrante(fila: pd.Series) -> str:
        claves = [
            "integrante_nombre1", "integrante_nombre2",
            "integrante_apellido1", "integrante_apellido2",
        ]
        partes = []
        for clave in claves:
            valor = str(fila.get(clave, "")).strip()
            if valor and valor.lower() not in ("nan", "none"):
                partes.append(valor)
        return " ".join(partes)

    @staticmethod
    def _armar_resultado(
        id_unico: str, modalidad: str, unidad_doc: str, integrante: str,
        docs_estado: Dict[str, Estado],
        resultado_visita: dict,
        resultado_visita2: dict,
        resultado_03: dict,
        resultado_04: dict,
        observaciones: list,
        tiene_error: bool,
        flujos: Optional[set] = None,
    ) -> dict:
        if flujos is None:
            flujos = {"00", "01", "02", "03", "04"}
        _OMITIDO = "—"  # valor en Excel cuando el flujo no fue ejecutado
        ia = resultado_visita.get("ia", {})
        dv = resultado_visita.get("docs_visita", {k: "N/A" for k in DOCS_VISITA_ORDEN})
        cm = resultado_visita.get("conteo_media", 0)
        encontrada = resultado_visita.get("encontrada", False)
        dv2 = resultado_visita2.get("docs_visita2", {})
        encontrada2 = resultado_visita2.get("encontrada", False)

        # ── 01_DOCUMENTOS: estado de existencia de los 3 docs ────────────────
        _NOMBRE_CORTO = {
            "ACTA_COMPROMISO":   "COMPROMISO",
            "ACTA_VISITA_1":     "VISITA",
            "TRATAMIENTO_DATOS": "TRATAMIENTO",
        }
        if "01" not in flujos:
            docs_01_valor = _OMITIDO
        elif not encontrada:
            docs_01_valor = "N/A"
        else:
            faltantes = [_NOMBRE_CORTO[k] for k in DOCS_VISITA_ORDEN if dv.get(k) == "FALTA"]
            presentes  = [_NOMBRE_CORTO[k] for k in DOCS_VISITA_ORDEN if dv.get(k) == "OK"]
            if faltantes:
                partes = [f"FALTA: {', '.join(faltantes)}"]
                if presentes:
                    partes.append(f"OK: {', '.join(presentes)}")
                docs_01_valor = " | ".join(partes)
            else:
                docs_01_valor = f"OK ({', '.join(presentes)})"

        # ── 01_FOTOS_VIDEOS ───────────────────────────────────────────────────
        if "01" not in flujos:
            fotos_01_valor = _OMITIDO
        elif not encontrada:
            fotos_01_valor = "N/A"
        elif cm >= MIN_ARCHIVOS_MEDIA:
            fotos_01_valor = f"OK ({cm} archivos)"
        else:
            fotos_01_valor = f"FALTA ({cm}/{MIN_ARCHIVOS_MEDIA} mínimo)"

        # ── 02_DOCUMENTOS ────────────────────────────────────────────────────────────
        _NOMBRE_CORTO_02 = {
            "ACTA_VISITA_2": "ACTA",
            "DIAGNOSTICO": "DIAGNOSTICO",
            "PLAN_NEGOCIO": "PLAN",
        }

        if "02" not in flujos:
            docs_02_valor = _OMITIDO
        elif not encontrada2:
            docs_02_valor = "N/A"
        else:
            faltantes = [
                _NOMBRE_CORTO_02[k]
                for k in DOCS_VISITA2_ORDEN
                if dv2.get(k) == "FALTA"
            ]

            presentes = [
                _NOMBRE_CORTO_02[k]
                for k in DOCS_VISITA2_ORDEN
                if dv2.get(k) == "OK"
            ]

            if faltantes:
                partes = [f"FALTA: {', '.join(faltantes)}"]

                if presentes:
                    partes.append(f"OK: {', '.join(presentes)}")

                docs_02_valor = " | ".join(partes)
            else:
                docs_02_valor = f"OK ({', '.join(presentes)})"

        # ── Revisión por documento (columnas individuales) ────────────────────
        def _valor_revision(ia_key: str) -> tuple[str, bool]:
            """Retorna (texto_celda, tiene_alerta) para una columna de revisión."""
            if not IA_HABILITADO:
                return "No aplica (IA desactivada)", False
            if not encontrada:
                return "N/A", False
            res = ia.get(ia_key, _NO_ANALIZADO)
            if res is None or "No analizado" in res.alerta:
                return "Error IA (ver log)", True
            if res.ok:
                return "OK", False
            return res.alerta, True

        if "01" not in flujos:
            val_compromiso = val_visita = val_tratamiento = _OMITIDO
            alerta_compromiso = alerta_visita = alerta_tratamiento = False
        else:
            val_compromiso,  alerta_compromiso  = _valor_revision("IA_COMPROMISO")
            val_visita,      alerta_visita      = _valor_revision("IA_VISITA")
            val_tratamiento, alerta_tratamiento = _valor_revision("IA_TRATAMIENTO")

        # ── Gestor ────────────────────────────────────────────────────────────
        gn  = resultado_visita.get("gestor_nombre", "")
        gc  = resultado_visita.get("gestor_cedula", "")
        gok = resultado_visita.get("gestor_ok", None)

        if "01" not in flujos:
            gestor_valor  = _OMITIDO
            alerta_gestor = False
        elif not encontrada or not IA_HABILITADO:
            gestor_valor  = "N/A"
            alerta_gestor = False
        elif gok is None:
            gestor_valor  = "No extraído"
            alerta_gestor = False
        elif gok:
            gestor_valor  = f"OK — {gn} ({gc})"
            alerta_gestor = False
        else:
            gestor_valor  = f"NO REGISTRADO — {gn} ({gc})"
            alerta_gestor = True

        # ── Acta Visita 2 ─────────────────────────────────────────────────────
        if "02" not in flujos:
            acta2_valor = _OMITIDO
            alerta_acta2 = False
        elif not encontrada2:
            acta2_valor = "N/A"
            alerta_acta2 = False
        else:
            acta2_valor = resultado_visita2.get("acta_visita_2", "N/A")
            alerta_acta2 = acta2_valor not in ("N/A",) and "OK" not in str(acta2_valor)

        # ── Diagnóstico ──────────────────────────────────────────────────────
        if "02" not in flujos:
            diagnostico_valor  = _OMITIDO
            alerta_diagnostico = False
        elif not encontrada2:
            diagnostico_valor  = "N/A"
            alerta_diagnostico = False
        else:
            diagnostico_valor  = resultado_visita2.get("diagnostico", "N/A")
            alerta_diagnostico = diagnostico_valor != "N/A" and "OK" not in str(diagnostico_valor)

        # ── Plan de Negocio ───────────────────────────────────────────────────
        if "02" not in flujos:
            plan_valor  = _OMITIDO
            alerta_plan = False
        elif not encontrada2:
            plan_valor  = "N/A"
            alerta_plan = False
        else:
            plan_valor  = resultado_visita2.get("plan_negocio", "N/A")
            alerta_plan = plan_valor != "N/A" and "OK" not in str(plan_valor)

        return {
            "ID_unico":        id_unico,
            "modalidad":       modalidad,
            "unidad_doc":      unidad_doc,
            "integrante":      integrante,
            # Carpeta 00
            "docs": (
                {doc: estado.value for doc, estado in docs_estado.items()}
                if "00" in flujos
                else {doc: _OMITIDO for doc in docs_estado}
            ),
            "docs_estado_raw": docs_estado if "00" in flujos else {},
            # Carpeta 01
            "01_documentos":      docs_01_valor,
            "01_fotos_videos":    fotos_01_valor,
            "01_acta_compromiso": val_compromiso,
            "01_acta_visita":     val_visita,
            "01_gestor":          gestor_valor,
            "01_tratamiento":     val_tratamiento,
            "01_alertas_por_doc": {
                "ACTA_COMPROMISO":   alerta_compromiso,
                "ACTA_VISITA":       alerta_visita,
                "GESTOR":            alerta_gestor,
                "TRATAMIENTO_DATOS": alerta_tratamiento,
            },
            # Carpeta 02
            "02_documentos":      docs_02_valor,
            "02_acta_visita":     acta2_valor,
            "02_diagnostico":     diagnostico_valor,
            "02_plan_negocio":    plan_valor,
            "02_alertas_por_doc": {
                "02_ACTA_VISITA":   alerta_acta2,
                "02_DIAGNOSTICO":   alerta_diagnostico,
                "02_PLAN_NEGOCIO":  alerta_plan,
            },
            # Carpeta 03
            "03_encuestas":    resultado_03.get("encuestas_valor",  _OMITIDO) if "03" in flujos else _OMITIDO,
            "03_grupal":       resultado_03.get("grupal_valor",     _OMITIDO) if "03" in flujos else _OMITIDO,
            "03_individual":   resultado_03.get("individual_valor", _OMITIDO) if "03" in flujos else _OMITIDO,
            "03_modulos":      resultado_03.get("modulos_valor",    _OMITIDO) if "03" in flujos else _OMITIDO,
            "03_asistencia":   resultado_03.get("asistencia_valor", _OMITIDO) if "03" in flujos else _OMITIDO,
            "03_alertas":      resultado_03.get("alertas", {}) if "03" in flujos else {},
            # Carpeta 04  (las claves se inyectan con ** en el dict padre)
            **( dict(zip(("04_xlsx", "04_pdf", "04_firmas", "04_cotizacion", "04_web"), _armar_04(resultado_04)))
                if "04" in flujos
                else {"04_xlsx": _OMITIDO, "04_pdf": _OMITIDO, "04_firmas": _OMITIDO, "04_cotizacion": _OMITIDO, "04_web": _OMITIDO} ),
            "observaciones":   "; ".join(observaciones),
            "_tiene_error":    tiene_error,
        }



    # ── Filas pendientes (omitidas por timeout / error crítico) ──────────────────

    @staticmethod
    def _guardar_pendientes(filas: list[dict], ruta: str) -> None:
        """Genera un Excel con las filas que no se procesaron para reintento manual."""
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pendientes"

        if not filas:
            wb.save(ruta)
            return

        columnas = list(filas[0].keys())

        fill_h = PatternFill(start_color="FF833C00", end_color="FF833C00", fill_type="solid")
        font_h = Font(name="Arial", size=9, bold=True, color="FFFFFFFF")
        alin_c = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, col_nombre in enumerate(columnas, start=1):
            c = ws.cell(row=1, column=col_idx, value=col_nombre)
            c.fill = fill_h
            c.font = font_h
            c.alignment = alin_c
        ws.row_dimensions[1].height = 22

        font_n = Font(name="Arial", size=9)
        for fila_idx, fila in enumerate(filas, start=2):
            for col_idx, col_nombre in enumerate(columnas, start=1):
                val = fila.get(col_nombre, "")
                if isinstance(val, float) and str(val) == "nan":
                    val = ""
                c = ws.cell(row=fila_idx, column=col_idx, value=val)
                c.font = font_n
                c.alignment = alin_c

        for col_idx, col_nombre in enumerate(columnas, start=1):
            max_ancho = max(len(str(col_nombre)), 12)
            for fila in filas:
                v = str(fila.get(col_nombre, "") or "")
                if len(v) > max_ancho:
                    max_ancho = len(v)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_ancho + 2, 60)

        ws.freeze_panes = "A2"
        wb.save(ruta)

    # ── Carga de la matriz ────────────────────────────────────────────────────

    @staticmethod
    def _cargar_matriz(ruta: str) -> pd.DataFrame:
        wb_tmp      = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        hoja_activa = wb_tmp.active.title
        wb_tmp.close()

        REQUERIDAS = {"ID_unico", "modalidad", "unidad_doc", "carpetas_link"}
        OPCIONALES = {
            "integrante_nombre1", "integrante_nombre2",
            "integrante_apellido1", "integrante_apellido2",
        }

        df = pd.read_excel(ruta, sheet_name=hoja_activa, dtype=str)
        faltantes = REQUERIDAS - set(df.columns)
        if faltantes:
            raise ValueError(
                f"Columnas requeridas no encontradas: {sorted(faltantes)}\n"
                f"Columnas presentes: {sorted(df.columns.tolist())}"
            )

        todas = list(REQUERIDAS | (OPCIONALES & set(df.columns)))
        return df[todas].fillna("")
