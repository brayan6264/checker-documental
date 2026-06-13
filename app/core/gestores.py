"""
Carga y búsqueda de gestores/talleristas aprobados desde el archivo xlsx de referencia.

Búsqueda en dos pasos:
  1. Por cédula (exacta, solo dígitos) — más confiable.
  2. Por nombre (flexible, ignora orden apellidos/nombres y acentos).
"""

import re
import logging
import unicodedata
from pathlib import Path

import openpyxl

from app.config import BASE_DIR

logger = logging.getLogger(__name__)

GESTORES_PATH = BASE_DIR / "data" / "Aprobación talleristas y gestores.xlsx"
_HOJAS = ["Base sin duplicados", "Cambio de roles"]  # se comparan con .strip()

_por_cedula: dict[str, str] = {}          # cedula_norm -> nombre_completo
_por_tokens: list[tuple[set, str]] = []   # (tokens_nombre, nombre_completo)
_cargado = False


def _norm_cedula(valor) -> str:
    return re.sub(r"\D", "", str(valor)) if valor else ""


def _tokens(nombre: str) -> set[str]:
    nfkd = unicodedata.normalize("NFKD", nombre.lower())
    sin_tilde = "".join(c for c in nfkd if not unicodedata.combining(c))
    return set(re.sub(r"[^a-z\s]", "", sin_tilde).split())


def _cargar() -> None:
    global _cargado
    if _cargado:
        return
    _cargado = True

    if not GESTORES_PATH.exists():
        logger.warning("Archivo de gestores no encontrado: %s", GESTORES_PATH)
        return

    try:
        wb = openpyxl.load_workbook(str(GESTORES_PATH), read_only=True, data_only=True)
        # Mapa nombre_limpio -> nombre_real (los títulos pueden tener espacios extra)
        hojas_disponibles = {s.strip(): s for s in wb.sheetnames}
        total = 0
        for nombre_hoja in _HOJAS:
            if nombre_hoja not in hojas_disponibles:
                continue
            ws   = wb[hojas_disponibles[nombre_hoja]]
            rows = ws.iter_rows(min_row=1, values_only=True)
            enc  = next(rows, None)
            if enc is None:
                continue

            enc_lower = [str(c).strip().lower() if c else "" for c in enc]
            idx_nombre = next((i for i, h in enumerate(enc_lower) if h == "nombre completo"), None)
            # Puede haber dos columnas "Numero de cedula"; la última suele ser más completa
            indices_cedula = [i for i, h in enumerate(enc_lower) if "cedula" in h or "cédula" in h]
            idx_cedula = indices_cedula[-1] if indices_cedula else None
            if idx_nombre is None:
                continue

            for row in rows:
                if not row or len(row) <= idx_nombre:
                    continue
                nombre = row[idx_nombre]
                if not nombre:
                    continue
                nombre_str = str(nombre).strip()
                # Si la última columna de cédula está vacía, intentar con la primera
                cedula_val = row[idx_cedula] if (idx_cedula is not None and len(row) > idx_cedula) else None
                if not cedula_val and len(indices_cedula) > 1:
                    cedula_val = row[indices_cedula[0]] if len(row) > indices_cedula[0] else None
                cedula_str = _norm_cedula(cedula_val)

                if cedula_str and cedula_str not in _por_cedula:
                    _por_cedula[cedula_str] = nombre_str

                toks = _tokens(nombre_str)
                if toks:
                    _por_tokens.append((toks, nombre_str))
                    total += 1

        wb.close()
        logger.info("Gestores cargados: %d por cédula, %d entradas de nombre", len(_por_cedula), total)
    except Exception as exc:
        logger.error("Error cargando archivo de gestores: %s", exc)


def buscar_gestor(
    cedula: str,
    nombre: str,
    cedulas_alt: list[str] | None = None,
) -> tuple[bool, str]:
    """
    Busca un gestor por cédula (principal + variantes alternativas) y/o nombre.

    Returns:
        (encontrado, nombre_en_bd)
    """
    _cargar()

    # 1. Cédula principal
    ced = _norm_cedula(cedula)
    if ced and ced in _por_cedula:
        return True, _por_cedula[ced]

    # 2. Cédulas alternativas (dígitos dudosos)
    for c_alt in (cedulas_alt or []):
        ced_alt = _norm_cedula(c_alt)
        if ced_alt and ced_alt in _por_cedula:
            logger.info("Gestor encontrado por cédula alternativa %s (original: %s)", ced_alt, ced)
            return True, _por_cedula[ced_alt]

    # 3. Nombre flexible (ignora orden apellidos/nombres y acentos)
    if nombre and nombre.strip():
        toks_busq = _tokens(nombre)
        if len(toks_busq) >= 2:
            for toks_bd, nombre_bd in _por_tokens:
                comunes = toks_busq & toks_bd
                # Coinciden si al menos (total_tokens - 1) palabras son iguales
                if len(comunes) >= max(2, len(toks_busq) - 1):
                    return True, nombre_bd

    return False, ""
