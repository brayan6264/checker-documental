"""
Búsqueda web de precios de referencia para cotizaciones seleccionadas.

Flujo por producto:
  1. buscar_links_openai()   — web_search_preview → URLs candidatas reales (no inventadas).
  2. tomar_screenshots()     — por ronda:
       a. Playwright visita las URLs → screenshot + precio DOM + base64.
       b. La visión recibe las capturas como IMÁGENES → valida producto + extrae precio.
       c. Si faltan válidas → busca URLs de reemplazo y repite.
       d. Se detiene al alcanzar el mínimo o agotar las rondas.
  3. generar_excel_cotizaciones() — Excel con capturas válidas + análisis de precio.
"""

import asyncio
import base64
import hashlib
import io
import json
import logging
import re
from datetime import date
from pathlib import Path
from typing import NamedTuple
from urllib.parse import urlparse

_RE_HOJA_INVALIDOS = re.compile(r'[\\/?*\[\]:]')

logger = logging.getLogger(__name__)

# Modelos diferenciados por tarea:
#  - BÚSQUEDA: gpt-4o devuelve URLs reales de tiendas mucho mejor que mini.
#  - ANÁLISIS (visión): mini es suficiente y mucho más barato para validar capturas.
#  - TÉRMINOS: tarea simple de texto, mini es suficiente.
_MODELO_BUSQUEDA = "gpt-4o"
# ANÁLISIS (visión): gpt-4o (no mini) — distingue mucho mejor el precio tachado/original
# del precio con descuento, que es una discriminación visual sutil (tamaño, línea encima).
_MODELO_ANALISIS = "gpt-4o"
_MODELO_TERMINOS = "gpt-4o-mini"

# Pool inicial de URLs por producto — cuanto mayor, más probabilidad de 3 válidos.
_URLS_INICIALES = 20

# URLs adicionales por ronda de reemplazo.
_URLS_POR_RONDA = 15

# Máximo de rondas (inicial + reemplazos). 6 rondas × 15 URLs = hasta 110 URLs por producto.
_MAX_RONDAS = 6

# Mínimo obligatorio de capturas válidas por producto.
_MIN_VALIDOS = 3

# ── Navegación Playwright ────────────────────────────────────────────────────
_NAV_TIMEOUT_MS  = 30_000   # más tiempo para páginas lentas colombianas
_NAV_WAIT_MS     = 3_000    # espera de render tras carga
_CONCURRENCIA_PW = 6        # reducido para evitar detección anti-bot simultánea

# Solo se bloquea linio.com.co (tienda cerrada en Colombia).
# Jumbo, Makro y Sodimac se dejan pasar — con mayor timeout suelen funcionar.
_DOMINIOS_LENTOS = {
    "linio.com.co",
}


def _url_lenta(url: str) -> bool:
    dominio = urlparse(url).netloc.lower().removeprefix("www.")
    return any(dominio == b or dominio.endswith("." + b) for b in _DOMINIOS_LENTOS)


# ── Fuentes adicionales de búsqueda ──────────────────────────────────────────

def _buscar_urls_serpapi(
    termino: str,
    n: int = 10,
    excluir_dominios: set | None = None,
    geo_nota: str = "",
) -> list[str]:
    """
    Busca URLs usando SerpAPI (índice Google Shopping + orgánico).

    Requiere en el entorno:
      SERPAPI_KEY — clave de API de serpapi.com

    Si la variable no está configurada, retorna [] sin lanzar excepción.
    La búsqueda se restringe a Colombia (gl=co) y en español (hl=es).
    Retorna URLs de fichas de producto reales encontradas en los resultados.
    """
    import os
    from urllib.parse import urlencode

    api_key = os.getenv("SERPAPI_KEY", "").strip()
    if not api_key:
        return []

    query = f"{termino} comprar Colombia precio{' ' + geo_nota if geo_nota else ''}"
    params = {
        "engine":  "google",
        "q":       query,
        "gl":      "co",
        "hl":      "es",
        "api_key": api_key,
        "num":     min(n + 5, 20),  # pedir de más para compensar filtros
    }
    endpoint = f"https://serpapi.com/search.json?{urlencode(params)}"
    try:
        import urllib.request as _req
        with _req.urlopen(endpoint, timeout=15) as resp:
            data = json.loads(resp.read())

        urls: list[str] = []
        seen_dominios: set[str] = set(excluir_dominios or set())

        # Shopping results (prioridad: fichas de producto directas)
        for r in (data.get("shopping_results") or []):
            link = r.get("link") or r.get("product_link") or ""
            if not link:
                continue
            dom = urlparse(link).netloc.lower().removeprefix("www.")
            if dom in seen_dominios or _url_bloqueada(link) or _url_lenta(link):
                continue
            seen_dominios.add(dom)
            urls.append(link)

        # Organic results
        for r in (data.get("organic_results") or []):
            link = r.get("link", "")
            if not link:
                continue
            dom = urlparse(link).netloc.lower().removeprefix("www.")
            if dom in seen_dominios or _url_bloqueada(link) or _url_lenta(link):
                continue
            seen_dominios.add(dom)
            urls.append(link)

        result = urls[:n]
        logger.info("  SerpAPI: %d URL(s) para '%s'", len(result), termino[:50])
        return result
    except Exception as exc:
        logger.warning("  SerpAPI falló ('%s'): %s", termino[:50], exc)
        return []


def _buscar_productos_serpapi_shopping(
    termino: str,
    n: int = 12,
    geo_nota: str = "",
) -> list["LinkProducto"]:
    """
    Busca productos usando SerpAPI Google Shopping (engine=google_shopping).

    A diferencia de la búsqueda orgánica, el motor Shopping devuelve datos
    ESTRUCTURADOS por producto: price, extracted_price, old_price,
    extracted_old_price (precio ANTES del descuento), source (tienda) y
    second_hand_condition (condición de segunda mano).

    Esto resuelve dos problemas de raíz:
      - Precio original: usamos extracted_old_price cuando hay descuento.
      - Productos usados: excluimos los que traen second_hand_condition.

    Retorna LinkProducto con precio_numero = precio ORIGINAL (si hay descuento)
    o el precio actual. Excluye usados, dominios bloqueados, Google y categorías.
    Si SERPAPI_KEY no está configurada, retorna [] sin lanzar excepción.
    """
    import os
    from urllib.parse import urlencode

    api_key = os.getenv("SERPAPI_KEY", "").strip()
    if not api_key:
        return []

    query = f"{termino}{' ' + geo_nota if geo_nota else ''}"
    params = {
        "engine":   "google_shopping",
        "q":        query,
        "gl":       "co",
        "hl":       "es",
        "location": "Colombia",
        "api_key":  api_key,
    }
    endpoint = f"https://serpapi.com/search.json?{urlencode(params)}"
    try:
        import urllib.request as _req
        with _req.urlopen(endpoint, timeout=20) as resp:
            data = json.loads(resp.read())
    except Exception as exc:
        logger.warning("  SerpAPI Shopping falló ('%s'): %s", termino[:50], exc)
        return []

    productos: list[LinkProducto] = []
    seen_dominios: set[str] = set()
    for r in (data.get("shopping_results") or []):
        # Excluir productos de segunda mano / usados
        if r.get("second_hand_condition"):
            continue
        link = r.get("link") or r.get("product_link") or ""
        if not link:
            continue
        dom = urlparse(link).netloc.lower().removeprefix("www.")
        # Saltar enlaces a Google (no son fichas de tienda real)
        if "google." in dom:
            continue
        if _url_bloqueada(link) or _url_lenta(link) or _url_es_categoria(link):
            continue
        if dom in seen_dominios:
            continue

        # Precio: preferir el ORIGINAL (antes del descuento) cuando exista
        precio_num = None
        for cand in (r.get("extracted_old_price"), r.get("extracted_price")):
            if cand is None:
                continue
            try:
                v = int(float(cand))
            except (ValueError, TypeError):
                continue
            if _es_precio_razonable(v):
                precio_num = v
                break
        if precio_num:
            precio_txt = _formatear_cop(precio_num)
        else:
            precio_txt = r.get("old_price") or r.get("price") or "N/A"

        seen_dominios.add(dom)
        productos.append(LinkProducto(url=link, precio_texto=precio_txt, precio_numero=precio_num))
        if len(productos) >= n:
            break

    logger.info("  SerpAPI Shopping: %d producto(s) para '%s'", len(productos), termino[:50])
    return productos


# Estado de Google CSE durante la corrida. Una vez que la cuota diaria devuelve
# 429, se marca cuota_agotada=True y no se vuelve a llamar (se reinicia al rearrancar).
_GCSE_ESTADO = {"cuota_agotada": False}


def _buscar_urls_google_cse(
    termino: str,
    n: int = 10,
    excluir_dominios: set | None = None,
    geo_nota: str = "",
) -> list[str]:
    """
    Busca URLs usando Google Custom Search API (CSE).

    Requiere en el entorno:
      GOOGLE_API_KEY  — clave de API de Google Cloud con Custom Search habilitado
      GOOGLE_CSE_ID   — ID del motor de búsqueda personalizado (cx)

    Si alguna variable no está configurada, retorna [] sin lanzar excepción.
    La búsqueda se restringe a Colombia (gl=co) y en español (lr=lang_es).
    """
    import os
    from urllib.parse import urlencode

    # Si la cuota diaria ya se agotó (429), no volver a llamar en esta corrida:
    # evita ~1s de latencia y ruido de logs por cada producto/reemplazo restante.
    if _GCSE_ESTADO["cuota_agotada"]:
        return []

    api_key = os.getenv("GOOGLE_API_KEY", "").strip()
    cse_id  = os.getenv("GOOGLE_CSE_ID",  "").strip()
    if not api_key or not cse_id:
        return []

    query = f"{termino} comprar Colombia precio{' ' + geo_nota if geo_nota else ''}"
    params = {
        "key": api_key,
        "cx":  cse_id,
        "q":   query,
        "num": min(n, 10),
        "gl":  "co",
        "lr":  "lang_es",
    }
    endpoint = f"https://www.googleapis.com/customsearch/v1?{urlencode(params)}"
    try:
        import urllib.request as _req
        import urllib.error as _uerr
        try:
            with _req.urlopen(endpoint, timeout=15) as resp:
                data = json.loads(resp.read())
        except _uerr.HTTPError as http_err:
            body = ""
            try:
                body = http_err.read().decode("utf-8", errors="replace")[:300]
            except Exception:
                pass
            if http_err.code == 429:
                # Cuota diaria agotada → deshabilitar CSE para el resto de la corrida.
                _GCSE_ESTADO["cuota_agotada"] = True
                logger.warning(
                    "  Google CSE: cuota diaria agotada (429) — deshabilitado para esta corrida."
                )
            else:
                logger.warning(
                    "  Google CSE HTTP %s para '%s': %s",
                    http_err.code, termino[:50], body or http_err.reason,
                )
            return []

        excl = set(excluir_dominios or set())
        urls = []
        for item in (data.get("items") or []):
            link = item.get("link", "")
            if not link:
                continue
            dom = urlparse(link).netloc.lower().removeprefix("www.")
            if dom in excl or _url_bloqueada(link) or _url_lenta(link):
                continue
            excl.add(dom)
            urls.append(link)

        result = urls[:n]
        logger.info("  Google CSE: %d URL(s) para '%s'", len(result), termino[:50])
        return result
    except Exception as exc:
        logger.warning("  Google CSE falló ('%s'): %s", termino[:50], exc)
        return []


# Estado de Brave Search durante la corrida (deshabilita tras 429 de rate-limit).
_BRAVE_ESTADO = {"cuota_agotada": False}


def _buscar_urls_brave(
    termino: str,
    n: int = 10,
    excluir_dominios: set | None = None,
    geo_nota: str = "",
) -> list[str]:
    """
    Busca URLs usando Brave Search API (alternativa a Google CSE con mejor cuota:
    plan gratuito de 2.000 consultas/mes vs 100/día de CSE).

    Requiere en el entorno:
      BRAVE_API_KEY — token de https://api.search.brave.com (Subscription Token)

    Si la variable no está configurada, retorna [] sin lanzar excepción.
    Restringe a Colombia (country=co) y español (search_lang=es).
    """
    import os
    from urllib.parse import urlencode

    if _BRAVE_ESTADO["cuota_agotada"]:
        return []

    api_key = os.getenv("BRAVE_API_KEY", "").strip()
    if not api_key:
        return []

    # OJO: Brave NO soporta Colombia en su parámetro 'country' (su enum solo tiene
    # AR/BR/CL/MX en LatAm). Enviar country=CO da HTTP 422. Por eso lo OMITIMOS y
    # sesgamos a Colombia con el texto de la consulta + search_lang=es. Los dominios
    # .co y el español surgen naturalmente; además filtramos por dominio después.
    query = f"{termino} comprar Colombia precio{' ' + geo_nota if geo_nota else ''}"
    params = {
        "q":           query,
        "search_lang": "es",
        "count":       min(max(n, 1), 20),
        "safesearch":  "off",
        "result_filter": "web",
    }
    endpoint = f"https://api.search.brave.com/res/v1/web/search?{urlencode(params)}"
    try:
        import urllib.request as _req
        import urllib.error as _uerr
        req = _req.Request(endpoint, headers={
            "Accept": "application/json",
            "X-Subscription-Token": api_key,
        })
        try:
            with _req.urlopen(req, timeout=15) as resp:
                data = json.loads(resp.read())
        except _uerr.HTTPError as http_err:
            if http_err.code == 429:
                _BRAVE_ESTADO["cuota_agotada"] = True
                logger.warning(
                    "  Brave Search: cuota/rate-limit agotado (429) — deshabilitado para esta corrida."
                )
            else:
                body = ""
                try:
                    body = http_err.read().decode("utf-8", errors="replace")[:200]
                except Exception:
                    pass
                logger.warning(
                    "  Brave Search HTTP %s para '%s': %s",
                    http_err.code, termino[:50], body or http_err.reason,
                )
            return []

        excl = set(excluir_dominios or set())
        urls = []
        for item in ((data.get("web") or {}).get("results") or []):
            link = item.get("url", "")
            if not link:
                continue
            dom = urlparse(link).netloc.lower().removeprefix("www.")
            if dom in excl or _url_bloqueada(link) or _url_lenta(link) or _url_no_navegable(link):
                continue
            excl.add(dom)
            urls.append(link)

        result = urls[:n]
        logger.info("  Brave Search: %d URL(s) para '%s'", len(result), termino[:50])
        return result
    except Exception as exc:
        logger.warning("  Brave Search falló ('%s'): %s", termino[:50], exc)
        return []


# ── Tipos ─────────────────────────────────────────────────────────────────────

class LinkProducto(NamedTuple):
    url:           str
    precio_texto:  str
    precio_numero: int | None


class ResultadoScreenshot(NamedTuple):
    ruta:          Path | None
    precio_texto:  str | None
    precio_numero: int | None
    sin_stock:     bool = False


class ProductoLinks(NamedTuple):
    item:        str
    descripcion: str | None
    links:       list[LinkProducto]


# Dominios bloqueados en código — el modelo los ignora a veces en el prompt
_DOMINIOS_BLOQUEADOS = {
    "mercadolibre.com.co",
    "mercadolibre.com",
    "mercadolibre.com.mx",
    "mercadolibre.com.ar",
    "mercadolibre.cl",
    "mercadolibre.com.pe",
    "mercadolibre.com.ve",
    "meli.com.co",
    # Amazon no es tienda colombiana — precios en USD o MXN
    "amazon.com",
    "amazon.com.mx",
    "amazon.com.co",
    "amazon.es",
    "amazon.co",
    # Redes sociales y marketplaces de usados — productos no son nuevos ni verificables
    "facebook.com",
    "fb.com",
    "instagram.com",
    "twitter.com",
    "x.com",
    "tiktok.com",
    "pinterest.com",
    "youtube.com",
    # Clasificados / segunda mano Colombia
    "olx.com.co",
    "olx.com",
    "tucarro.com.co",
    "tuinmueble.com.co",
    "clasificados.com.co",
    "milanuncios.com",
    "ebay.com",
    "ebay.co",
}

# Patrones de URL de categoría/búsqueda — no son fichas de producto
_RE_RUTA_CATEGORIA = re.compile(
    r'(?:^|/)(?:category|categories|categoria|categorias|'
    r'search|buscar|busqueda|resultados|collections|'
    r'listing|listings|brand|marcas)(?:/|$|\?)',
    re.IGNORECASE,
)
_RE_QUERY_BUSQUEDA = re.compile(
    r'(?:(?:^|&)(?:q|query|s|search|buscar|keyword)=)',
    re.IGNORECASE,
)


def _url_es_categoria(url: str) -> bool:
    """Retorna True si la URL es de categoría, listado o búsqueda (no ficha de producto)."""
    try:
        p = urlparse(url)
        return bool(_RE_RUTA_CATEGORIA.search(p.path) or _RE_QUERY_BUSQUEDA.search(p.query))
    except Exception:
        return False


# Extensiones de archivo que NO son páginas web navegables (Playwright las descarga
# y se cuelga el timeout completo). Fichas técnicas en PDF, catálogos, hojas de cálculo.
_RE_URL_DOCUMENTO = re.compile(
    r'\.(?:pdf|docx?|xlsx?|pptx?|zip|rar|7z|csv|txt|dwg|rtf)(?:$|\?)',
    re.IGNORECASE,
)
# Dominios de entidades oficiales: no venden productos (normas, fichas, resoluciones).
_DOMINIOS_NO_TIENDA = (
    "gov.co", "edu.co", "org.co", "minvivienda.gov.co", "ica.gov.co",
    "icbf.gov.co", "cra.gov.co", "normas.cra.gov.co",
)


def _url_no_navegable(url: str) -> bool:
    """Retorna True si la URL es un documento descargable o un dominio oficial (no tienda)."""
    try:
        p = urlparse(url)
        if _RE_URL_DOCUMENTO.search(p.path):
            return True
        dom = p.netloc.lower().removeprefix("www.")
        return any(dom == d or dom.endswith("." + d) for d in _DOMINIOS_NO_TIENDA)
    except Exception:
        return False


def _url_bloqueada(url: str) -> bool:
    """Retorna True si la URL pertenece a un dominio bloqueado."""
    dominio = urlparse(url).netloc.lower().removeprefix("www.")
    return any(dominio == b or dominio.endswith("." + b) for b in _DOMINIOS_BLOQUEADOS)


# Palabras clave que indican que el producto es un ser vivo (planta, semilla, animal, etc.)
# Cuando la descripción los contiene, la búsqueda debe restringirse al departamento del proyecto.
_RE_SERES_VIVOS = re.compile(
    r'\b(?:'
    r'planta[s]?|semilla[s]?|arbusto[s]?|árbol|arboles|árboles|palma[s]?|cactus|suculenta[s]?'
    r'|flor(?:es)?|orquídea[s]?|helecho[s]?|hierba[s]?|yerba[s]?|follaje'
    r'|abono[s]?|compost|sustrato[s]?|tierra\s+(?:fértil|para\s+(?:siembra|cultivo))'
    r'|almácigo[s]?|esquejes?|bulbo[s]?|tubérculo[s]?|rizoma[s]?'
    r'|animal(?:es)?|mascota[s]?|aves?|gallina[s]?|pollo[s]?|cerdo[s]?|conejo[s]?'
    r'|bovino[s]?|ovino[s]?|caprino[s]?|pez|peces|alevino[s]?'
    r'|cultivo[s]?|siembra|cosecha|vivero'
    r')\b',
    re.IGNORECASE,
)


def _es_producto_ser_vivo(descripcion: str) -> bool:
    """Retorna True si la descripción corresponde a plantas, semillas u otros seres vivos."""
    return bool(_RE_SERES_VIVOS.search(descripcion or ""))


_RE_PRECIO_COP = re.compile(
    r'(?:COP\s*|cop\s*|\$\s*)?(\d{1,3}(?:[.,]\d{3})+)(?:\s*(?:COP|cop))?'
)

_PRECIO_MIN_COP = 100
_PRECIO_MAX_COP = 50_000_000


# ── Helpers generales ─────────────────────────────────────────────────────────

def _normalizar_url(url: str) -> str:
    if "tienda.exito.com" in url:
        url = url.replace("tienda.exito.com", "www.exito.com")
    return url


# Parámetros de URL que son de TRACKING (no cambian la página). La misma ficha
# aparece con/sin estos (srsltid de Google Shopping, utm_*, gclid…) y se contaba
# como dos URLs distintas → capturas duplicadas. Se ignoran al deduplicar.
_PARAMS_TRACKING = {
    "srsltid", "gclid", "fbclid", "msclkid", "mc_eid", "_ga", "ref", "ref_",
    "igshid", "spm", "scm", "yclid", "dclid",
}


def _canonizar_url(url: str) -> str:
    """
    Clave canónica para DEDUPLICAR (no para navegar). Colapsa a la misma clave:
      - la misma página con/sin parámetros de tracking (srsltid, utm_*, gclid…)
      - barra final sobrante, host con/sin www, esquema http/https, fragmento #.
    Así dos enlaces a la MISMA ficha no producen dos capturas iguales.
    """
    try:
        from urllib.parse import urlsplit, urlunsplit, parse_qsl, urlencode
        p = urlsplit(url)
        host = p.netloc.lower().removeprefix("www.")
        path = p.path.rstrip("/") or "/"
        qs = [
            (k, v) for k, v in parse_qsl(p.query, keep_blank_values=False)
            if k.lower() not in _PARAMS_TRACKING and not k.lower().startswith("utm_")
        ]
        qs.sort()
        return urlunsplit(("", host, path, urlencode(qs), ""))
    except Exception:
        return (url or "").lower()


def _es_precio_razonable(num: int) -> bool:
    return _PRECIO_MIN_COP <= num <= _PRECIO_MAX_COP


def _parsear_precio_cop(texto: str) -> int | None:
    limpio = re.sub(r'[^\d.,]', '', texto.strip())
    if re.fullmatch(r'\d{1,3}(\.\d{3})+', limpio):
        return int(limpio.replace('.', ''))
    if re.fullmatch(r'\d{1,3}(,\d{3})+', limpio):
        return int(limpio.replace(',', ''))
    if re.fullmatch(r'\d+', limpio) and len(limpio) >= 4:
        return int(limpio)
    return None


def _formatear_cop(valor: int) -> str:
    return f"$ {valor:,}".replace(",", ".")


def _extraer_json(text: str) -> dict | list | None:
    text = text.strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass
    for pat in [
        r"```(?:json)?\s*(\{.*?\}|\[.*?\])\s*```",
        r"(\{.*\}|\[.*\])",
    ]:
        m = re.search(pat, text, re.DOTALL)
        if m:
            try:
                return json.loads(m.group(1))
            except json.JSONDecodeError:
                pass
    return None


# ── PASO 0 — Normalización del término de búsqueda con IA ────────────────────

def _derivar_terminos_busqueda(
    seleccionadas: list,
    api_key: str,
    fichas: dict | None = None,
) -> dict[str, str]:
    """
    Para CADA cotización deriva un término de búsqueda comercial preciso.

    Si se provee `fichas` (dict {item: FichaTecnica} leído de '1. Ficha técnica'),
    se usa la descripción general + especificaciones técnicas de la ficha para
    generar un término de búsqueda mucho más específico que coincida con el producto
    real (no solo el nombre genérico de la cotización).

    Retorna {item: termino_busqueda}. Si algo falla, cae a un recorte simple.
    """
    def _fallback(sel) -> str:
        # Si hay ficha técnica, usar la denominación como fallback (más precisa)
        if fichas:
            ficha = fichas.get(sel.item)
            if ficha and ficha.denominacion:
                return ficha.denominacion[:120]
        base = (sel.item or sel.descripcion or "").strip()
        return re.split(r'[.\n]', base)[0].strip()[:120] or base[:120]

    if not api_key:
        return {sel.item: _fallback(sel) for sel in seleccionadas}

    try:
        from openai import OpenAI
        client = OpenAI(api_key=api_key)
    except Exception:
        return {sel.item: _fallback(sel) for sel in seleccionadas}

    import json as _json

    entradas = []
    for i, sel in enumerate(seleccionadas):
        entrada: dict = {
            "id":     i,
            "nombre": (sel.item or "")[:300],
        }
        # Enriquecer con datos de la ficha técnica si están disponibles
        if fichas:
            ficha = fichas.get(sel.item)
            if ficha:
                if ficha.denominacion:
                    entrada["denominacion"] = ficha.denominacion[:200]
                if ficha.descripcion_general:
                    entrada["descripcion_general"] = ficha.descripcion_general[:400]
                if ficha.especificaciones:
                    # Truncar para no exceder el contexto del prompt
                    entrada["especificaciones"] = ficha.especificaciones[:400]
        elif sel.descripcion:
            entrada["descripcion"] = sel.descripcion[:400]
        entradas.append(entrada)

    tiene_fichas = fichas and any(fichas.get(sel.item) for sel in seleccionadas)

    if tiene_fichas:
        instruccion_extra = (
            "Cada ítem puede tener 'denominacion' (nombre oficial), 'descripcion_general' "
            "(para qué sirve) y 'especificaciones' (detalles técnicos como voltaje, capacidad, "
            "dimensiones, material, potencia). USA ESTOS DATOS para hacer el término muy "
            "específico e incluir las características técnicas clave que diferencian este "
            "producto de similares (ej: potencia en W, capacidad en L, voltaje, material)."
        )
    else:
        instruccion_extra = (
            "Si el texto es un prospecto o instrucción, extrae solo el NOMBRE del producto."
        )

    prompt = (
        "Eres un experto en compras colombianas. Para cada ítem genera el TÉRMINO DE BÚSQUEDA "
        "más efectivo para encontrarlo en tiendas online colombianas.\n\n"
        "REGLAS:\n"
        "- El término debe ser el nombre comercial del producto + sus características técnicas "
        "CLAVE (capacidad, potencia, voltaje, tamaño, material, modelo — las que diferencian "
        "este producto específico de otros similares).\n"
        "- Máximo 15 palabras. Específico, como lo buscaría un comprador técnico.\n"
        "- INCLUIR valores numéricos cuando estén disponibles (ej: '1500W', '25L', '110V').\n"
        "- NO incluir: instrucciones de uso, justificaciones, garantías, condiciones de entrega.\n"
        "- No inventar marcas que no aparezcan en el texto.\n"
        f"- {instruccion_extra}\n\n"
        f"Ítems:\n{_json.dumps(entradas, ensure_ascii=False, indent=2)}\n\n"
        "Responde SOLO con JSON:\n"
        '{"terminos": [{"id": 0, "busqueda": "..."}]}'
    )

    try:
        response = client.responses.create(
            model=_MODELO_TERMINOS,
            input=[{"role": "user", "content": prompt}],
        )
        data = _extraer_json(response.output_text) or {}
        por_id = {
            int(t["id"]): str(t.get("busqueda", "")).strip()
            for t in (data.get("terminos") or [])
            if "id" in t
        }
    except Exception as exc:
        logger.warning("  Normalización de términos falló: %s — usando fallback", exc)
        por_id = {}

    resultado: dict[str, str] = {}
    for i, sel in enumerate(seleccionadas):
        termino = por_id.get(i, "").strip()
        resultado[sel.item] = termino or _fallback(sel)
        if termino:
            logger.info("  Término búsqueda '%s' → '%s'", (sel.item or "")[:40], termino)
    return resultado


# ── PASO 1 — Búsqueda multi-fuente de URLs candidatas ────────────────────────

def _llamar_openai_web_search(
    client,
    prompt: str,
    n_max: int,
    dominios_vistos: set[str],
    item_label: str,
    query_label: str,
) -> list[LinkProducto]:
    """
    Ejecuta UNA llamada a OpenAI web_search_preview y retorna LinkProducto filtrados.
    Reutilizable para las dos queries distintas por producto.
    """
    try:
        response = client.responses.create(
            model=_MODELO_BUSQUEDA,
            tools=[{"type": "web_search_preview"}],
            tool_choice={"type": "web_search_preview"},
            input=[{"role": "user", "content": prompt}],
        )
        text = response.output_text
    except Exception as exc:
        logger.warning("  OpenAI web_search (%s) '%s': %s", query_label, item_label[:40], exc)
        return []

    data = _extraer_json(text)
    if not data:
        logger.warning("  OpenAI web_search (%s) JSON inválido para '%s'", query_label, item_label[:40])
        return []

    links: list[LinkProducto] = []
    for lk in (data.get("links") or []):
        if len(links) >= n_max:
            break
        url = str(lk.get("url", "")).strip()
        if not url:
            continue
        if _url_bloqueada(url):
            continue
        if _url_lenta(url):
            continue
        dominio = urlparse(url).netloc.lower().removeprefix("www.")
        if dominio in dominios_vistos:
            continue
        dominios_vistos.add(dominio)

        precio_txt = str(lk.get("precio", "") or "")
        precio_num = lk.get("precio_numero")

        if any(s in precio_txt.upper() for s in ("USD", "US$", "EUR", "€", "MXN")):
            precio_num = None
            precio_txt = "N/A"
        if isinstance(precio_num, float):
            precio_num = int(precio_num)
        if precio_num is not None and not _es_precio_razonable(precio_num):
            precio_num = None

        links.append(LinkProducto(url=url, precio_texto=precio_txt or "N/A", precio_numero=precio_num))

    return links


def buscar_links_openai(
    seleccionadas: list,
    api_key: str,
    departamento: str = "",
    fichas: dict | None = None,
) -> tuple[bool, list[ProductoLinks], str]:
    """
    Recopila URLs candidatas para cada producto usando MÚLTIPLES fuentes:
      1. OpenAI web_search_preview — query principal
      2. OpenAI web_search_preview — query alternativa (ángulo diferente)
      3. SerpAPI (Google Shopping + orgánico) — si SERPAPI_KEY está en entorno
      4. Brave Search — si BRAVE_API_KEY está en entorno
      5. Google Custom Search — si GOOGLE_API_KEY + GOOGLE_CSE_ID están en entorno

    Si se provee `fichas` (dict {item: FichaTecnica}), los prompts de búsqueda incluyen
    la descripción general y especificaciones técnicas para búsquedas más precisas.

    Las URLs se deduplican por clave canónica (ignora params de tracking) y se limita
    a 2 por dominio. Retorna (ok, productos, resumen).
    """
    try:
        from openai import OpenAI
    except ImportError:
        return False, [], "FALTA — openai no instalado"

    if not api_key:
        return False, [], "FALTA — OPENAI_API_KEY no configurada"

    client  = OpenAI(api_key=api_key)
    hoy_str = date.today().strftime("%d de %B de %Y")

    terminos = _derivar_terminos_busqueda(seleccionadas, api_key, fichas)

    productos_out: list[ProductoLinks] = []
    errores: list[str] = []

    for sel in seleccionadas:
        desc = terminos.get(sel.item) or (sel.item or "")[:120]

        # Construir descripción técnica enriquecida para los prompts si hay ficha
        detalle_tecnico = ""
        if fichas:
            ficha = fichas.get(sel.item)
            if ficha:
                partes = []
                if ficha.denominacion:
                    partes.append(f"Denominación oficial: {ficha.denominacion}")
                if ficha.descripcion_general:
                    partes.append(f"Descripción: {ficha.descripcion_general[:300]}")
                if ficha.especificaciones:
                    partes.append(f"Especificaciones técnicas: {ficha.especificaciones[:300]}")
                if partes:
                    detalle_tecnico = "\n" + "\n".join(partes)

        es_vivo   = _es_producto_ser_vivo(desc)
        geo_regla = ""
        geo_nota  = ""
        geo_serp  = ""
        if es_vivo and departamento:
            geo_regla = (
                f"6. RESTRICCIÓN GEOGRÁFICA OBLIGATORIA: este producto es un ser vivo "
                f"(planta, semilla, animal, etc.) y DEBE comprarse localmente. "
                f"TODAS las URLs deben ser de vendedores, viveros, granjas o distribuidores "
                f"ubicados en el departamento de {departamento} (Colombia). "
                f"No incluyas tiendas nacionales de e-commerce sin presencia en {departamento}."
            )
            geo_nota = (
                f"\nBUSCA ESPECÍFICAMENTE en {departamento}: viveros, productores agropecuarios, "
                f"agrotiendas, cooperativas o distribuidores locales."
            )
            geo_serp = departamento
            logger.info("  '%s': ser vivo → búsqueda restringida a %s", sel.item, departamento)

        # ── FUENTE A: OpenAI web_search — query principal ─────────────────
        prompt_1 = f"""Hoy es {hoy_str}. Necesito comprar este producto en Colombia.

PRODUCTO: {desc}{detalle_tecnico}{geo_nota}

USA LA HERRAMIENTA DE BÚSQUEDA WEB. Busca tiendas colombianas que vendan EXACTAMENTE este producto con las especificaciones indicadas.
Devuelve hasta 9 URLs de fichas de producto que REALMENTE hayas encontrado en los resultados.

⛔ REGLA #1 — NO INVENTES URLs:
- Cada URL DEBE haber aparecido en los resultados de tu búsqueda web.
- PROHIBIDO construir o adivinar URLs. PROHIBIDO inventar IDs/slugs de producto.
- Si solo encuentras 3 URLs reales, devuelve 3. Nunca rellenes con inventadas.

REGLAS adicionales:
1. URL directa a la ficha de UN producto (no categorías ni búsquedas).
2. Cada URL de un dominio distinto si es posible.
3. PROHIBIDO: MercadoLibre, Facebook, Instagram, OLX, clasificados, eBay, Amazon, redes sociales.
4. Solo tiendas que vendan en Colombia con precios en pesos colombianos (COP).
5. El producto debe ser NUEVO (no usado, no reacondicionado, no segunda mano).
6. El producto debe coincidir con las especificaciones técnicas indicadas arriba.
{geo_regla}

Responde ÚNICAMENTE con JSON (sin texto extra, sin ```):
{{"links":[{{"url":"https://...","precio":"$ ...","precio_numero":null}}]}}"""

        # ── FUENTE B: OpenAI web_search — query alternativa ───────────────
        prompt_2 = f"""Hoy es {hoy_str}. Busco dónde comprar este artículo en Colombia.

ARTÍCULO: {desc}{detalle_tecnico}{geo_nota}

USA LA HERRAMIENTA DE BÚSQUEDA WEB con una búsqueda diferente a la anterior.
Busca en tiendas en línea colombianas distintas a las usuales (Éxito, Alkosto, Falabella).
El producto DEBE cumplir con las especificaciones técnicas indicadas.
Devuelve hasta 6 URLs de fichas de producto que REALMENTE hayas encontrado.

⛔ REGLA CRÍTICA — NO INVENTES URLs. Solo URLs que aparecieron en tu búsqueda.

REGLAS:
1. Tiendas colombianas especializadas, ferretería, tecnología, agro, o similar según el producto.
2. URL directa a la ficha del producto con especificaciones que coincidan.
3. PROHIBIDO: MercadoLibre, Facebook, Instagram, OLX, clasificados, eBay, Amazon, redes sociales.
4. Solo precios en pesos colombianos (COP). Producto NUEVO (no usado ni segunda mano).
{geo_regla}

Responde ÚNICAMENTE con JSON:
{{"links":[{{"url":"https://...","precio":"$ ...","precio_numero":null}}]}}"""

        # Contador de URLs por dominio — permitimos hasta 2 por dominio para
        # tener más opciones sin abusar de un solo proveedor.
        _MAX_POR_DOMINIO = 2
        dominios_conteo: dict[str, int] = {}
        dominios_vistos: set[str] = set()   # para la deduplicación de OpenAI (1 por dominio)
        links_merged: list[LinkProducto] = []
        claves_vistas: set[str] = set()     # claves canónicas (dedup robusta de URLs)

        def _puede_agregar(url: str) -> bool:
            if not url or _canonizar_url(url) in claves_vistas:
                return False
            if _url_bloqueada(url) or _url_lenta(url) or _url_es_categoria(url) or _url_no_navegable(url):
                return False
            dom = urlparse(url).netloc.lower().removeprefix("www.")
            return dominios_conteo.get(dom, 0) < _MAX_POR_DOMINIO

        def _registrar(url: str, lk: LinkProducto):
            dom = urlparse(url).netloc.lower().removeprefix("www.")
            dominios_conteo[dom] = dominios_conteo.get(dom, 0) + 1
            claves_vistas.add(_canonizar_url(url))
            links_merged.append(lk)

        # Llamada A
        links_a = _llamar_openai_web_search(
            client, prompt_1, _URLS_INICIALES, dominios_vistos, sel.item, "query-1"
        )
        for lk in links_a:
            if _puede_agregar(lk.url):
                _registrar(lk.url, lk)

        # Llamada B (diferente query)
        links_b = _llamar_openai_web_search(
            client, prompt_2, 10, dominios_vistos, sel.item, "query-2"
        )
        for lk in links_b:
            if _puede_agregar(lk.url):
                _registrar(lk.url, lk)

        # ── FUENTE C0: SerpAPI Google Shopping (datos estructurados) ───────
        # Fuente PRIMARIA de precio: trae el precio original (old_price) y la
        # tienda directamente, sin necesidad de visitar la página.
        prods_shop = _buscar_productos_serpapi_shopping(desc, n=12, geo_nota=geo_serp)
        for lk in prods_shop:
            if _puede_agregar(lk.url):
                _registrar(lk.url, lk)

        # ── FUENTE C: SerpAPI orgánico (complemento para fichas no indexadas) ──
        urls_serp = _buscar_urls_serpapi(desc, n=12, excluir_dominios=set(), geo_nota=geo_serp)
        for u in urls_serp:
            if _puede_agregar(u):
                _registrar(u, LinkProducto(url=u, precio_texto="N/A", precio_numero=None))

        # Si SerpAPI retornó muy pocas URLs con el término técnico, reintenta
        # con un término más corto/genérico (primeras 4 palabras del nombre del ítem)
        if len(urls_serp) < 3:
            desc_corto = " ".join((sel.item or desc).split()[:4])
            if desc_corto and desc_corto != desc:
                urls_serp2 = _buscar_urls_serpapi(
                    desc_corto, n=8, excluir_dominios=set(), geo_nota=geo_serp
                )
                for u in urls_serp2:
                    if _puede_agregar(u):
                        _registrar(u, LinkProducto(url=u, precio_texto="N/A", precio_numero=None))
                logger.info(
                    "  SerpAPI retry (término corto '%s'): %d URL(s) adicionales",
                    desc_corto[:40], len(urls_serp2),
                )

        # ── FUENTE D: Brave Search (mejor cuota que CSE) ──────────────────
        urls_brave = _buscar_urls_brave(desc, n=12, excluir_dominios=set(), geo_nota=geo_serp)
        for u in urls_brave:
            if _puede_agregar(u):
                _registrar(u, LinkProducto(url=u, precio_texto="N/A", precio_numero=None))

        # ── FUENTE E: Google Custom Search (respaldo, si queda cuota) ──────
        urls_gcs = _buscar_urls_google_cse(desc, n=12, excluir_dominios=set(), geo_nota=geo_serp)
        for u in urls_gcs:
            if _puede_agregar(u):
                _registrar(u, LinkProducto(url=u, precio_texto="N/A", precio_numero=None))

        logger.info(
            "  '%s': pool total de %d URL(s) candidatas (A=%d B=%d shop=%d serp=%d brave=%d gcs=%d)",
            sel.item, len(links_merged), len(links_a), len(links_b),
            len(prods_shop), len(urls_serp), len(urls_brave), len(urls_gcs),
        )

        if links_merged:
            productos_out.append(ProductoLinks(item=sel.item, descripcion=desc, links=links_merged))
        else:
            logger.warning("  '%s': sin enlaces de ninguna fuente", sel.item)
            errores.append(f"Sin enlaces para {sel.item}")

    if not productos_out:
        return False, [], "Ninguna fuente retornó enlaces para ningún producto"

    total   = sum(len(p.links) for p in productos_out)
    resumen = f"OK — {len(productos_out)} producto(s), {total} URL(s) candidatas"
    if errores:
        resumen += f" | errores: {'; '.join(errores)}"
    return True, productos_out, resumen


def _capturar_concurrente(urls: list[str], dir_out: Path) -> dict:
    """
    Captura varias URLs EN PARALELO (hasta _CONCURRENCIA_PW a la vez).
    Retorna {url: (ruta_png, base64_png, None, None)}.
    El precio NO se extrae aquí: lo lee la visión desde la captura (fuente única).
    Reemplaza el bucle serial: corta el tiempo de cada ronda ~N veces.
    """
    if not urls:
        return {}
    try:
        return asyncio.run(_capturar_async(list(urls), dir_out))
    except Exception as exc:
        logger.error("Captura concurrente falló: %s", exc)
        return {u: (None, None, None, None) for u in urls}


_USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:125.0) Gecko/20100101 Firefox/125.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
]

# JS para cerrar banners de cookies/GDPR comunes en tiendas colombianas
_JS_CERRAR_COOKIES = """
() => {
    const SELS = [
        'button[id*="accept"]','button[id*="Accept"]','button[id*="aceptar"]',
        'button[class*="accept"]','button[class*="Accept"]','button[class*="aceptar"]',
        'button[id*="cookie"]','button[class*="cookie"]',
        '[aria-label*="accept"]','[aria-label*="Accept"]','[aria-label*="aceptar"]',
        '.cookie-accept','#cookie-accept','[data-cy="cookie-accept"]',
        '.btn-cookies-accept','#btn-accept-cookies',
        'button:has-text("Aceptar")','button:has-text("Aceptar todo")',
        'button:has-text("Accept")','button:has-text("Accept all")',
        '[class*="CookieBanner"] button','[class*="cookieBanner"] button',
        '[class*="gdpr"] button[class*="primary"]',
    ];
    for (const s of SELS) {
        try {
            const el = document.querySelector(s);
            if (el && el.offsetParent !== null) { el.click(); return true; }
        } catch(e) {}
    }
    return false;
}
"""


# JS que espera a que aparezca un precio del PRODUCTO PRINCIPAL (no de cards relacionadas),
# luego posiciona el viewport mostrando el título/imagen desde ARRIBA (no centrado en precio).
_JS_ESPERAR_Y_CENTRAR_PRECIO = r"""
async () => {
    // Excluir precios de productos relacionados / carruseles
    const EXCLUIR = '[class*="productCard"],[class*="ProductCard"],[class*="shelf-item"],' +
                    '[class*="ShelfItem"],[class*="carousel"],[class*="Carousel"],' +
                    '[class*="related"],[class*="Related"],[class*="recommendation"],' +
                    '[class*="Recommendation"],[class*="sugerido"],[class*="también"]';
    const noEsCard = el => !el.closest(EXCLUIR);

    const PRICE_SELS = [
        '[class*="price"],[class*="Price"],[class*="precio"],[class*="Precio"]',
        '[class*="original"],[class*="list-price"],[class*="compare-at"]',
        '[itemprop="price"],[data-price],[data-testid*="price"]',
        '.product-price,.product__price,.pdp-price',
    ];

    // Paso 1: ir al inicio para que el producto esté visible desde el tope
    window.scrollTo({top: 0, behavior: 'instant'});
    await new Promise(r => setTimeout(r, 300));

    // Paso 2: esperar hasta 5s a que aparezca un precio fuera de cards
    const start = Date.now();
    let precioEl = null;
    while (Date.now() - start < 5000) {
        for (const sel of PRICE_SELS) {
            try {
                const els = Array.from(document.querySelectorAll(sel));
                for (const el of els) {
                    if (!noEsCard(el)) continue;
                    const txt = el.innerText || el.textContent || "";
                    if (/[$]?\s*\d{2,3}[.,]\d{3}/.test(txt) || /\d{5,}/.test(txt)) {
                        precioEl = el;
                        break;
                    }
                }
            } catch(e) {}
            if (precioEl) break;
        }
        if (precioEl) break;
        await new Promise(r => setTimeout(r, 400));
    }

    if (!precioEl) {
        // No encontró precio — quedarse en el tope
        return false;
    }

    // Paso 3: mostrar la sección producto desde ARRIBA del viewport.
    // Usar el H1 como ancla (título + imagen + precio son visibles juntos).
    const h1 = document.querySelector('h1');
    let scrollY = 0;
    if (h1) {
        const rect = h1.getBoundingClientRect();
        // Dejar ~70px de margen para la barra de navegación superior
        scrollY = Math.max(0, window.scrollY + rect.top - 70);
    } else {
        // Sin H1: mostrar desde ~150px antes del precio
        const rect = precioEl.getBoundingClientRect();
        scrollY = Math.max(0, window.scrollY + rect.top - 250);
    }
    window.scrollTo({top: scrollY, behavior: 'smooth'});
    await new Promise(r => setTimeout(r, 700));
    return true;
}
"""


async def _capturar_async(urls: list[str], dir_out: Path) -> dict:
    import random
    from playwright.async_api import async_playwright

    resultados: dict = {}
    sem = asyncio.Semaphore(_CONCURRENCIA_PW)

    async def _ruta_handler(route):
        try:
            if route.request.resource_type in ("font", "media"):
                await route.abort()
            else:
                await route.continue_()
        except Exception:
            pass

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(
            headless=True,
            args=[
                "--disable-blink-features=AutomationControlled",
                "--no-sandbox",
                "--disable-dev-shm-usage",
                "--disable-gpu",
                "--disable-web-security",
                "--lang=es-CO",
                "--window-size=1280,900",
            ],
        )
        ctx = await browser.new_context(
            viewport={"width": 1280, "height": 900},
            locale="es-CO",
            timezone_id="America/Bogota",
            user_agent=random.choice(_USER_AGENTS),
            extra_http_headers={
                "Accept-Language": "es-CO,es;q=0.9,en;q=0.8",
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
                "sec-ch-ua-platform": '"Windows"',
            },
        )
        await ctx.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
            Object.defineProperty(navigator, 'plugins', {get: () => [1,2,3,4,5]});
            Object.defineProperty(navigator, 'languages', {get: () => ['es-CO','es','en-US','en']});
            window.chrome = {runtime: {}};
        """)
        await ctx.route("**/*", _ruta_handler)

        async def _navegar_y_capturar(page, url: str):
            """Toda la lógica de navegación + captura para una URL."""
            resp = await page.goto(url, timeout=_NAV_TIMEOUT_MS, wait_until="load")
            # Rechazar páginas con código HTTP de error (404, 410, 500, etc.)
            if resp and resp.status >= 400:
                raise Exception(f"HTTP {resp.status} — página no disponible")
            await page.wait_for_timeout(2_000)
            try:
                await page.evaluate(_JS_CERRAR_COOKIES)
                await page.wait_for_timeout(400)
            except Exception:
                pass
            try:
                precio_encontrado = await page.evaluate(_JS_ESPERAR_Y_CENTRAR_PRECIO)
                if not precio_encontrado:
                    await page.evaluate(
                        "() => window.scrollTo({top: Math.min(document.documentElement.scrollHeight * 0.3, 600), behavior: 'smooth'})"
                    )
                    await page.wait_for_timeout(800)
            except Exception:
                pass
            # El precio NO se extrae del DOM: la visión lo lee de la propia captura
            # (fuente única de verdad). Así el precio del Excel siempre coincide con
            # lo que se ve en la imagen — se elimina toda la clase de bugs "Excel ≠ captura".
            ruta = dir_out / (hashlib.md5(url.encode()).hexdigest()[:12] + ".png")
            await page.screenshot(path=str(ruta), full_page=False)
            b64 = base64.b64encode(ruta.read_bytes()).decode()
            return ruta, b64, None, None

        # Timeout absoluto por URL = NAV_TIMEOUT + margen para JS/screenshot
        _TIMEOUT_ABSOLUTO = (_NAV_TIMEOUT_MS / 1000) + 15  # segundos

        async def _una(url: str):
            async with sem:
                page = None
                try:
                    # new_page() TAMBIÉN puede colgarse si el navegador quedó en mal
                    # estado tras un "Page crashed" → lo acotamos con timeout.
                    page = await asyncio.wait_for(ctx.new_page(), timeout=15)
                    # Si el renderer se cae a mitad de operación, el evento 'crash'
                    # rompe el await en curso para no esperar el timeout completo.
                    page.on("crash", lambda _p: logger.warning("  Página crasheó: %s", url))
                    ruta, b64, p_txt, p_num = await asyncio.wait_for(
                        _navegar_y_capturar(page, url),
                        timeout=_TIMEOUT_ABSOLUTO,
                    )
                    resultados[url] = (ruta, b64, p_txt, p_num)
                except asyncio.TimeoutError:
                    logger.warning("  Screenshot timeout absoluto (%ds): %s", int(_TIMEOUT_ABSOLUTO), url)
                    resultados[url] = (None, None, None, None)
                except Exception as exc:
                    logger.warning("  Screenshot error %s: %s", url, exc)
                    resultados[url] = (None, None, None, None)
                finally:
                    if page is not None:
                        try:
                            await asyncio.wait_for(page.close(), timeout=5)
                        except Exception:
                            pass

        # Backstop global de la ronda: aunque cada URL ya está acotada, esto
        # garantiza que la ronda completa nunca quede colgada indefinidamente.
        _timeout_ronda = 60 + _TIMEOUT_ABSOLUTO * (len(urls) // _CONCURRENCIA_PW + 2)
        try:
            await asyncio.wait_for(
                asyncio.gather(*[_una(u) for u in urls]),
                timeout=_timeout_ronda,
            )
        except asyncio.TimeoutError:
            logger.warning(
                "  Ronda de capturas excedió %ds — se continúa con lo obtenido.",
                int(_timeout_ronda),
            )
        # Limpieza acotada: un navegador con un renderer crasheado puede dejar
        # COLGADO el cierre indefinidamente. Lo acotamos; si no responde, el
        # context manager `async with async_playwright()` mata el proceso al salir.
        try:
            await asyncio.wait_for(ctx.unroute_all(behavior="ignoreErrors"), timeout=10)
        except Exception:
            pass
        try:
            await asyncio.wait_for(ctx.close(), timeout=10)
        except Exception:
            pass
        try:
            await asyncio.wait_for(browser.close(), timeout=15)
        except Exception:
            logger.warning("  browser.close() no respondió a tiempo — el navegador se forzará al salir.")

    return resultados


# ── PASO 2b/2c — Verificación visual de capturas (IMÁGENES, no PDF) ───────────

# Imágenes por llamada de visión. Lotes pequeños evitan que la respuesta JSON
# se trunque: con muchas imágenes el modelo cortaba el array y se perdían TODAS
# las capturas de la ronda (caso Compost: 15 imágenes → JSON truncado → 0 válidas).
_VISION_BATCH = 6


def _verificar_con_vision(
    descripcion: str,
    capturas: list[tuple[str, str]],   # [(url, base64_png), ...]
    api_key: str,
) -> list[dict]:
    """Divide las capturas en lotes pequeños y verifica cada uno (evita truncar el JSON)."""
    from openai import OpenAI
    client = OpenAI(api_key=api_key)
    salida: list[dict] = []
    for inicio in range(0, len(capturas), _VISION_BATCH):
        lote = capturas[inicio:inicio + _VISION_BATCH]
        salida.extend(_verificar_vision_lote(descripcion, lote, client))
    return salida


def _verificar_vision_lote(
    descripcion: str,
    capturas: list[tuple[str, str]],   # [(url, base64_png), ...]
    client,
) -> list[dict]:
    """
    Envía las capturas como IMÁGENES a la visión para verificar cada página y
    extraer el precio. Enviar imágenes individuales (no un PDF) es más fiable:
    evita las negativas del modelo ("no puedo analizar PDFs") y mejora la lectura.

    Criterio de validez (NO ser rígido — productos similares SÍ valen):
      - válido: ficha de producto igual, similar, equivalente o del mismo sector.
      - inválido SOLO si: página en blanco / error 404 / CAPTCHA / listado-categoría /
        producto de categoría completamente distinta.

    Retorna lista de dicts: {url, valido, precio_texto, precio_numero, en_stock, motivo}.
    """
    n = len(capturas)
    prompt_texto = f"""Eres un verificador de páginas de e-commerce colombianas.

Producto buscado: "{descripcion}"

Analiza las {n} captura(s) de pantalla y para CADA UNA devuelve un objeto JSON con:

1. "valido" (boolean):
   - true ÚNICAMENTE si la captura muestra la FICHA INDIVIDUAL de UN solo producto
     que corresponde a "{descripcion}" o a un producto SIMILAR/EQUIVALENTE del mismo sector,
     CON al menos un precio visible en la pantalla (en pesos colombianos COP, $ o número
     de 5+ dígitos que razonablemente sea COP).
     Distinta marca, modelo o tamaño SÍ vale — sé generoso con el producto.
     Producto agotado con precio visible SÍ vale (el precio de referencia sirve igual).
   - false en CUALQUIERA de estos casos:
     • CUADRÍCULA o LISTADO con DOS o más productos distintos visibles → SIEMPRE false,
       aunque cada producto muestre su precio. Esto incluye páginas de categoría, resultados
       de búsqueda, páginas de marca, carruseles de recomendaciones como pantalla principal.
     • Página de error del sitio: "404", "Página no encontrada", "Este producto ya no está
       disponible", "Hay error", "producto no disponible", "enlace roto", o cualquier mensaje
       que indique que el producto o la URL ya no existe → SIEMPRE false aunque se vea precio.
     • Página en blanco, CAPTCHA, pantalla de login/acceso obligatorio sin contenido.
     • Producto de categoría COMPLETAMENTE distinta (buscando refrigerador, aparece ropa).
     • Imagen de carga / spinner sin contenido visible.
     • El único precio visible está en moneda EXTRANJERA confirmada (USD $X, €X, MXN $X).
     • NO hay NINGÚN precio visible en la captura (producto sin stock sin precio, ficha
       sin precio cargado).

PRECIOS — MUY IMPORTANTE. Mira SOLO los precios del producto PRINCIPAL (cerca del título
y del botón de compra; ignora precios de "productos relacionados"/"también compraron").
Extrae estos DOS campos por separado, OBSERVANDO bien cuál está tachado:

2. "precio_normal_numero" (integer o null): el precio REGULAR / DE LISTA / ANTES del descuento.
   • Cuando hay descuento, es el precio MÁS ALTO, casi siempre TACHADO (con una línea encima),
     en letra más pequeña o gris, a veces con etiqueta "Antes", "Precio normal" o "P. de lista".
   • Si NO hay descuento (un solo precio en pantalla), este campo = ese único precio.
   • Entero sin puntos ni símbolos: "$ 289.900" → 289900.

3. "precio_oferta_numero" (integer o null): el precio CON descuento / de oferta, si existe.
   • Es el precio MÁS BAJO, generalmente grande y resaltado, a veces con "-23%", "Oferta" o "Ahorra $X".
   • Si NO hay descuento → null.

   ⛔ REGLA DE ORO: el precio que necesitamos es el NORMAL (sin descuento). Por eso DEBES
   separar ambos. NUNCA pongas el precio de oferta en "precio_normal_numero".
   Ejemplo real: ves "$ 239.900 und" grande + "Ahorra $50.000" + "$ 289.900" tachado debajo →
   precio_normal_numero = 289900, precio_oferta_numero = 239900.
   Otro ejemplo: ves "$ 37.198" con "-23%" y "$ 48.500" tachado →
   precio_normal_numero = 48500, precio_oferta_numero = 37198.
   Si el precio está en moneda extranjera (USD/EUR/MXN) → ambos null.
   Si NO hay precio visible → ambos null.

4. "moneda" (string o null): la moneda del precio si es visible ("COP", "USD", "EUR", etc.).
   Si no se distingue pero es una tienda colombiana, asumir "COP".

5. "en_stock" (boolean): ¿el producto está DISPONIBLE para comprar?
   - false si ves CLARAMENTE: "Agotado", "Producto agotado", "Sin stock disponible",
     "Producto sin stock", "Sold out", "No disponible", "Vendido", "Próximamente",
     o el aviso de "Estamos preparando la imagen de este producto" (ficha incompleta).
   - true si el producto se puede comprar (botón "Agregar al carro"/"Comprar" activo),
     INCLUSO si dice "Sin stock en tienda" pero ofrece "Despacho a domicilio"/envío
     (eso significa disponible online = true).
   - Si no hay ninguna señal de agotado, asumir true.

6. "motivo" (string o null): si "valido" es false, explica brevemente por qué; si true, null.

Responde ÚNICAMENTE con un array JSON de {n} objeto(s) en el MISMO ORDEN que las imágenes,
sin texto adicional. Ejemplo para 2 imágenes:
[
  {{"valido": true, "precio_normal_numero": 289900, "precio_oferta_numero": 239900, "moneda": "COP", "en_stock": true, "motivo": null}},
  {{"valido": false, "precio_normal_numero": null, "precio_oferta_numero": null, "moneda": null, "en_stock": false, "motivo": "producto agotado"}}
]"""

    content: list[dict] = [{"type": "text", "text": prompt_texto}]
    for _url, b64 in capturas:
        content.append({
            "type": "image_url",
            "image_url": {"url": f"data:image/png;base64,{b64}", "detail": "high"},
        })

    try:
        resp = client.chat.completions.create(
            model=_MODELO_ANALISIS,
            messages=[{"role": "user", "content": content}],
            max_tokens=2000,
        )
        texto = resp.choices[0].message.content or ""
        logger.debug("Vision respuesta: %s", texto[:500])
        datos = _extraer_json(texto)

        # Parser TOLERANTE: si el array viene incompleto (p.ej. truncado), se mapea
        # lo que llegó y las faltantes se marcan inválidas — NO se descartan todas.
        if isinstance(datos, list) and datos:
            if len(datos) != n:
                logger.warning(
                    "Vision: se esperaban %d objeto(s), llegaron %d — se mapea lo disponible.",
                    n, len(datos),
                )
            resultado = []
            for i in range(n):
                d = datos[i] if i < len(datos) else {}

                def _num(v):
                    if v is None:
                        return None
                    try:
                        return int(float(v))
                    except (ValueError, TypeError):
                        return None

                precio_normal = _num(d.get("precio_normal_numero"))
                precio_oferta = _num(d.get("precio_oferta_numero"))

                # El precio que se publica es SIEMPRE el NORMAL (sin descuento).
                # Solo si no hay normal se cae al de oferta (caso degenerado).
                precio_num = precio_normal if precio_normal else precio_oferta
                if precio_normal and precio_oferta and precio_oferta > precio_normal:
                    # La visión los invirtió: el normal nunca es menor que la oferta.
                    precio_num = precio_oferta
                    logger.warning(
                        "Vision: normal(%s) < oferta(%s) — corregido al mayor.",
                        precio_normal, precio_oferta,
                    )

                moneda = (d.get("moneda") or "").upper()
                if moneda and moneda not in ("COP", ""):
                    logger.warning("Vision: precio en %s descartado (no COP)", moneda)
                    precio_num = None

                if precio_num and not _es_precio_razonable(precio_num):
                    logger.warning("Vision: precio fuera de rango (%s) descartado", precio_num)
                    precio_num = None

                precio_txt = _formatear_cop(precio_num) if precio_num else None

                # en_stock: por defecto True si la visión no lo determina
                en_stock_raw = d.get("en_stock")
                en_stock = True if en_stock_raw is None else bool(en_stock_raw)

                resultado.append({
                    "url":           capturas[i][0],
                    "valido":        bool(d.get("valido")),
                    "precio_texto":  precio_txt,
                    "precio_numero": precio_num,
                    "en_stock":      en_stock,
                    "motivo":        d.get("motivo") or "",
                })
            return resultado

        logger.warning(
            "Vision: array de longitud inesperada (esperado %d). Respuesta: %s",
            n, texto[:200],
        )
    except Exception as exc:
        logger.error("Vision API error: %s", exc)

    # Fallback conservador: marcar inválidos (no inflar el Excel con capturas dudosas)
    return [
        {"url": url, "valido": False, "precio_texto": None, "precio_numero": None,
         "en_stock": True, "motivo": "verificación de visión no disponible"}
        for url, _ in capturas
    ]


# ── PASO 2d — Búsqueda de URLs adicionales ───────────────────────────────────

def _buscar_links_reemplazo(
    descripcion: str,
    n: int,
    excluir: set[str],
    api_key: str,
    departamento: str = "",
) -> list[str]:
    """
    Busca hasta N URLs nuevas reales usando MÚLTIPLES fuentes, excluyendo dominios
    ya intentados. Usa OpenAI web_search + SerpAPI + Google CSE.
    """
    excluir_dominios = {
        urlparse(u).netloc.lower().removeprefix("www.")
        for u in excluir
    }

    es_vivo   = _es_producto_ser_vivo(descripcion)
    geo_regla = ""
    geo_nota  = ""
    geo_serp  = ""
    if es_vivo and departamento:
        geo_regla = (
            f"5. RESTRICCIÓN GEOGRÁFICA OBLIGATORIA: este producto es un ser vivo. "
            f"TODOS los links deben ser de vendedores en el departamento de "
            f"{departamento} (Colombia).\n"
        )
        geo_nota = f" en el departamento de {departamento}"
        geo_serp = departamento

    urls_finales: list[str] = []
    dominios_ok: set[str]   = set(excluir_dominios)
    excluir_canon: set[str] = {_canonizar_url(u) for u in excluir}

    def _agregar(url: str) -> bool:
        if not url or _canonizar_url(url) in excluir_canon:
            return False
        if _url_bloqueada(url) or _url_lenta(url) or _url_es_categoria(url) or _url_no_navegable(url):
            return False
        dom = urlparse(url).netloc.lower().removeprefix("www.")
        if dom in dominios_ok:
            return False
        dominios_ok.add(dom)
        excluir_canon.add(_canonizar_url(url))
        urls_finales.append(url)
        return True

    # ── Fuente A: OpenAI web_search ───────────────────────────────────────
    if api_key:
        try:
            from openai import OpenAI
            client = OpenAI(api_key=api_key)
            excl_str = ", ".join(sorted(excluir_dominios)) or "ninguno"
            prompt = (
                f"USA LA HERRAMIENTA DE BÚSQUEDA WEB. Busca tiendas colombianas{geo_nota} que vendan:\n"
                f"PRODUCTO: {descripcion}\n\n"
                f"Ya intenté estos dominios y fallaron (NO repetirlos): {excl_str}\n\n"
                f"⛔ NO INVENTES URLs. Solo URLs que aparecieron en tu búsqueda real.\n"
                f"REGLAS: URL directa a ficha de UN producto NUEVO (no usado). "
                f"PROHIBIDO: MercadoLibre, Facebook, Instagram, OLX, clasificados, eBay, Amazon. "
                f"Solo precios en pesos colombianos (COP).\n"
                f"{geo_regla}"
                f"\nResponde SOLO con JSON (pueden ser menos de {n} si no encuentras más):\n"
                f'{{"links":[{{"url":"https://...","precio":"$ ...","precio_numero":null}}]}}'
            )
            response = client.responses.create(
                model=_MODELO_BUSQUEDA,
                tools=[{"type": "web_search_preview"}],
                tool_choice={"type": "web_search_preview"},
                input=[{"role": "user", "content": prompt}],
            )
            data = _extraer_json(response.output_text)
            if data and isinstance(data, dict):
                for lk in (data.get("links") or [])[:n]:
                    _agregar(str(lk.get("url", "")).strip())
        except Exception as exc:
            logger.warning("  Reemplazo OpenAI error: %s", exc)

    # ── Fuente B0: SerpAPI Google Shopping (productos nuevos con precio) ───
    for lk in _buscar_productos_serpapi_shopping(descripcion, n=n, geo_nota=geo_serp):
        _agregar(lk.url)

    # ── Fuente B: SerpAPI orgánico ────────────────────────────────────────
    for u in _buscar_urls_serpapi(descripcion, n=n, excluir_dominios=dominios_ok, geo_nota=geo_serp):
        _agregar(u)

    # ── Fuente C: Brave Search (mejor cuota que CSE) ──────────────────────
    for u in _buscar_urls_brave(descripcion, n=n, excluir_dominios=dominios_ok, geo_nota=geo_serp):
        _agregar(u)

    # ── Fuente D: Google CSE (respaldo, si queda cuota) ───────────────────
    for u in _buscar_urls_google_cse(descripcion, n=n, excluir_dominios=dominios_ok, geo_nota=geo_serp):
        _agregar(u)

    logger.info(
        "  Reemplazo: %d URL(s) nuevas encontradas para '%s'",
        len(urls_finales), descripcion[:50],
    )
    return urls_finales[:n]


# ── PASO 2 — Orquestador principal ───────────────────────────────────────────

def tomar_screenshots(
    productos: list[ProductoLinks],
    dir_out: Path,
    api_key: str = "",
    departamento: str = "",
) -> tuple[list[ProductoLinks], dict[str, ResultadoScreenshot]]:
    """
    Para cada producto:
      Ronda 1+:
        a. Playwright visita todas las URLs pendientes → screenshot + precio DOM + base64.
        b. La visión recibe las capturas como IMÁGENES → valida producto + extrae precio.
        c. Los válidos se acumulan; si faltan, se piden URLs de reemplazo → nueva ronda.
        d. Se detiene al alcanzar el mínimo de capturas válidas o agotar las rondas.
      Máx _MAX_RONDAS rondas en total.

    Retorna (productos_actualizados, {url: ResultadoScreenshot}).
    """
    dir_out.mkdir(parents=True, exist_ok=True)

    res: dict[str, ResultadoScreenshot]  = {}
    productos_actualizados: list[ProductoLinks] = []

    try:
        for producto in productos:
            desc            = producto.descripcion or producto.item
            # Dedup canónica del pool inicial (misma ficha con/sin srsltid, utm…).
            urls_pendientes = []
            _claves_pend: set[str] = set()
            for lk in producto.links:
                if not lk.url:
                    continue
                u = _normalizar_url(lk.url)
                c = _canonizar_url(u)
                if c not in _claves_pend:
                    _claves_pend.add(c)
                    urls_pendientes.append(u)
            # Conjunto de claves ya intentadas (canónicas) y de capturas ya aceptadas.
            claves_intentadas: set[str] = set(_claves_pend)
            claves_validas:    set[str] = set()
            # Validez en dos niveles:
            #  - validos_stock: producto disponible (preferidos).
            #  - validos_agotado: válido pero agotado → solo como RESERVA si no hay 3 en stock.
            validos_stock:   list[LinkProducto] = []
            validos_agotado: list[LinkProducto] = []
            rondas_ok = 0

            # Precio previo conocido por URL (de SerpAPI Shopping: old_price/price).
            # Se usa SOLO como verificación cruzada del precio que lee la visión.
            prior_por_url: dict[str, int | None] = {
                _normalizar_url(lk.url): lk.precio_numero
                for lk in producto.links if lk.url
            }

            for ronda in range(1, _MAX_RONDAS + 1):
                if not urls_pendientes:
                    break

                logger.info(
                    "  [%s] Ronda %d/%d — visitando %d URL(s) en paralelo",
                    producto.item, ronda, _MAX_RONDAS, len(urls_pendientes),
                )

                # ── a. Playwright CONCURRENTE: screenshot + base64 (sin precio DOM)
                datos_ronda = _capturar_concurrente(urls_pendientes, dir_out)

                # ── b. Verificación visual por IMÁGENES (no PDF) ──────────
                capturas_img = [
                    (url, b64) for url, (_, b64, _, _) in datos_ronda.items() if b64
                ]
                if api_key and capturas_img:
                    verifs = _verificar_con_vision(desc, capturas_img, api_key)
                    verif_por_url = {v["url"]: v for v in verifs}
                else:
                    # Sin IA: aceptar todas las que tienen captura
                    verif_por_url = {
                        url: {"valido": True, "precio_texto": None, "precio_numero": None,
                              "en_stock": True, "motivo": ""}
                        for url, (ruta, _, _, _) in datos_ronda.items() if ruta
                    }

                # ── c. Clasificar válidos / inválidos ─────────────────────
                for url, (ruta, _b64, _p_txt, _p_num) in datos_ronda.items():
                    verif = verif_por_url.get(url)

                    if not verif or not verif.get("valido"):
                        res[url] = ResultadoScreenshot(None, None, None, False)
                        motivo = (verif or {}).get("motivo", "captura no disponible")
                        logger.warning("  [%s] ✗ inválido (%s): %s", producto.item, motivo, url)
                        continue

                    # PRECIO = lo que la visión LEE de la propia captura (fuente única
                    # de verdad). El precio del Excel siempre coincide con la imagen.
                    precio_num = verif.get("precio_numero")
                    precio_txt = verif.get("precio_texto")

                    if precio_num and not _es_precio_razonable(precio_num):
                        precio_num = None
                        precio_txt = None

                    # Verificar que el precio sea COP: si el texto tiene moneda extranjera, descartar
                    if precio_txt and any(s in precio_txt.upper() for s in ("USD", "US$", "EUR", "€", "MXN")):
                        logger.warning("  [%s] ✗ precio en moneda extranjera: %s — %s", producto.item, precio_txt, url)
                        res[url] = ResultadoScreenshot(None, None, None, False)
                        continue

                    if not ruta:
                        res[url] = ResultadoScreenshot(None, None, None, False)
                        logger.warning("  [%s] ✗ sin archivo de captura: %s", producto.item, url)
                        continue

                    # Sin precio visible → rechazar (no publicar precio que no se ve en
                    # la captura, p.ej. producto sin stock con metadata oculta).
                    if not precio_num:
                        res[url] = ResultadoScreenshot(None, None, None, False)
                        logger.warning("  [%s] ✗ sin precio visible en la captura: %s", producto.item, url)
                        continue

                    # Evitar capturas DUPLICADAS: la misma ficha (clave canónica) ya
                    # aceptada no se vuelve a incluir, aunque la URL difiera por tracking.
                    clave = _canonizar_url(url)
                    if clave in claves_validas:
                        res[url] = ResultadoScreenshot(None, None, None, False)
                        logger.info("  [%s] ↺ duplicada (misma ficha ya capturada): %s", producto.item, url)
                        continue

                    # Verificación cruzada con el precio estructurado de Shopping:
                    # si difieren mucho, es señal de que la visión pudo leer mal dígitos.
                    prior_num = prior_por_url.get(url)
                    if prior_num and prior_num > 0:
                        ratio = max(precio_num, prior_num) / min(precio_num, prior_num)
                        if ratio > 1.6:
                            logger.warning(
                                "  [%s] ⚠ precio visión %s difiere de Shopping %s (%.1f×): %s",
                                producto.item, precio_num, prior_num, ratio, url,
                            )

                    en_stock = verif.get("en_stock", True)
                    lk = LinkProducto(
                        url=url,
                        precio_texto=precio_txt or "N/A",
                        precio_numero=precio_num,
                    )
                    res[url] = ResultadoScreenshot(
                        ruta=ruta,
                        precio_texto=precio_txt,
                        precio_numero=precio_num,
                        sin_stock=not en_stock,
                    )
                    claves_validas.add(clave)
                    if en_stock:
                        validos_stock.append(lk)
                        logger.info(
                            "  [%s] ✓ en stock (%d/%d): %s → %s",
                            producto.item, len(validos_stock), _MIN_VALIDOS, url, precio_txt,
                        )
                    else:
                        validos_agotado.append(lk)
                        logger.info(
                            "  [%s] ○ agotado (reserva, %d): %s → %s",
                            producto.item, len(validos_agotado), url, precio_txt,
                        )

                rondas_ok += 1

                # ── d. ¿Tenemos suficientes EN STOCK? ─────────────────────
                needed = _MIN_VALIDOS - len(validos_stock)
                if needed <= 0:
                    logger.info("  [%s] %d/%d en stock — listo.", producto.item, len(validos_stock), _MIN_VALIDOS)
                    break

                if ronda >= _MAX_RONDAS:
                    logger.warning(
                        "  [%s] Límite de rondas con %d/%d en stock (+%d agotados de reserva).",
                        producto.item, len(validos_stock), _MIN_VALIDOS, len(validos_agotado),
                    )
                    break

                if not api_key:
                    break

                # Pedir generosamente: cada fallo de navegación/anti-bot/agotado cuesta
                # una URL, así que pedimos muchas más de las que necesitamos estrictamente.
                n_pedir = min(_URLS_POR_RONDA, needed * 6)
                logger.info(
                    "  [%s] Faltan %d en stock — buscando %d URL(s) de reemplazo...",
                    producto.item, needed, n_pedir,
                )
                nuevas_raw = _buscar_links_reemplazo(desc, n_pedir, claves_intentadas, api_key, departamento)
                nuevas = []
                for u in nuevas_raw:
                    un = _normalizar_url(u)
                    c = _canonizar_url(un)
                    if c not in claves_intentadas:
                        claves_intentadas.add(c)
                        nuevas.append(un)
                if not nuevas:
                    logger.warning("  [%s] Sin reemplazos disponibles.", producto.item)
                    break

                urls_pendientes = nuevas

            # ── Selección final: en stock primero, agotados solo para completar 3 ──
            final_links = list(validos_stock[:_MIN_VALIDOS])
            if len(final_links) < _MIN_VALIDOS and validos_agotado:
                faltan = _MIN_VALIDOS - len(final_links)
                relleno = validos_agotado[:faltan]
                final_links.extend(relleno)
                logger.info(
                    "  [%s] Completando con %d captura(s) de productos agotados (reserva).",
                    producto.item, len(relleno),
                )

            if len(final_links) < _MIN_VALIDOS:
                logger.warning(
                    "  [%s] ⚠ Final: solo %d/%d capturas tras %d ronda(s) — NO cumple el mínimo.",
                    producto.item, len(final_links), _MIN_VALIDOS, rondas_ok,
                )
            else:
                logger.info(
                    "  [%s] Final: %d/%d capturas (%d en stock) tras %d ronda(s).",
                    producto.item, len(final_links), _MIN_VALIDOS,
                    len(validos_stock), rondas_ok,
                )
            productos_actualizados.append(ProductoLinks(
                item=producto.item,
                descripcion=producto.descripcion,
                links=final_links,
            ))

    except Exception as exc:
        logger.error("Playwright error general: %s", exc)
        for producto in productos:
            if not any(p.item == producto.item for p in productos_actualizados):
                productos_actualizados.append(producto)

    return productos_actualizados, res


# ── Análisis de desviación de precio ─────────────────────────────────────────

_RE_UNIDADES = re.compile(
    r'[-–]\s*(\d+)\s*(?:unidades?|und\.?|uni\.?|uds?\.?|piezas?|pcs?\.?)',
    re.IGNORECASE,
)


def _extraer_unidades(nombre_item: str) -> int:
    """
    Extrae la cantidad de unidades del nombre del ítem.
    Ejemplos: "Hilos Macrame - 22 Unidades" → 22
              "Lámpara - 1 unidad"           → 1
              "Computador"                   → 1  (default)
    """
    m = _RE_UNIDADES.search(nombre_item or "")
    if m:
        try:
            return max(1, int(m.group(1)))
        except ValueError:
            pass
    return 1


def _mediana(valores: list[int]) -> int:
    """Mediana de enteros: más robusta que el promedio ante precios atípicos."""
    s = sorted(valores)
    n = len(s)
    mid = n // 2
    return s[mid] if n % 2 == 1 else (s[mid - 1] + s[mid]) // 2


class AlertaDesviacion:
    __slots__ = ("mediana_unitaria_internet", "mediana_total_internet",
                 "precio_cotizacion", "unidades",
                 "desviacion_pct", "limite_pct", "supera_limite", "mensaje")

    def __init__(self, mediana_unitaria_internet, mediana_total_internet,
                 precio_cotizacion, unidades,
                 desviacion_pct, limite_pct, supera_limite, mensaje):
        self.mediana_unitaria_internet = mediana_unitaria_internet
        self.mediana_total_internet    = mediana_total_internet
        self.precio_cotizacion         = precio_cotizacion
        self.unidades                  = unidades
        self.desviacion_pct            = desviacion_pct
        self.limite_pct                = limite_pct
        self.supera_limite             = supera_limite
        self.mensaje                   = mensaje


def _analizar_desviacion(
    precio_cotizacion: int,
    precios_unitarios_internet: list[int],
    nombre_item: str = "",
) -> "AlertaDesviacion | None":
    """
    Compara el total cotizado con la mediana de internet ajustada por unidades.

    Se usa la mediana (no el promedio) para evitar que precios atípicos
    distorsionen la referencia de mercado.

    Lógica:
      - precio_cotizacion  = valor_total de la cotización (incluye N unidades)
      - precios_internet   = precios UNITARIOS encontrados en internet
      - unidades           = parseado del nombre del ítem ("22 Unidades")
      - mediana_total      = mediana_unitaria × unidades

    Reglas de tolerancia (sobre precio_cotizacion / unidades, es decir, por unidad):
      - precio_cotizacion/unidad < 500.000 COP → tolerancia 50 %
      - precio_cotizacion/unidad ≥ 500.000 COP → tolerancia 20 %

    Alerta cuando el precio cotizado supera la mediana de internet ajustada
    en más del % de tolerancia.
    """
    validos = [p for p in precios_unitarios_internet if p and p > 0]
    if not validos or not precio_cotizacion:
        return None

    unidades       = _extraer_unidades(nombre_item)
    mediana_unit   = _mediana(validos)
    mediana_total  = mediana_unit * unidades

    precio_unit_cotiz = precio_cotizacion / unidades
    limite            = 50.0 if precio_unit_cotiz < 500_000 else 20.0

    desviacion = (precio_cotizacion - mediana_total) / mediana_total * 100
    supera     = desviacion > limite

    mtotal_fmt = _formatear_cop(mediana_total)
    munit_fmt  = _formatear_cop(mediana_unit)
    pcotiz_fmt = _formatear_cop(precio_cotizacion)

    unid_str = f"{unidades} und." if unidades > 1 else "1 und."

    if supera:
        exceso  = desviacion - limite
        mensaje = (
            f"⚠ ALERTA: la cotización ({pcotiz_fmt} / {unid_str}) supera la mediana "
            f"de internet en {desviacion:.1f}% — "
            f"mediana internet: {munit_fmt}/und. × {unidades} = {mtotal_fmt} — "
            f"límite: {limite:.0f}%, exceso: {exceso:.1f}%"
        )
    else:
        mensaje = (
            f"✓ OK: cotización ({pcotiz_fmt} / {unid_str}) dentro del rango permitido — "
            f"mediana internet: {munit_fmt}/und. × {unidades} = {mtotal_fmt} — "
            f"desviación: {desviacion:+.1f}%, límite: {limite:.0f}%"
        )

    return AlertaDesviacion(
        mediana_unitaria_internet=mediana_unit,
        mediana_total_internet=mediana_total,
        precio_cotizacion=precio_cotizacion,
        unidades=unidades,
        desviacion_pct=desviacion,
        limite_pct=limite,
        supera_limite=supera,
        mensaje=mensaje,
    )


# ── PASO 3 — Excel ────────────────────────────────────────────────────────────

def generar_excel_cotizaciones(
    id_unico: str,
    productos: list[ProductoLinks],
    screenshots: dict[str, ResultadoScreenshot],
    ruta_salida: Path,
    seleccionadas: list | None = None,
) -> list[str]:
    """
    Genera el Excel de referencia de precios web (una hoja por producto + índice).
    Retorna lista de mensajes de alerta de desviación de precio disparados
    (uno por producto que supere el límite), para incluir en el checklist.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.drawing.image import Image as XLImage

    precio_cotizacion_por_item: dict[str, int] = {}
    descripcion_original_por_item: dict[str, str] = {}
    if seleccionadas:
        for sel in seleccionadas:
            if sel.valor_total:
                precio_cotizacion_por_item[sel.item] = sel.valor_total
            if sel.descripcion:
                descripcion_original_por_item[sel.item] = sel.descripcion

    alertas_desviacion: list[str] = []

    hoy = date.today().strftime("%d/%m/%Y")

    def _sheet_name(name: str) -> str:
        return _RE_HOJA_INVALIDOS.sub("_", name)[:31]

    _C_DARK  = "1F5C2E"
    _C_MED   = "2E7D32"
    _C_ODD   = "E8F5E9"
    _C_EVEN  = "F5F5F5"
    _C_WHITE = "FFFFFF"

    def _fill(hex6):
        return PatternFill("solid", start_color=hex6, end_color=hex6)

    def _font(bold=False, size=11, color="000000", name="Arial"):
        return Font(bold=bold, size=size, color=color, name=name)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Índice ───────────────────────────────────────────────────────────
    ws_i = wb.create_sheet("Índice")
    ws_i.merge_cells("A1:E1")
    ws_i["A1"].value     = f"Cotizaciones de referencia web — {id_unico}"
    ws_i["A1"].font      = _font(bold=True, size=14, color=_C_WHITE)
    ws_i["A1"].fill      = _fill(_C_DARK)
    ws_i["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_i.row_dimensions[1].height = 32

    ws_i["A2"].value = f"Generado el  {hoy}"
    ws_i["A2"].font  = _font(size=10, color="555555")
    ws_i.row_dimensions[2].height = 16

    for col, hdr in enumerate(["Producto", "Descripción", "Hoja"], 1):
        c = ws_i.cell(row=4, column=col, value=hdr)
        c.font  = _font(bold=True, color=_C_WHITE)
        c.fill  = _fill(_C_MED)
        c.alignment = Alignment(horizontal="center")
    ws_i.row_dimensions[4].height = 18

    for r, prod in enumerate(productos, 5):
        desc_display = descripcion_original_por_item.get(prod.item) or prod.descripcion or ""
        ws_i.cell(row=r, column=1, value=prod.item)
        ws_i.cell(row=r, column=2, value=desc_display)
        ws_i.cell(row=r, column=3, value=_sheet_name(prod.item))
    ws_i.column_dimensions["A"].width = 35
    ws_i.column_dimensions["B"].width = 60
    ws_i.column_dimensions["C"].width = 35

    # ── Una hoja por producto ────────────────────────────────────────────
    IMG_ROWS = 24
    IMG_H_PX = 460
    IMG_W_PX = 900

    for prod in productos:
        ws = wb.create_sheet(title=_sheet_name(prod.item))

        ws.merge_cells("A1:E1")
        ws["A1"].value     = f"COTIZACIONES DE REFERENCIA — {prod.item.upper()}"
        ws["A1"].font      = _font(bold=True, size=13, color=_C_WHITE)
        ws["A1"].fill      = _fill(_C_DARK)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        ws.merge_cells("A2:C2")
        _desc_orig = descripcion_original_por_item.get(prod.item) or prod.descripcion or "N/A"
        ws["A2"].value = f"Descripción: {_desc_orig}"
        ws["A2"].font  = _font(size=10)
        ws.merge_cells("D2:E2")
        ws["D2"].value     = f"Fecha: {hoy}  |  ID: {id_unico}"
        ws["D2"].font      = _font(size=10, color="555555")
        ws["D2"].alignment = Alignment(horizontal="right")
        ws.row_dimensions[2].height = 16

        for col, hdr in enumerate(["#", "URL", "Precio actual", "Captura de pantalla", "Fuente precio"], 1):
            c = ws.cell(row=3, column=col, value=hdr)
            c.font      = _font(bold=True, color=_C_WHITE)
            c.fill      = _fill(_C_MED)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[3].height = 20

        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 70
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 120
        ws.column_dimensions["E"].width = 22

        current_row = 4
        for i, link in enumerate(prod.links, 1):
            bg = _C_ODD if i % 2 == 1 else _C_EVEN

            lbl           = ws.cell(row=current_row, column=1, value=f"Op. {i}")
            lbl.font      = _font(bold=True)
            lbl.fill      = _fill(bg)
            lbl.alignment = Alignment(horizontal="center", vertical="center")

            url_cell           = ws.cell(row=current_row, column=2, value=link.url)
            url_cell.hyperlink = link.url
            url_cell.font      = Font(name="Arial", size=10, color="0563C1", underline="single")
            url_cell.fill      = _fill(bg)

            ss = screenshots.get(link.url, ResultadoScreenshot(None, None, None, False))
            # El precio SOLO se acepta si la visión lo leyó de la captura real
            # (ss.precio_numero). Nunca se usa el precio "adivinado" por el web search.
            if ss.precio_numero:
                precio_mostrar = ss.precio_texto or _formatear_cop(ss.precio_numero)
                precio_fuente  = "precio leído de la página"
                if ss.sin_stock:
                    # Captura de reserva: producto agotado (se usó por no haber 3 en stock)
                    precio_fuente = "precio de referencia — PRODUCTO AGOTADO"
            else:
                precio_mostrar = "No disponible"
                precio_fuente  = "no se pudo leer de la página"

            precio_cell           = ws.cell(row=current_row, column=3, value=precio_mostrar)
            precio_cell.font      = _font(bold=True, size=11)
            precio_cell.fill      = _fill(bg)
            precio_cell.alignment = Alignment(horizontal="center", vertical="center")

            fuente_cell       = ws.cell(row=current_row, column=5, value=precio_fuente)
            fuente_cell.font  = _font(size=8, color="777777")
            fuente_cell.fill  = _fill(bg)
            fuente_cell.alignment = Alignment(horizontal="center", vertical="center")

            ws.row_dimensions[current_row].height = 22
            current_row += 1

            img_row = current_row
            for r in range(IMG_ROWS):
                ws.row_dimensions[current_row + r].height = 20

            img_path = ss.ruta
            if img_path and img_path.exists():
                try:
                    xl_img        = XLImage(str(img_path))
                    xl_img.width  = IMG_W_PX
                    xl_img.height = IMG_H_PX
                    ws.add_image(xl_img, f"D{img_row}")
                except Exception as exc:
                    logger.warning("No se pudo insertar imagen en Excel: %s", exc)
                    ws.cell(row=img_row, column=4, value="[Captura no disponible]").font = _font(color="999999")
            else:
                ws.cell(row=img_row, column=4, value="[Captura no disponible]").font = _font(color="999999")

            current_row += IMG_ROWS + 1

        # ── Análisis de desviación ────────────────────────────────────────
        # Los precios de internet son UNITARIOS; el valor_total de la cotización
        # incluye N unidades (parseadas del nombre del ítem).
        # SOLO se usan precios que la visión leyó de la captura real, nunca el
        # precio adivinado por el web search.
        precios_unitarios_internet = [
            (screenshots.get(lk.url) or ResultadoScreenshot(None, None, None, False)).precio_numero
            for lk in prod.links
        ]
        precio_cotiz = precio_cotizacion_por_item.get(prod.item)
        unidades     = _extraer_unidades(prod.item)

        current_row += 1

        ws.merge_cells(f"A{current_row}:E{current_row}")
        h = ws.cell(row=current_row, column=1, value="ANÁLISIS DE PRECIO")
        h.font      = _font(bold=True, size=11, color=_C_WHITE)
        h.fill      = _fill(_C_DARK)
        h.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        # Fila: total cotizado
        ws.cell(row=current_row, column=1, value="Total cotización elegida:").font = _font(bold=True, size=10)
        ws.cell(row=current_row, column=2,
                value=_formatear_cop(precio_cotiz) if precio_cotiz else "No disponible"
                ).font = _font(bold=True, size=10)
        unid_label = f"({unidades} unidad{'es' if unidades != 1 else ''})"
        ws.cell(row=current_row, column=3, value=unid_label).font = _font(size=9, color="555555")
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        # Filas: precios unitarios de internet + total ajustado
        for idx, (lk, p_unit) in enumerate(zip(prod.links, precios_unitarios_internet), 1):
            ws.cell(row=current_row, column=1,
                    value=f"  Precio unitario internet #{idx}:").font = _font(size=10, color="444444")
            if p_unit:
                ws.cell(row=current_row, column=2,
                        value=_formatear_cop(p_unit)).font = _font(size=10)
                if unidades > 1:
                    ws.cell(row=current_row, column=3,
                            value=f"× {unidades} = {_formatear_cop(p_unit * unidades)}"
                            ).font = _font(size=9, color="555555")
            else:
                ws.cell(row=current_row, column=2, value="No extraído").font = _font(size=10, color="999999")
            ws.row_dimensions[current_row].height = 16
            current_row += 1

        # Fila: mediana unitaria + mediana total ajustada
        precios_validos = [p for p in precios_unitarios_internet if p]
        ws.cell(row=current_row, column=1,
                value="Mediana unitaria internet:").font = _font(bold=True, size=10)
        if precios_validos:
            med_unit  = _mediana(precios_validos)
            med_total = med_unit * unidades
            ws.cell(row=current_row, column=2,
                    value=_formatear_cop(med_unit)).font = _font(bold=True, size=10)
            if unidades > 1:
                ws.cell(row=current_row, column=3,
                        value=f"× {unidades} = {_formatear_cop(med_total)} (total)"
                        ).font = _font(size=9, color="555555")
        else:
            ws.cell(row=current_row, column=2, value="Sin datos").font = _font(size=10, color="999999")
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        # Fila: resultado del análisis
        alerta = (
            _analizar_desviacion(precio_cotiz, precios_unitarios_internet, prod.item)
            if precio_cotiz else None
        )
        ws.merge_cells(f"A{current_row}:E{current_row}")
        if alerta:
            color_fondo = "FFCCCC" if alerta.supera_limite else "CCFFCC"
            color_texto = "CC0000" if alerta.supera_limite else "1A5C1A"
            rc = ws.cell(row=current_row, column=1, value=alerta.mensaje)
            rc.font      = Font(name="Arial", size=10, bold=True, color=color_texto)
            rc.fill      = _fill(color_fondo)
            rc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.row_dimensions[current_row].height = 36
            if alerta.supera_limite:
                alertas_desviacion.append(
                    f"04_VALIDACION_PRECIO_MERCADO [{prod.item}]: {alerta.mensaje}"
                )
        else:
            nc = ws.cell(row=current_row, column=1,
                         value="Sin datos suficientes para el análisis de desviación.")
            nc.font = _font(size=10, color="999999")
            ws.row_dimensions[current_row].height = 18

    ruta_salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(ruta_salida))
    logger.info("Excel cotizaciones web guardado: %s", ruta_salida)
    return alertas_desviacion
