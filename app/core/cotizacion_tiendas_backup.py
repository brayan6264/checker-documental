"""
Búsqueda web de precios de referencia para cotizaciones seleccionadas.

Flujo:
  1. buscar_links_tiendas()  — busca directamente en alkosto, homecenter, exito, etc.
                               usando Playwright (sin IA, sin costo).
  2. tomar_screenshots()     — visita cada URL, verifica stock, guarda captura.
  3. generar_excel_cotizaciones() — genera un Excel con una hoja por producto.
"""

import hashlib
import logging
import re
from datetime import date
from pathlib import Path
from typing import NamedTuple
from urllib.parse import quote_plus, urlparse

# Caracteres inválidos en nombres de hoja Excel: \ / ? * [ ] :
_RE_HOJA_INVALIDOS = re.compile(r'[\\/?*\[\]:]')

logger = logging.getLogger(__name__)


# ── Tipos ─────────────────────────────────────────────────────────────────────

class LinkProducto(NamedTuple):
    url:           str
    precio_texto:  str
    precio_numero: int | None


class ResultadoScreenshot(NamedTuple):
    ruta:          Path | None
    precio_texto:  str | None
    precio_numero: int | None
    sin_stock:     bool = False


class ProductoLinks(NamedTuple):
    item:        str
    descripcion: str | None
    links:       list[LinkProducto]   # hasta 5 candidatos


# Regex para precios colombianos: $ 89.900 / 1.234.567 / COP 89.900
_RE_PRECIO_COP = re.compile(
    r'(?:COP\s*|cop\s*|\$\s*)?(\d{1,3}(?:[.,]\d{3})+)(?:\s*(?:COP|cop))?'
)

# ── Tiendas colombianas ───────────────────────────────────────────────────────

# Tiendas con API JSON directa (sin Playwright — rápido y confiable)
# tipo "vtex": usa endpoint /api/catalog_system/pub/products/search
_TIENDAS_API: list[dict] = [
    {
        "nombre":  "Éxito",
        "tipo":    "vtex",
        "api_url": "https://www.exito.com/api/catalog_system/pub/products/search",
    },
]

# Tiendas con scraping via Playwright
_TIENDAS_PLAYWRIGHT: list[dict] = [
    {
        "nombre":      "Alkosto",
        "dominio":     "alkosto.com",
        "url":         "https://www.alkosto.com/search?text={q}",
        "espera":      4000,
        "patron_url":  "/p/",    # URLs de producto: /nombre-producto/p/SKU
    },
    # Homecenter y Falabella muestran contenido patrocinado antes que los resultados
    # de búsqueda; se omiten para evitar links irrelevantes.
]


# JS que extrae el primer link de producto de la página.
# Recibe {dominio, patronUrl} para filtrar solo URLs que parezcan fichas de producto.
_JS_PRIMER_PRODUCTO = """
(cfg) => {
    const {dominio, patronUrl} = cfg;

    const isProduct = href => {
        if (!href || !href.includes(dominio)) return false;
        if (!href.includes(patronUrl)) return false;
        try {
            const path = new URL(href).pathname;
            // Rechazar páginas de cuenta, búsqueda, etc.
            if (/account|login|register|cart|carrito|checkout|search|buscar/i.test(path)) return false;
            return path.length > 10;
        } catch(e) { return false; }
    };

    const visto = new Set();
    for (const a of document.querySelectorAll('a[href]')) {
        const href = a.href;
        if (!isProduct(href) || visto.has(href)) continue;
        if (a.closest('nav, footer, header, [class*="breadcrumb"], [class*="menu"]')) continue;
        visto.add(href);

        // Buscar precio en la tarjeta que contiene este link
        const card = a.closest(
            '[class*="product"], [class*="Product"], article, li[class], [class*="card"], [class*="result"]'
        );
        let precio = '';
        if (card) {
            const pSels = [
                '[class*="price"]:not([class*="original"]):not([class*="old"]):not([class*="before"])',
                '[class*="precio"]:not([class*="antes"])',
                '[itemprop="price"]',
            ];
            for (const s of pSels) {
                const el = card.querySelector(s);
                if (el && el.textContent.trim()) { precio = el.textContent.trim(); break; }
            }
        }
        return {url: href, precio};
    }
    return null;
}
"""


# ── Búsqueda directa en tiendas ───────────────────────────────────────────────

def buscar_links_tiendas(
    seleccionadas: list,
    _api_key: str = "",   # ignorado — mantenido para compatibilidad con la llamada en procesador.py
) -> tuple[bool, list[ProductoLinks], str]:
    """
    Busca el primer resultado disponible en tiendas colombianas:
    - Tiendas con API JSON (ej. Éxito/VTEX): consulta directa, sin browser.
    - Tiendas con Playwright: scraping de página de resultados.
    Gratis, sin IA.

    Retorna (ok, productos, resumen).
    """
    import requests as _req

    try:
        from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
    except ImportError:
        return False, [], "FALTA — playwright no instalado (pip install playwright && playwright install chromium)"

    _HEADERS_REQ = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124",
        "Accept": "application/json",
    }

    def _buscar_vtex(api_url: str, nombre: str, q: str) -> LinkProducto | None:
        """Consulta el API VTEX de búsqueda de una tienda y retorna el primer resultado."""
        try:
            r = _req.get(
                api_url,
                params={"ft": q, "_from": 0, "_to": 0},
                headers=_HEADERS_REQ,
                timeout=10,
            )
            if not r.ok:
                return None
            data = r.json()
            if not data:
                return None
            prod = data[0]
            url = prod.get("link", "")
            if not url:
                return None
            # Precio del primer seller
            precio_num: int | None = None
            precio_txt = "N/A"
            try:
                precio_raw = (
                    prod.get("items", [{}])[0]
                    .get("sellers", [{}])[0]
                    .get("commertialOffer", {})
                    .get("Price")
                )
                if precio_raw:
                    precio_num = int(precio_raw)
                    if not _es_precio_razonable(precio_num):
                        precio_num = None
                    else:
                        precio_txt = _formatear_cop(precio_num)
            except Exception:
                pass
            logger.info("  API %s → %s (%s)", nombre, url[:80], precio_txt)
            return LinkProducto(url=url, precio_texto=precio_txt, precio_numero=precio_num)
        except Exception as exc:
            logger.warning("  API %s error: %s", nombre, exc)
            return None

    productos_out: list[ProductoLinks] = []
    errores: list[str] = []

    try:
        with sync_playwright() as pw:
            browser = pw.chromium.launch(headless=True)
            ctx = browser.new_context(
                viewport={"width": 1280, "height": 900},
                locale="es-CO",
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0.0.0 Safari/537.36"
                ),
            )
            page = ctx.new_page()
            page.route(
                "**/*",
                lambda route: route.abort()
                if route.request.resource_type in ("font", "media", "image")
                else route.continue_(),
            )

            for sel in seleccionadas:
                desc = sel.descripcion or sel.item
                q = quote_plus(desc)
                links_encontrados: list[LinkProducto] = []
                dominios_vistos: set[str] = set()

                def _agregar(lk: LinkProducto) -> bool:
                    dominio = urlparse(lk.url).netloc.lower().removeprefix("www.")
                    if dominio in dominios_vistos:
                        return False
                    dominios_vistos.add(dominio)
                    links_encontrados.append(lk)
                    return True

                # ── 1. Tiendas con API JSON ───────────────────────────────────
                for t in _TIENDAS_API:
                    if len(links_encontrados) >= 5:
                        break
                    if t["tipo"] == "vtex":
                        lk = _buscar_vtex(t["api_url"], t["nombre"], desc)
                        if lk:
                            _agregar(lk)

                # ── 2. Tiendas con Playwright ─────────────────────────────────
                for tienda in _TIENDAS_PLAYWRIGHT:
                    if len(links_encontrados) >= 5:
                        break
                    url_busqueda = tienda["url"].format(q=q)
                    try:
                        page.goto(url_busqueda, timeout=25_000, wait_until="domcontentloaded")
                        page.wait_for_timeout(tienda["espera"])

                        resultado = page.evaluate(
                            _JS_PRIMER_PRODUCTO,
                            {"dominio": tienda["dominio"], "patronUrl": tienda["patron_url"]},
                        )
                        if not resultado:
                            logger.info("  [%s] %s — sin resultados", sel.item, tienda["nombre"])
                            continue

                        url = resultado.get("url", "").strip()
                        if not url:
                            continue

                        precio_txt = resultado.get("precio", "").strip()
                        precio_num = _parsear_precio_cop(precio_txt) if precio_txt else None
                        if precio_num and not _es_precio_razonable(precio_num):
                            precio_num = None
                            precio_txt = ""

                        _agregar(LinkProducto(
                            url=url,
                            precio_texto=precio_txt or "N/A",
                            precio_numero=precio_num,
                        ))
                        logger.info("  [%s] %s → %s (%s)", sel.item, tienda["nombre"], url[:80], precio_txt or "sin precio")

                    except PWTimeout:
                        logger.warning("  [%s] Timeout en %s", sel.item, tienda["nombre"])
                    except Exception as exc:
                        logger.warning("  [%s] Error en %s: %s", sel.item, tienda["nombre"], exc)

                if links_encontrados:
                    productos_out.append(ProductoLinks(
                        item=sel.item, descripcion=desc, links=links_encontrados
                    ))
                else:
                    logger.warning("  Producto '%s': sin enlaces en tiendas", sel.item)
                    errores.append(f"Sin enlaces para '{sel.item}'")

            browser.close()

    except Exception as exc:
        logger.error("Playwright error en búsqueda de tiendas: %s", exc)
        return False, [], f"Error Playwright: {exc}"

    if not productos_out:
        return False, [], "No se encontraron enlaces en ninguna tienda"

    total = sum(len(p.links) for p in productos_out)
    resumen = f"OK — {len(productos_out)} producto(s), {total} candidato(s) de tiendas directas"
    if errores:
        resumen += f" | sin resultados: {'; '.join(errores)}"
    return True, productos_out, resumen


# ── Playwright — verificación stock + screenshots ─────────────────────────────

_PALABRAS_SIN_STOCK = [
    "sin stock", "agotado", "no disponible", "producto no disponible",
    "out of stock", "notify me", "avísame", "avisame",
    "producto no encontrado", "no se encontró", "estamos preparando",
    "no está disponible", "no esta disponible",
]

# Límites razonables para precios COP de productos cotizados
_PRECIO_MIN_COP = 100
_PRECIO_MAX_COP = 50_000_000


def tomar_screenshots(
    productos: list[ProductoLinks],
    dir_out: Path,
) -> dict[str, ResultadoScreenshot]:
    """
    Para cada producto:
      1. Visita los candidatos en orden.
      2. Verifica stock (descarta páginas sin stock).
      3. Extrae el precio real de la página.
      4. Toma screenshot.
      5. Guarda hasta 3 resultados válidos por producto.

    Retorna {url: ResultadoScreenshot}.
    """
    from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

    dir_out.mkdir(parents=True, exist_ok=True)

    candidatos_por_producto: list[tuple[str, list[str]]] = [
        (p.item, [lk.url for lk in p.links if lk.url])
        for p in productos
    ]

    res: dict[str, ResultadoScreenshot] = {}
    _vacio = ResultadoScreenshot(ruta=None, precio_texto=None, precio_numero=None, sin_stock=False)

    try:
        with sync_playwright() as pw:
            browser = pw.chromium.launch(headless=True)
            ctx = browser.new_context(
                viewport={"width": 1280, "height": 900},
                locale="es-CO",
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0.0.0 Safari/537.36"
                ),
            )
            page = ctx.new_page()
            page.route(
                "**/*",
                lambda route: route.abort()
                if route.request.resource_type in ("font", "media")
                else route.continue_(),
            )

            for item, urls in candidatos_por_producto:
                validos = 0
                for url in urls:
                    if validos >= 3:
                        break
                    fname = hashlib.md5(url.encode()).hexdigest()[:12] + ".png"
                    ruta  = dir_out / fname
                    try:
                        page.goto(url, timeout=30_000, wait_until="domcontentloaded")
                        page.wait_for_timeout(2500)

                        # ── Verificar stock ──────────────────────────────────
                        texto_pagina = (page.evaluate("document.body.innerText") or "").lower()
                        sin_stock = any(p in texto_pagina for p in _PALABRAS_SIN_STOCK)
                        if sin_stock:
                            logger.warning("  Sin stock: %s", url)
                            res[url] = ResultadoScreenshot(ruta=None, precio_texto=None, precio_numero=None, sin_stock=True)
                            continue

                        # ── Captura y precio ─────────────────────────────────
                        page.screenshot(path=str(ruta), full_page=False)
                        precio_txt, precio_num = _extraer_precio_pagina(page)
                        res[url] = ResultadoScreenshot(
                            ruta=ruta,
                            precio_texto=precio_txt,
                            precio_numero=precio_num,
                            sin_stock=False,
                        )
                        validos += 1
                        logger.info("  [%s] OK (%d/3): %s → %s", item, validos, url, precio_txt or "precio no extraído")

                    except PWTimeout:
                        logger.warning("  Timeout: %s", url)
                        res[url] = _vacio
                    except Exception as exc:
                        logger.warning("  Error %s: %s", url, exc)
                        res[url] = _vacio

                if validos < 3:
                    logger.warning("  [%s] Solo %d/3 links con stock disponibles", item, validos)

            browser.close()
    except Exception as exc:
        logger.error("Playwright error general: %s", exc)
        all_urls = [u for _, urls in candidatos_por_producto for u in urls]
        for url in all_urls:
            res.setdefault(url, _vacio)

    return res


# ── helpers internos de precio ────────────────────────────────────────────────

def _es_precio_razonable(num: int) -> bool:
    return _PRECIO_MIN_COP <= num <= _PRECIO_MAX_COP


def _extraer_precio_pagina(page) -> tuple[str | None, int | None]:
    """
    Extrae el precio más prominente y razonable de la página abierta.
    Estrategia: selectores semánticos → escaneo de texto completo.
    """
    selectores = [
        "[class*='price']:not([class*='original']):not([class*='old']):not([class*='before']):not([class*='tachado'])",
        "[class*='precio']:not([class*='antes']):not([class*='tachado'])",
        "[itemprop='price']",
        "[data-price]",
        ".product-price", ".pdp-price", ".price-box",
        ".vtex-product-price", ".ProductPrice",
    ]
    for sel in selectores:
        try:
            el = page.locator(sel).first
            if el.count() and el.is_visible(timeout=500):
                texto = (el.text_content(timeout=500) or "").strip()
                num = _parsear_precio_cop(texto)
                if num and _es_precio_razonable(num):
                    return _formatear_cop(num), num
        except Exception:
            continue

    # Escaneo de texto completo
    try:
        texto_pagina = page.evaluate("document.body.innerText") or ""
        candidatos = []
        for m in _RE_PRECIO_COP.finditer(texto_pagina):
            num = _parsear_precio_cop(m.group(0))
            if num and _es_precio_razonable(num):
                candidatos.append(num)
        if candidatos:
            from collections import Counter
            mas_comun = Counter(candidatos).most_common(1)[0][0]
            return _formatear_cop(mas_comun), mas_comun
    except Exception:
        pass

    return None, None


def _parsear_precio_cop(texto: str) -> int | None:
    """Convierte '$ 89.900' o '89,900' o '89.900' a entero 89900."""
    limpio = re.sub(r'[^\d.,]', '', texto.strip())
    if re.fullmatch(r'\d{1,3}(\.\d{3})+', limpio):
        return int(limpio.replace('.', ''))
    if re.fullmatch(r'\d{1,3}(,\d{3})+', limpio):
        return int(limpio.replace(',', ''))
    if re.fullmatch(r'\d+', limpio) and len(limpio) >= 4:
        return int(limpio)
    return None


def _formatear_cop(valor: int) -> str:
    """Convierte 89900 a '$ 89.900'."""
    s = f"{valor:,}".replace(",", ".")
    return f"$ {s}"


# ── Excel ────────────────────────────────────────────────────────────────────

def generar_excel_cotizaciones(
    id_unico: str,
    productos: list[ProductoLinks],
    screenshots: dict[str, ResultadoScreenshot],
    ruta_salida: Path,
) -> None:
    """Genera el Excel de referencia de precios web (una hoja por producto + índice)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.drawing.image import Image as XLImage

    hoy = date.today().strftime("%d/%m/%Y")

    def _sheet_name(name: str) -> str:
        return _RE_HOJA_INVALIDOS.sub("_", name)[:31]

    _C_DARK   = "1F5C2E"
    _C_MED    = "2E7D32"
    _C_ODD    = "E8F5E9"
    _C_EVEN   = "F5F5F5"
    _C_WHITE  = "FFFFFF"

    def _fill(hex6: str) -> PatternFill:
        return PatternFill("solid", start_color=hex6, end_color=hex6)

    def _font(bold=False, size=11, color="000000", name="Arial") -> Font:
        return Font(bold=bold, size=size, color=color, name=name)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Hoja índice ──────────────────────────────────────────────────────
    ws_i = wb.create_sheet("Índice")
    ws_i.merge_cells("A1:E1")
    ws_i["A1"].value     = f"Cotizaciones de referencia web — {id_unico}"
    ws_i["A1"].font      = _font(bold=True, size=14, color=_C_WHITE)
    ws_i["A1"].fill      = _fill(_C_DARK)
    ws_i["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_i.row_dimensions[1].height = 32

    ws_i["A2"].value = f"Generado: {hoy}"
    ws_i["A2"].font  = _font(size=10, color="555555")
    ws_i.row_dimensions[2].height = 16

    for col, hdr in enumerate(["Producto", "Descripción", "Hoja"], 1):
        c = ws_i.cell(row=4, column=col, value=hdr)
        c.font  = _font(bold=True, color=_C_WHITE)
        c.fill  = _fill(_C_MED)
        c.alignment = Alignment(horizontal="center")
    ws_i.row_dimensions[4].height = 18

    for r, prod in enumerate(productos, 5):
        ws_i.cell(row=r, column=1, value=prod.item)
        ws_i.cell(row=r, column=2, value=prod.descripcion or "")
        ws_i.cell(row=r, column=3, value=_sheet_name(prod.item))
    ws_i.column_dimensions["A"].width = 35
    ws_i.column_dimensions["B"].width = 60
    ws_i.column_dimensions["C"].width = 35

    # ── Una hoja por producto ─────────────────────────────────────────────
    IMG_ROWS  = 24
    IMG_H_PX  = 460
    IMG_W_PX  = 900

    for prod in productos:
        ws = wb.create_sheet(title=_sheet_name(prod.item))

        ws.merge_cells("A1:E1")
        ws["A1"].value     = f"COTIZACIONES DE REFERENCIA — {prod.item.upper()}"
        ws["A1"].font      = _font(bold=True, size=13, color=_C_WHITE)
        ws["A1"].fill      = _fill(_C_DARK)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        ws.merge_cells("A2:C2")
        ws["A2"].value = f"Descripción: {prod.descripcion or 'N/A'}"
        ws["A2"].font  = _font(size=10)
        ws.merge_cells("D2:E2")
        ws["D2"].value     = f"Fecha consulta: {hoy}  |  ID: {id_unico}"
        ws["D2"].font      = _font(size=10, color="555555")
        ws["D2"].alignment = Alignment(horizontal="right")
        ws.row_dimensions[2].height = 16

        for col, hdr in enumerate(["#", "URL", "Precio actual", "Captura de pantalla", "Fuente precio"], 1):
            c = ws.cell(row=3, column=col, value=hdr)
            c.font      = _font(bold=True, color=_C_WHITE)
            c.fill      = _fill(_C_MED)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[3].height = 20

        ws.column_dimensions["A"].width  = 8
        ws.column_dimensions["B"].width  = 70
        ws.column_dimensions["C"].width  = 22
        ws.column_dimensions["D"].width  = 120
        ws.column_dimensions["E"].width  = 22

        current_row = 4
        for i, link in enumerate(prod.links, 1):
            bg = _C_ODD if i % 2 == 1 else _C_EVEN

            lbl = ws.cell(row=current_row, column=1, value=f"Op. {i}")
            lbl.font  = _font(bold=True)
            lbl.fill  = _fill(bg)
            lbl.alignment = Alignment(horizontal="center", vertical="center")

            url_cell           = ws.cell(row=current_row, column=2, value=link.url)
            url_cell.hyperlink = link.url
            url_cell.font      = Font(name="Arial", size=10, color="0563C1", underline="single")
            url_cell.fill      = _fill(bg)

            ss = screenshots.get(link.url, ResultadoScreenshot(None, None, None, False))
            if ss.sin_stock:
                precio_mostrar = "SIN STOCK"
                precio_fuente  = "verificado en página"
            elif ss.precio_texto:
                precio_mostrar = ss.precio_texto
                precio_fuente  = "precio actual página"
            else:
                precio_mostrar = link.precio_texto or "No disponible"
                precio_fuente  = "precio búsqueda"

            precio_cell           = ws.cell(row=current_row, column=3, value=precio_mostrar)
            precio_cell.font      = _font(bold=True, size=11)
            precio_cell.fill      = _fill(bg)
            precio_cell.alignment = Alignment(horizontal="center", vertical="center")

            fuente_cell       = ws.cell(row=current_row, column=5, value=precio_fuente)
            fuente_cell.font  = _font(size=8, color="777777")
            fuente_cell.fill  = _fill(bg)
            fuente_cell.alignment = Alignment(horizontal="center", vertical="center")

            ws.row_dimensions[current_row].height = 22
            current_row += 1

            img_row = current_row
            for r in range(IMG_ROWS):
                ws.row_dimensions[current_row + r].height = 20

            img_path = ss.ruta if not ss.sin_stock else None
            if img_path and img_path.exists():
                try:
                    xl_img        = XLImage(str(img_path))
                    xl_img.width  = IMG_W_PX
                    xl_img.height = IMG_H_PX
                    ws.add_image(xl_img, f"D{img_row}")
                except Exception as exc:
                    logger.warning("No se pudo insertar imagen en Excel: %s", exc)
                    ws.cell(row=img_row, column=4, value="[Captura no disponible]").font = _font(color="999999")
            elif ss.sin_stock:
                c = ws.cell(row=img_row, column=4, value="⚠ PRODUCTO SIN STOCK EN ESTA TIENDA")
                c.font = Font(name="Arial", size=12, bold=True, color="CC0000")
            else:
                ws.cell(row=img_row, column=4, value="[Captura no disponible]").font = _font(color="999999")

            current_row += IMG_ROWS + 1

    ruta_salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(ruta_salida))
    logger.info("Excel cotizaciones web guardado: %s", ruta_salida)
