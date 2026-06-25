"""
Validación del archivo PLAN_INVERSION.xlsx — hoja "2. Revisión compra (cotización)".

Estructura de la hoja:
  - Fila 10  : nombre del ítem/producto (cols D, G, J, M... — salto de 3)
  - Fila 11  : "Cotización 1 / 2 / 3" por ítem
  - Fila 12  : Bien/Servicio/Ítem
  - Fila 13  : Nombre proveedor       ← campo obligatorio
  - Fila 17  : Descripción            ← campo obligatorio
  - Fila 18  : Valor total            ← campo obligatorio
  - Fila 32  : Cotización seleccionada (X marca la cotización elegida por producto)

Un ítem es "activo" si al menos una de sus 3 cotizaciones tiene proveedor con valor.
Los ítems de plantilla vacíos ("Ítem 4", "Ítem 5"...) sin datos se omiten.
"""

import base64
import logging
import re
from pathlib import Path
from typing import NamedTuple
from app.ia.extractor import extraer_texto

logger = logging.getLogger(__name__)

# Índices de fila (1-based, como en openpyxl)
_FILA_ITEM_NOMBRE    = 10
_FILA_BIEN           = 12
_FILA_PROVEEDOR      = 13
_FILA_DESCRIPCION    = 17
_FILA_VALOR_TOTAL    = 18
_FILA_COT_SELEC      = 32   # fila con la "X" de cotización seleccionada

_COL_INICIO          = 4    # columna D
_SALTO               = 3    # cada ítem ocupa 3 columnas (una por cotización)
_HOJA_NOMBRE         = "2. Revisión compra (cotización)"
_HOJA_FICHA_TECNICA  = "1. Ficha técnica"
_COL_FICHA_LABEL     = 4   # columna D — etiquetas de fila
_COL_FICHA_DATA      = 5   # columna E — primer producto (los demás siguen en E+)

_RE_ITEM_GENERICO    = re.compile(r"^item\s*\d+$")   # se aplica sobre texto ya normalizado
# Número colombiano: 195.000  /  1.796.826  /  2.990.000
_RE_NUM_CO           = re.compile(r"\b\d{1,3}(?:\.\d{3})+\b")

# Si el texto nativo del PDF tiene menos de esta cantidad de chars se considera escaneado
_CHARS_ESCANEADO_UMBRAL = 80


class AlertaCotizacion(NamedTuple):
    item:       str   # nombre del ítem
    cotizacion: int   # 1, 2 o 3
    campo:      str   # "proveedor", "descripcion", "valor_total" o "no_seleccionada_en_pdf"


class CotizacionSeleccionada(NamedTuple):
    item:        str
    cotizacion:  int          # 1, 2 o 3
    proveedor:   str | None
    descripcion: str | None
    valor_total: int | None   # valor numérico sin puntos


class ResultadoPlanInversion(NamedTuple):
    ok:                   bool
    alertas:              list[AlertaCotizacion]
    items_ok:             list[str]
    resumen:              str
    seleccionadas:        list[CotizacionSeleccionada]   # cotizaciones elegidas por producto
    todas_cotizaciones:   list[CotizacionSeleccionada] = []   # TODAS las cotizaciones (1, 2 y 3)


class ResultadoFirmas(NamedTuple):
    ok:             bool
    firmas_halladas: int
    resumen:        str


class ResultadoCotizacionPDF(NamedTuple):
    ok:       bool
    alertas:  list[str]   # una por producto cuyo valor total no se encontró en el PDF
    resumen:  str


# ── Helpers ──────────────────────────────────────────────────────────────────

def _valor(ws, fila: int, col: int):
    """Devuelve el valor de una celda como str limpio, o None si está vacía."""
    v = ws.cell(row=fila, column=col).value
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _valor_num(ws, fila: int, col: int) -> int | None:
    """Devuelve el valor numérico de una celda (entero), o None."""
    v = ws.cell(row=fila, column=col).value
    if v is None:
        return None
    try:
        return int(float(str(v).strip()))
    except (ValueError, TypeError):
        return None


def _es_item_generico(nombre: str | None) -> bool:
    if not nombre:
        return True
    from app.core.normalizacion import normalizar
    return bool(_RE_ITEM_GENERICO.match(normalizar(str(nombre).strip())))


def _co_num_a_int(texto: str) -> int:
    """Convierte '1.796.826' (formato colombiano) a entero 1796826."""
    return int(texto.replace(".", ""))


def _extraer_numeros_pdf(texto: str) -> set[int]:
    """Extrae todos los números en formato colombiano del texto del PDF."""
    return {_co_num_a_int(m) for m in _RE_NUM_CO.findall(texto)}


# ── Validaciones ──────────────────────────────────────────────────────────────

def validar_plan_inversion(ruta: Path) -> ResultadoPlanInversion:
    """
    Lee el PLAN_INVERSION.xlsx y valida la hoja de cotizaciones.
    También extrae la cotización seleccionada (X) por producto.
    """
    try:
        import openpyxl
    except ImportError:
        msg = "openpyxl no instalado"
        return ResultadoPlanInversion(ok=False, alertas=[], items_ok=[], resumen=msg, seleccionadas=[])

    try:
        wb = openpyxl.load_workbook(str(ruta), data_only=True)
    except Exception as exc:
        msg = f"No se pudo abrir {ruta.name}: {exc}"
        return ResultadoPlanInversion(ok=False, alertas=[], items_ok=[], resumen=msg, seleccionadas=[])

    if _HOJA_NOMBRE not in wb.sheetnames:
        msg = f"Hoja '{_HOJA_NOMBRE}' no encontrada en {ruta.name}"
        return ResultadoPlanInversion(ok=False, alertas=[], items_ok=[], resumen=msg, seleccionadas=[])

    ws      = wb[_HOJA_NOMBRE]
    max_col = ws.max_column

    alertas:       list[AlertaCotizacion]      = []
    items_ok:      list[str]                   = []
    seleccionadas: list[CotizacionSeleccionada] = []
    todas_cotizaciones: list[CotizacionSeleccionada] = []
    items_revisados = 0

    col = _COL_INICIO
    while col + 2 <= max_col:
        nombre_item = _valor(ws, _FILA_ITEM_NOMBRE, col)
        proveedores = [_valor(ws, _FILA_PROVEEDOR, col + i) for i in range(3)]
        activo      = any(p for p in proveedores)

        if not activo:
            if _es_item_generico(nombre_item):
                break   # fin de ítems reales

        label = nombre_item or f"Ítem en col {col}"
        items_revisados += 1
        alertas_item: list[AlertaCotizacion] = []

        # ── Verificar completitud de las 3 cotizaciones + recolectarlas ───
        for i in range(3):
            num_cot = i + 1
            c = col + i
            prov_c = _valor(ws, _FILA_PROVEEDOR,   c)
            desc_c = _valor(ws, _FILA_DESCRIPCION, c)
            vt_c   = _valor_num(ws, _FILA_VALOR_TOTAL, c)
            if not prov_c:
                alertas_item.append(AlertaCotizacion(label, num_cot, "proveedor"))
            if not desc_c:
                alertas_item.append(AlertaCotizacion(label, num_cot, "descripcion"))
            if vt_c is None:
                alertas_item.append(AlertaCotizacion(label, num_cot, "valor_total"))
            # Recolectar la cotización (aunque esté incompleta) si tiene proveedor o valor
            if prov_c or vt_c is not None:
                todas_cotizaciones.append(CotizacionSeleccionada(
                    item        = label,
                    cotizacion  = num_cot,
                    proveedor   = prov_c,
                    descripcion = desc_c,
                    valor_total = vt_c,
                ))

        # ── Detectar cotización seleccionada (X en fila 32) ───────────────
        selec_idx = None   # 0, 1 o 2
        for i in range(3):
            v = _valor(ws, _FILA_COT_SELEC, col + i)
            if v and v.strip().upper() == "X":
                selec_idx = i
                break

        if selec_idx is not None:
            c_sel = col + selec_idx
            seleccionadas.append(CotizacionSeleccionada(
                item        = label,
                cotizacion  = selec_idx + 1,
                proveedor   = _valor(ws, _FILA_PROVEEDOR,   c_sel),
                descripcion = _valor(ws, _FILA_DESCRIPCION, c_sel),
                valor_total = _valor_num(ws, _FILA_VALOR_TOTAL, c_sel),
            ))
        else:
            alertas_item.append(AlertaCotizacion(label, 0, "sin_cotizacion_seleccionada"))
            logger.warning("  PLAN_INVERSION '%s': no tiene X en cotización seleccionada", label)

        if alertas_item:
            alertas.extend(alertas_item)
            logger.info("  PLAN_INVERSION '%s': %d alerta(s)", label, len(alertas_item))
        else:
            items_ok.append(label)

        col += _SALTO

    if items_revisados == 0:
        resumen = "PLAN_INVERSION: no se encontraron ítems con datos"
        return ResultadoPlanInversion(ok=False, alertas=[], items_ok=[], resumen=resumen, seleccionadas=[])

    if not alertas:
        resumen = f"OK ({items_revisados} ítem(s), todas las cotizaciones completas)"
    else:
        items_con_alerta: dict[str, list[str]] = {}
        for a in alertas:
            items_con_alerta.setdefault(a.item, []).append(
                f"Cot.{a.cotizacion}/{a.campo}" if a.cotizacion else a.campo
            )
        partes = [f"{item}: {', '.join(errs)}" for item, errs in items_con_alerta.items()]
        resumen = "FALTA — " + " | ".join(partes)

    return ResultadoPlanInversion(
        ok                = len(alertas) == 0,
        alertas           = alertas,
        items_ok          = items_ok,
        resumen           = resumen,
        seleccionadas     = seleccionadas,
        todas_cotizaciones= todas_cotizaciones,
    )


_FIRMAS_REQUERIDAS = 2


def verificar_firmas_pdf(ruta_pdf: Path) -> ResultadoFirmas:
    """
    Cuenta las imágenes embebidas en el PDF como indicador de firmas manuscritas.

    El documento FIRMA_UP_*.pdf tiene exactamente 2 imágenes de firma
    (representante legal + consultor). No usa IA ni widgets digitales.

    Retorna ok=True solo si se encuentran exactamente _FIRMAS_REQUERIDAS imágenes.
    """
    try:
        import fitz
        doc = fitz.open(str(ruta_pdf))
    except Exception as exc:
        return ResultadoFirmas(ok=False, firmas_halladas=0, resumen=f"Error al leer PDF: {exc}")

    total_imgs = sum(len(page.get_image_info()) for page in doc)
    doc.close()

    logger.debug("  Firmas PDF '%s': %d imagen(s) encontradas", ruta_pdf.name, total_imgs)

    if total_imgs >= _FIRMAS_REQUERIDAS:
        return ResultadoFirmas(
            ok=True,
            firmas_halladas=total_imgs,
            resumen=f"OK — {total_imgs} firma(s) encontradas",
        )

    faltantes = _FIRMAS_REQUERIDAS - total_imgs
    if total_imgs == 0:
        msg = "FALTA — no se encontraron firmas en el PDF"
    else:
        msg = f"FALTA — solo {total_imgs} de {_FIRMAS_REQUERIDAS} firma(s) requeridas"

    return ResultadoFirmas(ok=False, firmas_halladas=total_imgs, resumen=msg)


def _texto_nativo_pdf(ruta_pdf: Path) -> str:
    """Extrae texto nativo del PDF con PyMuPDF (sin OCR)."""
    try:
        import fitz
        doc = fitz.open(str(ruta_pdf))
        texto = "\n".join(p.get_text() for p in doc)
        doc.close()
        return texto
    except Exception as exc:
        logger.warning("  PDF texto nativo '%s': %s", ruta_pdf.name, exc)
        return ""


from app.ia.extractor import extraer_texto

def _verificar_valores_con_gpt(
    ruta_pdf: Path,
    pendientes: list["CotizacionSeleccionada"],
) -> tuple[list["CotizacionSeleccionada"], list["CotizacionSeleccionada"]]:
    """
    Envía el PDF a GPT-4o-mini y pregunta específicamente si cada valor
    de las cotizaciones pendientes aparece en el documento.

    Retorna (confirmados, no_confirmados).
    """
    try:
        from openai import OpenAI
        from app.config import OPENAI_API_KEY
    except Exception:
        logger.warning("  GPT fallback: openai/config no disponible")
        return [], pendientes

    if not OPENAI_API_KEY:
        logger.warning("  GPT fallback: OPENAI_API_KEY no configurada")
        return [], pendientes

    try:
        pdf_b64 = base64.b64encode(ruta_pdf.read_bytes()).decode()
    except Exception as exc:
        logger.error("  GPT fallback: no se pudo leer '%s': %s", ruta_pdf.name, exc)
        return [], pendientes

    # Construir la lista de valores a buscar
    items_json = [
        {
            "item":              sel.item,
            "cotizacion":        sel.cotizacion,
            "proveedor":         sel.proveedor or "",
            "valor_total":       sel.valor_total,
            "valor_total_fmt":   f"{sel.valor_total:,}".replace(",", ".") if sel.valor_total else "N/A",
        }
        for sel in pendientes
        if sel.valor_total is not None
    ]

    if not items_json:
        return [], pendientes

    import json as _json
    prompt = (
        "Este es un documento de cotización con múltiples ítems.\n\n"
        "TAREA: Para cada ítem de la lista, busca en el documento el VALOR TOTAL "
        "del ítem (no el precio unitario, no el precio por unidad). "
        "El valor total es el monto final que se paga por la cantidad completa del producto "
        "y suele aparecer en la columna 'Total', 'Valor total' o 'Subtotal' de la cotización.\n\n"
        "IMPORTANTE:\n"
        "- Busca ÚNICAMENTE el valor total, NO el precio unitario.\n"
        "- El valor puede aparecer con o sin separadores de miles "
        "(formato colombiano '1.234.000' o sin puntos '1234000').\n"
        "- Si en el documento aparece tanto un precio unitario como un total diferente, "
        "verifica SOLO contra el total.\n\n"
        f"Ítems a verificar:\n{_json.dumps(items_json, ensure_ascii=False, indent=2)}\n\n"
        "Responde ÚNICAMENTE con este JSON (sin texto adicional):\n"
        '{"resultados": [{"item": "...", "cotizacion": 1, "encontrado": true/false}]}'
    )

    try:
        client   = OpenAI(api_key=OPENAI_API_KEY)
        response = client.responses.create(
            model="gpt-4o-mini",
            input=[{
                "role": "user",
                "content": [
                    {
                        "type": "input_file",
                        "filename": ruta_pdf.name,
                        "file_data": f"data:application/pdf;base64,{pdf_b64}",
                    },
                    {"type": "input_text", "text": prompt},
                ],
            }],
        )
        texto = (response.output_text or "").strip()
        logger.info("  GPT fallback cotización '%s': respuesta recibida", ruta_pdf.name)
    except Exception as exc:
        logger.error("  GPT fallback cotización error: %s", exc)
        return [], pendientes

    # Parsear respuesta
    try:
        import re as _re, json as _json2
        m = _re.search(r"\{.*\}", texto, _re.DOTALL)
        data = _json2.loads(m.group(0)) if m else {}
    except Exception:
        data = {}

    encontrados_set: set[tuple[str, int]] = set()
    for r in (data.get("resultados") or []):
        if r.get("encontrado"):
            encontrados_set.add((r.get("item", ""), int(r.get("cotizacion", 0))))

    confirmados:     list[CotizacionSeleccionada] = []
    no_confirmados:  list[CotizacionSeleccionada] = []

    for sel in pendientes:
        clave = (sel.item, sel.cotizacion)
        if clave in encontrados_set:
            confirmados.append(sel)
            logger.info("  GPT: ✓ valor confirmado para '%s' Cot.%d", sel.item, sel.cotizacion)
        else:
            no_confirmados.append(sel)
            logger.warning("  GPT: ✗ valor NO confirmado para '%s' Cot.%d", sel.item, sel.cotizacion)

    return confirmados, no_confirmados


def validar_cotizacion_seleccionada_en_pdf(
    seleccionadas: list[CotizacionSeleccionada],
    ruta_pdf: Path,
) -> ResultadoCotizacionPDF:
    """
    Verifica que el valor total de cada cotización seleccionada aparezca
    en el PDF.

    Flujo:

      1. Detecta si el PDF es escaneado (solo para decidir el fallback GPT).
      2. Extrae el texto mediante el pipeline OCR centralizado.
      3. Busca los valores en el texto extraído.
      4. Si el PDF es escaneado y algún valor no aparece, consulta GPT-4o-mini.
    """

    # ------------------------------------------------------------------
    # Detectar si el PDF es escaneado (solo para el fallback GPT)
    # ------------------------------------------------------------------
    texto_nativo = _texto_nativo_pdf(ruta_pdf)
    es_escaneado = len(texto_nativo.strip()) < _CHARS_ESCANEADO_UMBRAL

    # ------------------------------------------------------------------
    # Extraer texto usando el pipeline centralizado
    # ------------------------------------------------------------------
    texto_activo = extraer_texto(ruta_pdf)

    origen = "OCR" if es_escaneado else "texto nativo"

    logger.info(
        "  Texto utilizado (%s): %d caracteres",
        origen,
        len(texto_activo),
    )

    numeros_pdf = _extraer_numeros_pdf(texto_activo)

    logger.debug(
        "  Números extraídos (%s): %s",
        origen,
        sorted(numeros_pdf),
    )

    alertas_pdf: list[str] = []
    ok_items: list[str] = []
    pendientes_gpt: list[CotizacionSeleccionada] = []

    # ------------------------------------------------------------------
    # Validación de valores
    # ------------------------------------------------------------------
    for sel in seleccionadas:

        if sel.valor_total is None:
            alertas_pdf.append(
                f"{sel.item} (Cot.{sel.cotizacion}): valor total no disponible en Excel"
            )
            continue

        if sel.valor_total in numeros_pdf:

            ok_items.append(
                f"{sel.item}: ${sel.valor_total:,}".replace(",", ".")
            )

            logger.info(
                "  PDF [%s]: ✓ valor %d de '%s' (Cot.%d) encontrado",
                origen,
                sel.valor_total,
                sel.item,
                sel.cotizacion,
            )

        else:

            logger.warning(
                "  PDF [%s]: ✗ valor %d de '%s' (Cot.%d) NO encontrado",
                origen,
                sel.valor_total,
                sel.item,
                sel.cotizacion,
            )

            pendientes_gpt.append(sel)

    # ------------------------------------------------------------------
    # Segunda opinión con GPT solo para PDFs escaneados
    # ------------------------------------------------------------------
    if pendientes_gpt and es_escaneado:

        logger.info(
            "  PDF escaneado: %d valor(es) sin confirmar — enviando a GPT-4o-mini...",
            len(pendientes_gpt),
        )

        confirmados_gpt, no_confirmados_gpt = _verificar_valores_con_gpt(
            ruta_pdf,
            pendientes_gpt,
        )

        for sel in confirmados_gpt:
            ok_items.append(
                f"{sel.item}: ${sel.valor_total:,}".replace(",", ".")
            )

        for sel in no_confirmados_gpt:
            alertas_pdf.append(
                (
                    f"⛔ {sel.item} "
                    f"(Cot.{sel.cotizacion} — {sel.proveedor or 'sin proveedor'}): "
                    f"valor ${sel.valor_total:,} NO confirmado por OCR ni por IA — "
                    f"los valores del PDF escaneado podrían ser incorrectos o ilegibles"
                ).replace(",", ".")
            )

    elif pendientes_gpt:

        # PDF digital: el valor simplemente no aparece
        for sel in pendientes_gpt:

            alertas_pdf.append(
                (
                    f"{sel.item} "
                    f"(Cot.{sel.cotizacion} — {sel.proveedor or 'sin proveedor'}): "
                    f"valor ${sel.valor_total:,} no encontrado en PDF"
                ).replace(",", ".")
            )

    # ------------------------------------------------------------------
    # Resumen
    # ------------------------------------------------------------------
    if not alertas_pdf:
        resumen = f"OK — {len(ok_items)} producto(s) verificados en PDF"
    else:
        resumen = "FALTA — " + " | ".join(alertas_pdf)

    return ResultadoCotizacionPDF(
        ok=len(alertas_pdf) == 0,
        alertas=alertas_pdf,
        resumen=resumen,
    )

# ── Verificación de cotizaciones en carpeta 02_COTIZACIONES_Y_COMPRA ──────────

class ResultadoCotizacionesCarpeta(NamedTuple):
    ok:       bool
    detiene:  bool          # True si falta alguna cotización → se debe detener el proceso
    alertas:  list[str]
    resumen:  str


# Palabras genéricas de razón social que NO sirven para identificar la empresa
_STOPWORDS_EMPRESA = {
    "sas", "sa", "s", "a", "ltda", "cia", "compania", "company", "co",
    "ingenieria", "comercializadora", "comercial", "distribuciones", "distribuidora",
    "ci", "eu", "e", "u", "grupo", "group", "soluciones", "inversiones", "servicios",
    "y", "de", "del", "la", "el", "los", "las", "para", "industrias", "industrial",
}


def _tokens_empresa(nombre: str) -> list[str]:
    """Tokens distintivos del nombre del proveedor (sin palabras genéricas)."""
    from app.core.normalizacion import normalizar
    return [
        t for t in re.findall(r'[a-z0-9]+', normalizar(nombre or ""))
        if len(t) >= 3 and t not in _STOPWORDS_EMPRESA
    ]


def validar_cotizaciones_en_carpeta(
    cotizaciones: list[CotizacionSeleccionada],
    archivos_02: list[Path],
) -> ResultadoCotizacionesCarpeta:
    """
    Verifica las cotizaciones del plan contra los PDFs de 02_COTIZACIONES_Y_COMPRA.

    1. Localiza el PDF de cada proveedor por coincidencia PARCIAL del nombre de la
       empresa (tokens distintivos) en el nombre del archivo o en su texto.
       El texto se extrae nativamente (PDFs con caracteres) — sin OCR ni IA.
    2. Si falta el PDF de algún proveedor → detiene el proceso.
    3. Para cada cotización valida que su valor_total aparezca en el PDF del proveedor.
    """
    from app.core.normalizacion import normalizar

    pdfs = [a for a in archivos_02 if str(a).lower().endswith(".pdf")]
    if not pdfs:
        return ResultadoCotizacionesCarpeta(
            ok=False, detiene=True,
            alertas=["⛔ No hay PDFs de cotización en 02_COTIZACIONES_Y_COMPRA"],
            resumen="FALTA — sin PDFs de cotización",
        )

    # Palabras para descartar documentos administrativos
    EXCLUIR = {
        "rut",
        "camara",
        "cámara",
        "certificado",
        "certificacion",
        "certificación",
        "existencia",
        "representacion",
        "representación",
        "bancaria",
    }

    # Cache de texto nativo por PDF
    _texto_cache: dict[Path, str] = {}
    def _texto(p: Path) -> str:
        if p not in _texto_cache:
            _texto_cache[p] = _texto_nativo_pdf(p)
        return _texto_cache[p]

    # Filtrar PDFs administrativos (RUT, Cámara, Certificados, etc.)
    pdfs_cotizacion = []

    for p in pdfs:
        nombre = normalizar(p.name)

        if any(x in nombre for x in EXCLUIR):
            logger.info("  02_COTIZACIONES: descartado %s", p.name)
            continue

        pdfs_cotizacion.append(p)

    logger.info(
        "  02_COTIZACIONES: PDFs candidatos a cotización: %s",
        [p.name for p in pdfs_cotizacion]
    )

    if not pdfs_cotizacion:
        return ResultadoCotizacionesCarpeta(
            ok=False,
            detiene=True,
            alertas=[
                "⛔ No se encontraron PDFs de cotización válidos "
                "(solo RUT, Cámara, Certificados, etc.)"
            ],
            resumen="FALTA — sin PDFs de cotización válidos",
        )

    # Búsqueda del PDF de cada proveedor (por tokens en nombre de archivo + contenido)
    proveedores = sorted({c.proveedor for c in cotizaciones if c.proveedor})
    pdf_por_proveedor: dict[str, Path] = {}
    alertas: list[str] = []
    detiene = False

    for prov in proveedores:
        toks = _tokens_empresa(prov)
        mejor, mejor_score = None, 0

        for p in pdfs_cotizacion:
            heno = normalizar(p.name) + " " + normalizar(_texto(p))
            score = sum(1 for t in toks if t in heno)

            logger.info(
                "PDF=%s | score=%s",
                p.name,
                score
            )

            if score > mejor_score:
                mejor, mejor_score = p, score

        if mejor and mejor_score > 0:
            pdf_por_proveedor[prov] = mejor
            logger.info(
                "  02_COTIZACIONES: proveedor '%s' → %s",
                prov,
                mejor.name
            )
        else:
            alertas.append(
                f"⛔ Cotización de '{prov}' NO encontrada en 02_COTIZACIONES_Y_COMPRA"
            )
            detiene = True
            logger.warning(
                "  02_COTIZACIONES: proveedor '%s' SIN PDF",
                prov
            )

    # Validar el valor_total de cada cotización contra el PDF de su proveedor
    ok_items: list[str] = []
    for c in cotizaciones:
        if not c.proveedor:
            continue
        pdf = pdf_por_proveedor.get(c.proveedor)
        if pdf is None:
            continue   # ya alertado como faltante
        if c.valor_total is None:
            alertas.append(
                f"{c.item} (Cot.{c.cotizacion} — {c.proveedor}): sin valor en Excel"
            )
            continue
        numeros = _extraer_numeros_pdf(_texto(pdf))
        if c.valor_total in numeros:
            ok_items.append(
                f"{c.item}: ${c.valor_total:,}".replace(",", ".")
            )

            logger.info(
                "  02_COTIZACIONES: ✓ %s (Cot.%d, %s) valor %d coincide en %s",
                c.item, c.cotizacion, c.proveedor, c.valor_total, pdf.name,
            )
        else:
            alertas.append(
                f"{c.item} (Cot.{c.cotizacion} — {c.proveedor}): "
                f"valor ${c.valor_total:,} no coincide con el PDF {pdf.name}".replace(",", ".")
            )
            logger.warning(
                "  02_COTIZACIONES: ✗ %s (Cot.%d, %s) valor %d NO está en %s",
                c.item, c.cotizacion, c.proveedor, c.valor_total, pdf.name,
            )

    # Resumen CORTO para la celda del checklist (el detalle va en observaciones)
    n_cotiz = sum(1 for c in cotizaciones if c.proveedor)
    n_faltantes = len(proveedores) - len(pdf_por_proveedor)
    n_problema = len(alertas) - n_faltantes

    if not alertas:
        resumen = f"OK — {len(ok_items)}/{n_cotiz} precios coinciden"
    else:
        partes = []
        if n_faltantes:
            partes.append(f"{n_faltantes} cotización(es) no hallada(s)")
        if n_problema > 0:
            partes.append(f"{n_problema} precio(s) no coinciden")
        resumen = "REVISAR — " + ", ".join(partes)

    return ResultadoCotizacionesCarpeta(
        ok=len(alertas) == 0,
        detiene=detiene,
        alertas=alertas,
        resumen=resumen,
    )


# ── Lectura de Ficha Técnica ──────────────────────────────────────────────────

class FichaTecnica(NamedTuple):
    """Datos técnicos de un producto extraídos de la hoja '1. Ficha técnica'."""
    denominacion:       str | None   # "Denominación del bien" — nombre limpio del producto
    descripcion_general: str | None  # descripción funcional del producto
    especificaciones:   str | None   # especificaciones técnicas consolidadas


def leer_ficha_tecnica(ruta: Path) -> dict[str, FichaTecnica]:
    """
    Lee la hoja '1. Ficha técnica' del PLAN_INVERSION.xlsx y extrae
    para cada producto: denominación, descripción general y especificaciones.

    Estructura de la hoja:
      - Col D (4): etiquetas de fila
      - Cols E+ (5+): un producto por columna
      - Fila donde col D = "Descripciones": nombres de productos (igual que hoja 2)
      - Filas posteriores: "Denominación del bien/insumo/...", "Descripción general",
        "Especificaciones técnicas:", sub-especificaciones (" - Potencia", etc.)

    Retorna {nombre_producto: FichaTecnica}.
    Retorna {} si la hoja no existe o no se puede leer.
    """
    try:
        import openpyxl
        from app.core.normalizacion import normalizar as _norm
    except ImportError:
        return {}

    try:
        wb = openpyxl.load_workbook(str(ruta), data_only=True)
    except Exception as exc:
        logger.warning("  FichaTecnica: no se pudo abrir '%s': %s", ruta.name, exc)
        return {}

    if _HOJA_FICHA_TECNICA not in wb.sheetnames:
        logger.info("  FichaTecnica: hoja '%s' no encontrada en %s", _HOJA_FICHA_TECNICA, ruta.name)
        return {}

    ws = wb[_HOJA_FICHA_TECNICA]

    # Palabras que indican nombre genérico de ítem vacío → se omiten
    _RE_ITEM_VACIO = re.compile(r'^[íi]tem\s*\d+$', re.IGNORECASE)

    # Etiquetas normalizadas de "Denominación" (distintas según el tipo de capital)
    _ETIQ_DENOMINACION = {
        "denominacion del bien",
        "denominacion del insumo material o inventario",
        "denominacion de la certificacion permiso o licencia",
        "denominacion de la certificacion o licencia requerida",
        "denominacion",
    }
    # "Descripción general" aparece igual en todas las secciones
    _ETIQ_DESC_GENERAL = {"descripcion general"}
    # "Especificaciones técnicas:" y sus variantes
    _ETIQ_SPECS_MAIN   = {"especificaciones tecnicas", "especificaciones tecnicas:"}
    # Sub-especificaciones técnicas a incluir en el bloque de specs
    _ETIQ_SPECS_SUB    = {
        "potencia (especifique si hay alguna restriccion)",
        "capacidad (especifique si hay alguna restriccion)",
        "dimensiones (especifique si hay alguna restriccion)",
        "peso (especifique si hay alguna restriccion)",
        "adicione otras especificaciones tecnicas a tener en cuenta",
        "adicione las especificaciones a tener en cuenta",
        "material",
        "composicion",
        "referencia (especifique si hay alguna restriccion)",
    }

    # ── Paso 1: encontrar la fila de encabezado con los nombres de productos ──
    header_row_idx: int | None = None
    product_cols: dict[int, str] = {}  # {col_1based: nombre_producto}

    for row in ws.iter_rows(min_row=1, max_row=20):
        label_val = row[_COL_FICHA_LABEL - 1].value if len(row) >= _COL_FICHA_LABEL else None
        if label_val and "descripciones" in _norm(str(label_val)):
            header_row_idx = row[0].row
            for cell in row[_COL_FICHA_DATA - 1:]:
                val = cell.value
                if val is None:
                    continue
                nombre = str(val).strip()
                if not nombre or _RE_ITEM_VACIO.match(nombre) or nombre == ".":
                    continue
                product_cols[cell.column] = nombre
            break

    if not product_cols:
        logger.warning(
            "  FichaTecnica: no se encontró fila 'Descripciones' en '%s'", ruta.name
        )
        return {}

    logger.info(
        "  FichaTecnica: %d producto(s) en fila %s de '%s': %s",
        len(product_cols), header_row_idx, ruta.name,
        list(product_cols.values()),
    )

    # ── Paso 2: recolectar datos por etiqueta para cada columna de producto ──
    # Usamos dicts para acumular: tomamos el PRIMER valor no nulo por (col, campo)
    data: dict[int, dict[str, list[str]]] = {
        col: {"denominacion": [], "descripcion_general": [], "especificaciones": []}
        for col in product_cols
    }

    for row in ws.iter_rows(min_row=(header_row_idx or 1) + 1):
        label_cell = row[_COL_FICHA_LABEL - 1] if len(row) >= _COL_FICHA_LABEL else None
        if label_cell is None or label_cell.value is None:
            continue

        label_raw  = str(label_cell.value).strip()
        label_norm = _norm(label_raw.lstrip("- \t"))

        # Clasificar la etiqueta
        if label_norm in _ETIQ_DENOMINACION:
            campo = "denominacion"
        elif label_norm in _ETIQ_DESC_GENERAL:
            campo = "descripcion_general"
        elif label_norm in _ETIQ_SPECS_MAIN:
            campo = "especificaciones"
        elif label_norm in _ETIQ_SPECS_SUB:
            campo = "especificaciones"   # sub-specs se concatenan al bloque de specs
        else:
            continue

        for col, _ in product_cols.items():
            col_0 = col - 1
            if col_0 >= len(row):
                continue
            cell_val = row[col_0].value
            if cell_val is None:
                continue
            texto = str(cell_val).strip()
            if not texto:
                continue

            # Para denominación y descripción: solo guardar el primer valor
            if campo in ("denominacion", "descripcion_general"):
                if not data[col][campo]:
                    data[col][campo].append(texto)
            else:
                # Para specs: acumular pero con máximo razonable
                if len(data[col]["especificaciones"]) < 10:
                    data[col]["especificaciones"].append(texto)

    # ── Paso 3: ensamblar FichaTecnica por nombre de producto ──────────────
    resultado: dict[str, FichaTecnica] = {}
    for col, nombre in product_cols.items():
        d = data[col]
        denominacion       = d["denominacion"][0] if d["denominacion"] else None
        descripcion_general = d["descripcion_general"][0] if d["descripcion_general"] else None
        especificaciones   = " | ".join(d["especificaciones"]) if d["especificaciones"] else None

        ficha = FichaTecnica(
            denominacion        = denominacion,
            descripcion_general = descripcion_general,
            especificaciones    = especificaciones,
        )
        resultado[nombre] = ficha
        logger.debug(
            "  FichaTecnica '%s': denom=%s | desc=%s chars | specs=%s chars",
            nombre[:40],
            bool(denominacion), len(descripcion_general or ""), len(especificaciones or ""),
        )

    return resultado