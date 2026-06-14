"""
Analizadores de documentos de la carpeta
02_VISITA_2_DIAGNOSTICO.

- Plan de Negocio
- Diagnóstico
"""

import logging
from pathlib import Path
from typing import Dict, List
from app.core.normalizacion import normalizar

import openpyxl

from app.config import (
    IA_HABILITADO,
    IA_MAX_PAGINAS_VISITA,
    OPENAI_API_KEY,
)

from app.ia.analizador_visita import (
    ResultadoAnalisis,
    _SISTEMA_BASE,
    _llamar_gpt,
    _obtener_imagenes_b64,
    _parsear_json,
)

logger = logging.getLogger(__name__)


# ──────────────────────────────────────────────────────────────────────────────
# PLAN DE NEGOCIO
# ──────────────────────────────────────────────────────────────────────────────

CAMPOS_CONTROL = [
    "unidad productiva",
    "representante",
    "celular",
    "actividad economica principal",
    "necesidad o problema",
    "principales clientes",
    "otro tipo de clientes",
    "fortalezas",
    "oportunidades",
    "debilidades",
    "amenazas",
    "plan de inversion",
    "estrategias administrativas",
    "fortalecimiento productivo",
    "asistencia tecnica",
    "modalidad educativa",
    "organizacion y/o asociacion",
    "asociacion, red o plataforma",
]


def buscar_valor_cercano(ws, etiqueta):
    """
    Busca una etiqueta y localiza una respuesta cercana.

    Revisa un área alrededor de la etiqueta y evita
    tomar otras etiquetas como respuesta.
    """

    etiqueta_norm = normalizar(etiqueta)

    for fila in ws.iter_rows():

        for celda in fila:

            if celda.value is None:
                continue

            texto = normalizar(str(celda.value))

            if etiqueta_norm not in texto:
                continue

            fila_idx = celda.row
            col_idx = celda.column

            candidatos = []

            # Buscar en una ventana cercana
            for r in range(fila_idx, fila_idx + 6):

                for c in range(col_idx + 1, col_idx + 6):

                    valor = ws.cell(
                        row=r,
                        column=c,
                    ).value

                    if valor is None:
                        continue

                    if isinstance(valor, str):

                        valor = valor.strip()

                        if not valor:
                            continue

                    candidatos.append(str(valor))

            # Filtrar etiquetas conocidas
            for candidato in candidatos:

                candidato_norm = normalizar(candidato)

                es_etiqueta = any(
                    normalizar(campo) in candidato_norm
                    for campo in CAMPOS_CONTROL
                )

                if es_etiqueta:
                    continue

                return candidato

            return None

    return None

def analizar_plan_negocio(ruta_excel: str) -> Dict[str, object]:
    """
    Valida que los campos mínimos del plan de negocio estén diligenciados.
    """

    faltantes: List[str] = []

    try:
        wb = openpyxl.load_workbook(
            ruta_excel,
            data_only=True,
        )

        ws = wb.worksheets[0]

        for etiqueta in CAMPOS_CONTROL:

            valor = buscar_valor_cercano(
                ws,
                etiqueta,
            )

            if valor is None:

                print(
                    f"NO ENCONTRADO -> {etiqueta}"
                )

                faltantes.append(
                    etiqueta
                )

            else:

                print(
                    f"{etiqueta} -> {valor}"
                )

        wb.close()

        return {
            "completo": len(faltantes) == 0,
            "faltantes": faltantes,
        }

    except Exception as exc:
        return {
            "completo": False,
            "faltantes": [f"Error al leer archivo: {exc}"],
        }


# ──────────────────────────────────────────────────────────────────────────────
# DIAGNÓSTICO
# ──────────────────────────────────────────────────────────────────────────────

def analizar_diagnostico(
    ruta_pdf: str | Path,
) -> ResultadoAnalisis:
    """
    Verifica que todas las secciones principales del
    Diagnóstico DPS estén diligenciadas.
    """

    if not IA_HABILITADO or not OPENAI_API_KEY:
        return ResultadoAnalisis(ok=True, alerta="")

    ruta = Path(ruta_pdf)

    imagenes = _obtener_imagenes_b64(
        ruta,
        IA_MAX_PAGINAS_VISITA,
    )

    if not imagenes:
        return ResultadoAnalisis(
            ok=False,
            alerta="No se pudo leer el archivo para análisis IA",
        )

    prompt_usuario = """Analiza este formato DIAGNÓSTICO DE UNIDADES PRODUCTIVAS DPS escaneado.

Verifica los siguientes puntos:

1. ¿Todos los campos OBLIGATORIOS del formulario tienen contenido?
    - Un campo obligatorio está vacío SOLO si está completamente en blanco.

2. Para preguntas con opciones SI / NO:
   - Debe existir al menos una opción marcada.

3. Para preguntas que incluyan observaciones, descripción,
   comentarios o justificaciones:
   - Debe existir contenido escrito.

4. Para espacios de firma:
   - Cualquier rúbrica, garabato o marca visible cuenta como firma.

5. Ignora campos que indiquen:
   - si aplica
   - cuando aplique
   - opcional
   - no aplica
   - o cualquier variante equivalente

Responde SOLO con este JSON:

{
  "campos_completos": true/false,
  "campos_vacios": [
    "lista de campos obligatorios completamente vacíos"
  ],
  "alerta": "descripción breve del problema o null"
}
"""

    try:

        data = _parsear_json(
            _llamar_gpt(
                imagenes,
                _SISTEMA_BASE,
                prompt_usuario,
            )
        )

    except Exception as exc:

        logger.error(
            "Error IA diagnóstico '%s': %s",
            ruta.name,
            exc,
        )

        return ResultadoAnalisis(
            ok=False,
            alerta=f"Error en análisis IA: {exc}",
        )

    if not data.get("campos_completos", True):

        campos_vacios = data.get(
            "campos_vacios",
            [],
        )

        if campos_vacios:

            return ResultadoAnalisis(
                ok=False,
                alerta=(
                    "Campos vacíos: "
                    + ", ".join(campos_vacios[:5])
                ),
            )

        return ResultadoAnalisis(
            ok=False,
            alerta=data.get("alerta")
            or "Diagnóstico incompleto",
        )

    return ResultadoAnalisis(ok=True)


if __name__ == "__main__":

    ruta = r"C:\Users\Usuario\Downloads\PLAN DE NEGOCIO_20736200.xlsx"

    resultado = analizar_plan_negocio(ruta)

    print("\n=== RESULTADO ===")
    print(resultado)